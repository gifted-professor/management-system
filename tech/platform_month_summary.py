#!/usr/bin/env python3
from __future__ import annotations
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent
SRC = ROOT / '账单汇总_全部.xlsx'
SHEET = '汇总(全部)'
OUT_SHEET = '平台月汇总'

def norm_plat(x: object) -> str:
    s = str(x or '').strip().replace(' ', '')
    s = s.replace('一', '1').replace('二', '2')
    s = s.replace('咸鱼', '闲鱼')
    if s in {'闲鱼', '闲鱼1', '闲鱼一'}:
        return '闲鱼1'
    if s in {'闲鱼2', '闲鱼二', '咸鱼二', '咸鱼2'}:
        return '闲鱼2'
    return s

def is_cancel_row(row: pd.Series) -> bool:
    v1 = str(row.get('状态') or '')
    v2 = str(row.get('退款类型') or '')
    return ('取消' in v1) or ('取消' in v2)

def build_summary() -> Dict[Tuple[int, str], int]:
    try:
        df = pd.read_excel(SRC, sheet_name=SHEET, engine='openpyxl')
    except Exception:
        df = pd.read_excel(SRC, engine='openpyxl')
    for col in ['顾客付款日期', '出售平台', '状态', '退款类型', '收款额']:
        if col not in df.columns:
            df[col] = None
    df['pay_date'] = pd.to_datetime(df['顾客付款日期'], errors='coerce')
    valid = (pd.to_numeric(df['收款额'], errors='coerce').fillna(0) > 0) & (~df.apply(is_cancel_row, axis=1))
    df = df[valid & df['pay_date'].notna()].copy()
    df['plat'] = df['出售平台'].apply(norm_plat)
    df['month'] = df['pay_date'].dt.month
    subset = df[df['plat'].astype(str).str.contains('闲鱼')]
    grouped = subset.groupby(['month', 'plat']).size().reset_index(name='count')
    res: Dict[Tuple[int, str], int] = {}
    for _, r in grouped.iterrows():
        res[(int(r['month']), str(r['plat']))] = int(r['count'])
    return res

def write_sheet(res: Dict[Tuple[int, str], int]) -> None:
    wb = load_workbook(SRC)
    if OUT_SHEET in wb.sheetnames:
        ws = wb[OUT_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(OUT_SHEET)
    ws.append(['月份', '平台', '订单数'])
    for mon in sorted({m for (m, _) in res.keys()}):
        for plat in ['闲鱼1', '闲鱼2']:
            ws.append([mon, plat, res.get((mon, plat), 0)])
    wb.save(SRC)

def main() -> int:
    res = build_summary()
    write_sheet(res)
    print('已写入平台月汇总工作表:', OUT_SHEET)
    for k in sorted(res.keys()):
        print(k, res[k])
    return 0

if __name__ == '__main__':
    raise SystemExit(main())
