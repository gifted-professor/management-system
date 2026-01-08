#!/usr/bin/env python3
# relocated to tech/
"""
订单/退货单号 查询小工具

用法示例：
  python3 lookup_order.py --query  SF123456789012
  python3 lookup_order.py --query  123456789012 --export 订单查询结果.xlsx

默认数据源：tech/账单汇总_截至10月前.xlsx 的 "汇总(截至10月前)" 工作表。
可通过 --source / --sheet 指定其他表。
支持：订单号、退货单号、退货物流（含快递+单号混写）、备注里出现的号码。
查询会同时匹配原文与纯数字形式（忽略空格/横杠）。
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook, Workbook
try:
    from .common import resolve_sheet, normalize, digits_only
except Exception:
    import sys, os
    sys.path.append(os.path.dirname(__file__))
    from common import resolve_sheet, normalize, digits_only


DEFAULT_SOURCE = Path(__file__).resolve().parent / '账单汇总_截至10月前.xlsx'
DEFAULT_SHEET = '汇总(截至10月前)'


TARGET_COLS = [
    '退货单号', '退货物流', '退货快递', '退货快递单号', '退货物流单号', '退回单号', '退货运单号',
    '订单号', '单号', '出库单号', '备注'
]

EXTRA_OUTPUT_COLS = [
    '退款金额', '退款类型', '退货状态', '出售平台', '顾客付款日期', '手机号', '姓名', '商品名称', '货品名',
]


 


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description='按 订单号/退货单号 检索汇总账单')
    p.add_argument('--query', '-q', required=True, help='要查询的单号（支持模糊匹配）')
    p.add_argument('--source', default=str(DEFAULT_SOURCE), help='数据源 Excel 路径')
    p.add_argument('--sheet', default=DEFAULT_SHEET, help='工作表名')
    p.add_argument('--export', default=None, help='可选：导出命中的结果到此 Excel 文件')
    return p.parse_args()


 


def scan(path: Path, sheet: str, query: str) -> Tuple[List[str], List[List]]:
    wb, ws = resolve_sheet(path, sheet)
    try:
        iterator = ws.iter_rows(values_only=True)
        try:
            first_row = next(iterator)
        except StopIteration:
            return [], []
        header = [normalize(x) if x is not None else '' for x in first_row]
        name_to_idx: Dict[str, int] = {h: i for i, h in enumerate(header) if h}
        target_indices: List[int] = [name_to_idx[c] for c in TARGET_COLS if c in name_to_idx]
        if not target_indices:
            return header, []
        q_raw = query.strip()
        q_low = q_raw.lower()
        q_digits = digits_only(q_raw)
        results: List[List] = []
        for r in iterator:
            if r is None:
                continue
            hit = False
            for idx in target_indices:
                if idx >= len(r):
                    continue
                cell = normalize(r[idx])
                if not cell:
                    continue
                low = cell.lower()
                if q_low and q_low in low:
                    hit = True
                    break
                if q_digits:
                    if q_digits in digits_only(cell):
                        hit = True
                        break
            if hit:
                row_out: List = []
                for col in TARGET_COLS + EXTRA_OUTPUT_COLS:
                    if col in name_to_idx:
                        idx = name_to_idx[col]
                        row_out.append(r[idx] if idx < len(r) else None)
                    else:
                        row_out.append(None)
                results.append(row_out)
        return TARGET_COLS + EXTRA_OUTPUT_COLS, results
    finally:
        wb.close()


def export_results(headers: List[str], rows: List[List], out_path: Path) -> None:
    wb = Workbook(write_only=True)
    ws = wb.active
    ws.title = '订单查询结果'
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(out_path)


def main() -> None:
    args = parse_args()
    src = Path(args.source)
    if not src.exists():
        print(f'❌ 找不到数据源：{src}', file=sys.stderr)
        sys.exit(1)
    headers, rows = scan(src, args.sheet, args.query)
    if not rows:
        print('未找到匹配记录。')
        sys.exit(0)
    print(f'命中 {len(rows)} 条记录：')
    # 打印前 10 条作为预览
    preview = rows[:10]
    for i, row in enumerate(preview, 1):
        # 展示核心字段的文本摘要
        fields = dict(zip(headers, row))
        brief = [
            f"订单号={fields.get('订单号') or ''}",
            f"退货单号={fields.get('退货单号') or ''}",
            f"退货物流={fields.get('退货物流') or ''}",
            f"姓名={fields.get('姓名') or ''}",
            f"平台={fields.get('出售平台') or ''}",
            f"日期={fields.get('顾客付款日期') or ''}",
        ]
        print(f"  {i}. " + ' | '.join(brief))
    if args.export:
        out_path = Path(args.export)
        export_results(headers, rows, out_path)
        print(f'已导出全部命中记录到：{out_path}')


if __name__ == '__main__':
    main()
