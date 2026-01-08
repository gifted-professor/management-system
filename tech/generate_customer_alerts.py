#!/usr/bin/env python3
# relocated to tech/
"""
Generate customer engagement and risk alerts from the historical billing workbook.

Usage:
    python3 generate_customer_alerts.py \
        --source 账单汇总_截至10月前.xlsx \
        --sheet 汇总(截至10月前) \
        --output 客户预警输出.xlsx

The script aggregates per-customer metrics, detects high-value churn, spending drops,
and refund spikes, then writes actionable sheets for follow-up.
"""
from __future__ import annotations

import argparse
import math
from collections import Counter, defaultdict
import re
from datetime import date, datetime, timedelta, timezone
import os
import calendar
import json
from pathlib import Path
import os
import requests
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from html import escape
try:
    from .common import resolve_sheet as common_resolve_sheet, to_float as common_to_float, parse_excel_date as common_parse_excel_date, deduplicate_phone as common_deduplicate_phone, build_header_index as common_build_header_index, lookup_index as common_lookup_index
except Exception:
    import sys, os
    sys.path.append(os.path.dirname(__file__))
    from common import resolve_sheet as common_resolve_sheet, to_float as common_to_float, parse_excel_date as common_parse_excel_date, deduplicate_phone as common_deduplicate_phone, build_header_index as common_build_header_index, lookup_index as common_lookup_index

# Default column name candidates for robustness against slight header variations.
COLUMNS = {
    "name": ("姓名", "客户名称", "顾客姓名"),
    "phone": ("手机号", "电话", "联系方式"),
    "pay_date": ("顾客付款日期", "客户付款日期", "付款日期", "下单日期", "下单时间"),
    "status": ("状态",),
    "gross": ("收款额", "金额"),
    "net": ("净收款",),
    "profit": ("毛利", "利润估算"),
    "cost": ("打款金额", "打款", "打款价", "成本价", "成本"),
    "refund_amount": ("退款金额",),
    "refund_status": ("退货状态",),
    "refund_type": ("退款类型",),
    "refund_reason": ("退款原因",),
    "owner": ("负责人", "跟进人"),
    "platform": ("出售平台", "平台"),
    "address": ("地址", "收货地址"),
    "notes": ("备注", "货品备注"),
    "item": ("货品名", "商品名称"),
    "color": ("颜色", "色号"),
    "size": ("尺码", "规格", "码数"),
    "manufacturer": ("厂家", "有货厂家"),
    "order_no": ("单号", "订单号", "出库单号", "出单号"),
    "return_no": ("退货单号", "退货物流", "退货快递", "退货快递单号", "退货物流单号", "退回单号", "退货运单号"),
    "data_source": ("数据来源",),
}


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build customer alerts from billing data.")
    parser.add_argument("--source", required=True, help="Path to source Excel workbook.")
    parser.add_argument(
        "--sheet", default=None, help="Sheet name to read (defaults to first visible sheet)."
    )
    parser.add_argument("--output", default="客户预警输出.xlsx", help="Destination Excel file.")
    parser.add_argument(
        "--today",
        default=None,
        help="Override today's date (YYYY-MM-DD) for reproducible runs.",
    )
    parser.add_argument(
        "--value-top",
        type=float,
        default=0.2,
        help="Fraction of customers considered high value for churn alerts (default 0.2 → top 20%%).",
    )
    parser.add_argument(
        "--churn-days",
        type=int,
        default=90,
        help="Days without purchase that trigger high-value churn alerts.",
    )
    parser.add_argument(
        "--churn-multiplier",
        type=float,
        default=1.5,
        help="Multiplier applied to each customer's average repurchase cycle (个体复购周期倍数，默认 1.5 倍)。",
    )
    parser.add_argument(
        "--drop-threshold",
        type=float,
        default=0.5,
        help="Recent/previous spend ratio threshold for spend-drop alert (0.5 → drop >50%%).",
    )
    parser.add_argument(
        "--max-action",
        type=int,
        default=None,
        help="Limit number of high-priority customers surfaced in the action sheet (e.g. 20).",
    )
    parser.add_argument(
        "--anniversary-only",
        action="store_true",
        help="Only include customers with upcoming purchase anniversaries.",
    )
    parser.add_argument(
        "--anniversary-window",
        type=int,
        default=15,
        help="Window (days) for anniversary detection.",
    )
    parser.add_argument(
        "--anniversary-months",
        help="Comma-separated list of month offsets (e.g., '12' or '12,24') to check for anniversaries.",
    )
    parser.add_argument(
        "--config",
        default="tech/config.json",
        help="Path to business rules configuration JSON.",
    )
    parser.add_argument(
        "--html-output",
        default="客户预警仪表盘.html",
        help="Path to generate interactive HTML dashboard.",
    )
    parser.add_argument(
        "--allow-high-return",
        action="store_true",
        help="If set, customers with return rate > 30%% will NOT be filtered out of action list.",
    )
    parser.add_argument(
        "--contact-log",
        default="contact_log.xlsx",
        help="Path to contact_log.xlsx tracking recent outreach.",
    )
    parser.add_argument(
        "--exclude-recent-days",
        type=int,
        default=0,
        help="Exclude customers contacted within the last N days (requires contact_log).",
    )
    parser.add_argument(
        "--cooldown-days",
        type=int,
        default=0,
        help="Number of days for customer cooldown period after contact.",
    )
    parser.add_argument(
        "--cooldown-scope",
        default="action",
        help="Scope of cooldown: 'action' (only action list) or 'all' (entire dashboard).",
    )
    parser.add_argument(
        "--deepseek-key",
        default="sk-0d0e2d8d0a0141dcb4728068ba3d04ff",
        help="DeepSeek API Key for manufacturer analysis.",
    )
    return parser.parse_args()


def get_ai_manufacturer_analysis(mfr_name: str, sku_stats: List[Dict[str, Any]], api_key: str) -> str:
    """使用 DeepSeek API 生成厂家运营建议"""
    if not api_key:
        return ""
        
    url = "https://api.deepseek.com/chat/completions"
    prompt = f"""
    你是一位电商资深运营专家。请分析厂家【{mfr_name}】的货品表现数据，并给出具体、可操作的运营建议。
    
    数据分析重点：
    1. 识别出高风险款式（退货率高或毛利低）。
    2. 针对退货原因（如质量问题、色差、尺码不准等）给出改进建议。
    3. 语言要专业、干练，给运营看，不需要客套话。
    4. 结果控制在 250 字以内。
    
    待分析数据：
    {json.dumps(sku_stats, ensure_ascii=False, indent=2)}
    """
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    
    payload = {
        "model": "deepseek-chat",
        "messages": [
            {
                "role": "system",
                "content": "你是一个电商数据分析助手，擅长通过订单数据发现经营风险并提供战术建议。"
            },
            {
                "role": "user",
                "content": prompt
            }
        ],
        "temperature": 0.7,
        "stream": False
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=25)
        response.raise_for_status()
        result = response.json()
        return result['choices'][0]['message']['content']
    except Exception as e:
        return f"AI分析暂时不可用: {str(e)}"


def parse_month_offsets(text: str) -> List[int]:
    months: List[int] = []
    for chunk in text.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        try:
            value = int(chunk)
        except ValueError:
            continue
        if value > 0:
            months.append(value)
    return months


def shift_months(base: date, delta: int) -> date:
    total_months = base.year * 12 + base.month - 1 + delta
    year = total_months // 12
    month = total_months % 12 + 1
    last_day = calendar.monthrange(year, month)[1]
    day = min(base.day, last_day)
    return date(year, month, day)


def build_anniversary_dates(today: date, offsets: Iterable[int]) -> List[date]:
    unique_offsets = sorted({int(offset) for offset in offsets if offset > 0})
    return [shift_months(today, -offset) for offset in unique_offsets]


 


def normalize_header(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    return str(value).strip()


 


 


def try_get(row: Tuple, idx: Optional[int]):
    if idx is None:
        return None
    if idx >= len(row):
        return None
    return row[idx]


 


 


def load_contact_log(path: Path, today: date) -> Dict[str, date]:
    """
    Load a contact log (手机号, 最后联系日期) and return latest contact date per phone.
    Accepts optional header row; unknown columns default to first two columns.
    """
    contact_map: Dict[str, date] = {}
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    try:
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return contact_map

    header_candidates = rows[0]
    header_index = {}
    for idx, cell in enumerate(header_candidates):
        if cell is None:
            continue
        header_index[str(cell).strip()] = idx

    phone_idx = None
    date_idx = None
    phone_headers = ("手机号", "手机", "手机号码", "联系电话", "电话", "联系方式", "phone", "Phone")
    date_headers = ("最后联系日期", "最后联系日", "最近联系日期", "最近联系日", "员工联系日期", "联系日期", "last_contact", "LastContact")
    for header in phone_headers:
        if header in header_index:
            phone_idx = header_index[header]
            break
    for header in date_headers:
        if header in header_index:
            date_idx = header_index[header]
            break

    data_rows = rows[1:]
    if phone_idx is None or date_idx is None:
        # Assume first row is data as well when headers missing.
        data_rows = rows
        phone_idx = 0
        date_idx = 1 if len(rows[0]) > 1 else None

    if date_idx is None:
        return contact_map

    for row in data_rows:
        if row is None:
            continue
        phone_raw = row[phone_idx] if len(row) > phone_idx else None
        date_raw = row[date_idx] if len(row) > date_idx else None
        phone = common_deduplicate_phone(phone_raw)
        if not phone:
            continue
        contact_date = common_parse_excel_date(date_raw, today)
        if contact_date is None:
            continue
        prev = contact_map.get(phone)
        if prev is None or contact_date > prev:
            contact_map[phone] = contact_date
    return contact_map

def load_contact_log_extended(path: Path, today: date) -> Tuple[Dict[str, date], Dict[str, Dict[str, Any]]]:
    contact_map: Dict[str, date] = {}
    info_map: Dict[str, Dict[str, Any]] = {}
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    try:
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return contact_map, info_map
    header_candidates = rows[0]
    header_index = {}
    for idx, cell in enumerate(header_candidates):
        if cell is None:
            continue
        header_index[str(cell).strip()] = idx
    def idx_of(*names):
        for n in names:
            if n in header_index:
                return header_index[n]
        return None
    phone_idx = idx_of("手机号","手机","手机号码","联系电话","电话","联系方式","phone","Phone")
    date_idx = idx_of("最后联系日期","最后联系日","最近联系日期","最近联系日","员工联系日期","联系日期","last_contact","LastContact")
    emp_idx = idx_of("联系人","负责人","跟进人","employee","Employee")
    plat_idx = idx_of("联系平台","主要平台","平台","Platform","platform")
    status_idx = idx_of("回复状态","联系进程","状态标签","Status","status")
    note_idx = idx_of("备注","Note","note")
    next_idx = idx_of("下一次联系日","下次联系日","NextContact","next_contact")
    optout_idx = idx_of("不再联系","免打扰","OptOut","optout")
    happy_idx = idx_of("愉快值","满意度","Happiness","happiness")
    data_rows = rows[1:]
    for row in data_rows:
        if row is None:
            continue
        phone_raw = row[phone_idx] if (phone_idx is not None and len(row) > phone_idx) else None
        date_raw = row[date_idx] if (date_idx is not None and len(row) > date_idx) else None
        phone = common_deduplicate_phone(phone_raw)
        contact_date = common_parse_excel_date(date_raw, today)
        if phone and contact_date:
            prev = contact_map.get(phone)
            if prev is None or contact_date > prev:
                contact_map[phone] = contact_date
        if not phone:
            continue
        info: Dict[str, Any] = info_map.get(phone) or {}
        if emp_idx is not None and len(row) > emp_idx:
            info["employee"] = str(row[emp_idx]).strip() if row[emp_idx] else info.get("employee")
        if plat_idx is not None and len(row) > plat_idx:
            info["platform"] = str(row[plat_idx]).strip() if row[plat_idx] else info.get("platform")
        if status_idx is not None and len(row) > status_idx:
            info["status"] = str(row[status_idx]).strip() if row[status_idx] else info.get("status")
        if note_idx is not None and len(row) > note_idx:
            info["note"] = str(row[note_idx]).strip() if row[note_idx] else info.get("note")
        if next_idx is not None and len(row) > next_idx:
            nd = common_parse_excel_date(row[next_idx], today)
            if nd:
                info["next_contact"] = nd
        if optout_idx is not None and len(row) > optout_idx:
            val = str(row[optout_idx]).strip().lower() if row[optout_idx] else ""
            info["optout"] = val in ("是","yes","true","1")
        if happy_idx is not None and len(row) > happy_idx:
            try:
                info["happiness"] = float(row[happy_idx]) if row[happy_idx] is not None else info.get("happiness")
            except Exception:
                pass
        info_map[phone] = info
    return contact_map, info_map


def fetch_feishu_contact_log(app_token: str, table_id: str, today: date, *, token: Optional[str] = None, view_id: Optional[str] = None) -> Dict[str, date]:
    """从飞书多维表读取联系记录。

    为避免视图筛选丢失最新数据，默认“同时”读取：指定视图 + 全表，然后按手机号取最近日期。
    可通过环境变量 FEISHU_CONTACT_FETCH_MODE 控制：
      - both (默认)：视图 + 全表 合并
      - all：仅全表
      - view：仅指定视图
    """
    if not app_token or not table_id or not token:
        return {}

    base = f'https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records'
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json; charset=utf-8',
    }

    def _fetch_once(v_id: Optional[str]) -> Dict[str, date]:
        contact_map: Dict[str, date] = {}
        params = {'page_size': 200}
        if v_id:
            params['view_id'] = v_id
        page_token = None
        while True:
            if page_token:
                params['page_token'] = page_token
            resp = requests.get(base, headers=headers, params=params, timeout=30)
            try:
                resp.raise_for_status()
                data = resp.json()
                if isinstance(data, dict) and data.get('code') not in (None, 0):
                    break
            except Exception:
                break
            items = data.get('data', {}).get('items', []) or []
            for it in items:
                fields = it.get('fields', {}) or {}
                phone_raw = (
                    fields.get('手机号')
                    or fields.get('手机')
                    or fields.get('手机号码')
                    or fields.get('联系电话')
                    or fields.get('电话')
                    or fields.get('联系方式')
                )
                date_raw = (
                    fields.get('最后联系日期')
                    or fields.get('最后联系日')
                    or fields.get('最近联系日期')
                    or fields.get('最近联系日')
                    or fields.get('员工联系日期')
                    or fields.get('联系日期')
                    or fields.get('LastContact')
                    or fields.get('last_contact')
                )
                if not phone_raw:
                    try:
                        for k, v in fields.items():
                            if not isinstance(v, (str, int, float)):
                                continue
                            s = str(v)
                            digits = ''.join(ch for ch in s if ch.isdigit())
                            if len(digits) >= 7:
                                kn = str(k)
                                if any(t in kn for t in ('手', '电', 'phone', 'Phone')):
                                    phone_raw = v
                                    break
                    except Exception:
                        pass
                if date_raw is None:
                    try:
                        for k, v in fields.items():
                            if v is None:
                                continue
                            candidate: Optional[date] = None
                            if isinstance(v, (int, float)):
                                if float(v) > 10_000_000_000:
                                    try:
                                        candidate = datetime.fromtimestamp(float(v)/1000.0, tz=timezone.utc).astimezone().date()
                                    except Exception:
                                        candidate = None
                                else:
                                    try:
                                        candidate = datetime.fromtimestamp(float(v), tz=timezone.utc).astimezone().date()
                                    except Exception:
                                        candidate = None
                            else:
                                candidate = common_parse_excel_date(v, today)
                            if candidate is not None:
                                kn = str(k)
                                if any(t in kn for t in ('联系', '日期', '日', 'time', 'date')):
                                    date_raw = v
                                    break
                    except Exception:
                        pass
                phone = common_deduplicate_phone(phone_raw)
                contact_date: Optional[date]
                if isinstance(date_raw, (int, float)) and date_raw > 10_000_000_000:
                    try:
                        contact_date = datetime.fromtimestamp(float(date_raw)/1000.0, tz=timezone.utc).astimezone().date()
                    except Exception:
                        contact_date = None
                else:
                    # 支持 ISO8601 字符串
                    if isinstance(date_raw, str) and 'T' in date_raw:
                        try:
                            ss = str(date_raw).replace('Z', '+00:00')
                            dt = datetime.fromisoformat(ss)
                            contact_date = (dt.astimezone().date() if dt.tzinfo else dt.date())
                        except Exception:
                            contact_date = common_parse_excel_date(date_raw, today)
                    else:
                        contact_date = common_parse_excel_date(date_raw, today)
                if phone and contact_date:
                    prev = contact_map.get(phone)
                    if prev is None or contact_date > prev:
                        contact_map[phone] = contact_date
            page_token = data.get('data', {}).get('page_token')
            if not data.get('data', {}).get('has_more'):
                break
        return contact_map

    mode = (os.getenv('FEISHU_CONTACT_FETCH_MODE') or 'both').strip().lower()
    merged: Dict[str, date] = {}
    if mode in ('view', 'both') and view_id:
        view_map = _fetch_once(view_id)
        for ph, dt in view_map.items():
            if ph not in merged or dt > merged[ph]:
                merged[ph] = dt
        if not view_map:
            try:
                print(f"ℹ️  视图 {view_id} 未返回联系人或为空。")
            except Exception:
                pass
    if mode in ('all', 'both'):
        all_map = _fetch_once(None)
        for ph, dt in all_map.items():
            if ph not in merged or dt > merged[ph]:
                merged[ph] = dt
    return merged


class ConfigModel:
    __slots__ = (
        "defaults",
        "categories",
        "alias_map",
        "platform_touch_cost",
        "path",
        "priority_min",
        "priority_max",
        "orders_dampening",
        "single_order_enabled",
        "single_order_mode",
        "single_order_days",
        "timing_window_boost_config",
    )

    def __init__(self, raw: Dict[str, Any], path: Path):
        self.defaults: Dict[str, Any] = raw.get("defaults", {})
        self.categories: Dict[str, Dict[str, Any]] = raw.get("categories", {})
        self.platform_touch_cost: Dict[str, float] = raw.get("platform_touch_cost", {})
        # Optional per-order-count dampening mapping, e.g., {"1": 0.3, "2": 0.6, "default": 1.0}
        od_raw = raw.get("orders_dampening", {}) or {}
        orders_dampening: Dict[str, float] = {}
        if isinstance(od_raw, dict):
            for k, v in od_raw.items():
                try:
                    orders_dampening[str(k)] = float(v)
                except (TypeError, ValueError):
                    continue
        self.orders_dampening = orders_dampening
        # Single-order inclusion policy
        so = raw.get("single_order", {}) or {}
        self.single_order_enabled = bool(so.get("enabled", False))
        self.single_order_mode = str(so.get("mode", "previous_month")).strip() or "previous_month"
        try:
            self.single_order_days = int(so.get("days", 30))
        except Exception:
            self.single_order_days = 30
        alias_map: Dict[str, str] = {}
        for category, payload in self.categories.items():
            alias_map[category] = category
            for alias in payload.get("aliases", []) or []:
                alias_map[str(alias)] = category
        self.alias_map = alias_map
        self.path = path
        self.priority_min = float(self.defaults.get("min_priority_score", -50.0))
        self.priority_max = float(self.defaults.get("max_priority_score", 150.0))

        # Timing window boost parameters
        twb_raw = raw.get("timing_window_boost", {}) or {}
        self.timing_window_boost_config = {
            "enabled": bool(twb_raw.get("enabled", True)),
            "window_percentage": float(twb_raw.get("window_percentage", 0.20)),
            "peak_boost": float(twb_raw.get("peak_boost", 30.0)),
            "min_orders": int(twb_raw.get("min_orders", 2)),
            "max_return_rate": float(twb_raw.get("max_return_rate", 0.0)),
            "gaussian_k": float(twb_raw.get("gaussian_k", 2.0)),
        }

    def resolve_category_name(self, item_name: Optional[str]) -> Optional[str]:
        if not item_name:
            return None
        name = str(item_name).strip()
        if not name:
            return None
        return self.alias_map.get(name)

    def category_profile(self, category_name: Optional[str]) -> Dict[str, float]:
        defaults = {
            "gross_margin": float(self.defaults.get("gross_margin", 0.3)),
            "category_cycle_days": float(self.defaults.get("category_cycle_days", 60)),
            "expected_return_rate": float(self.defaults.get("expected_return_rate", 0.08)),
            "touch_cost": float(self.defaults.get("touch_cost", 6.0)),
            "max_estimated_margin": float(self.defaults.get("max_estimated_margin", 10000.0)),
            "max_estimated_uplift": float(self.defaults.get("max_estimated_uplift", 5.0)),
        }
        if category_name and category_name in self.categories:
            payload = self.categories[category_name]
            for key in (
                "gross_margin",
                "category_cycle_days",
                "expected_return_rate",
                "touch_cost",
                "max_estimated_margin",
                "max_estimated_uplift",
            ):
                if key in payload and payload[key] is not None:
                    try:
                        defaults[key] = float(payload[key])
                    except (TypeError, ValueError):
                        continue
        return defaults

    def uplift_params(self) -> Dict[str, float]:
        return {
            "base": float(self.defaults.get("uplift_base", 0.5)),
            "floor": float(self.defaults.get("uplift_floor", 0.1)),
            "ceiling": float(self.defaults.get("uplift_ceiling", 2.0)),
        }

    def platform_cost(self, platform_name: Optional[str], fallback: float) -> float:
        if not platform_name:
            return fallback
        return float(self.platform_touch_cost.get(str(platform_name).strip(), fallback))

    def priority_bounds(self) -> Tuple[float, float]:
        lower = min(self.priority_min, self.priority_max)
        upper = max(self.priority_min, self.priority_max)
        return lower, upper

    def orders_weight(self, orders: Optional[int]) -> float:
        """Return confidence weight based on order count, defaulting to 1.0.
        Looks up an explicit mapping by count string, falling back to 'default'.
        """
        try:
            count_key = str(int(orders)) if isinstance(orders, (int, float)) else None
        except Exception:
            count_key = None
        if not self.orders_dampening:
            return 1.0
        if count_key and count_key in self.orders_dampening:
            try:
                return float(self.orders_dampening[count_key])
            except (TypeError, ValueError):
                return float(self.orders_dampening.get("default", 1.0))
        return float(self.orders_dampening.get("default", 1.0))

    def timing_boost_params(self) -> Dict[str, Any]:
        """Return timing window boost configuration."""
        return self.timing_window_boost_config

    def allow_single_order(self, last_order: Optional[date], today: date) -> bool:
        """Return True if a single-order customer should be included in actions.
        Modes:
          - previous_month: only include if last_order falls in previous calendar month
          - days: include if last_order within last N days (single_order_days)
        When disabled or missing data, return True (no extra filter).
        """
        if not self.single_order_enabled:
            return True
        if last_order is None:
            return False
        mode = (self.single_order_mode or "previous_month").lower()
        if mode == "previous_month":
            prev_year = today.year
            prev_month = today.month - 1
            if prev_month == 0:
                prev_month = 12
                prev_year -= 1
            start = date(prev_year, prev_month, 1)
            last_day = calendar.monthrange(prev_year, prev_month)[1]
            end = date(prev_year, prev_month, last_day)
            return start <= last_order <= end
        if mode == "days":
            window_start = today - timedelta(days=max(0, self.single_order_days))
            return last_order >= window_start
        return True


def load_config(path: Path) -> ConfigModel:
    if not path.exists():
        raise FileNotFoundError(f"Config file not found: {path}")
    with path.open("r", encoding="utf-8") as fp:
        try:
            raw = json.load(fp)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Invalid JSON in config file {path}: {exc}") from exc
    return ConfigModel(raw, path)


def estimate_uplift(
    days_since: Optional[int],
    threshold: Optional[int],
    config_model: ConfigModel,
    extra_ceiling: Optional[float] = None,
    orders: Optional[int] = None,  # 订单数参数
    return_rate: Optional[float] = None,  # 退货率参数
    avg_order_value: float = 0.0,  # 客单价参数
) -> float:
    params = config_model.uplift_params()
    base = params["base"]
    floor = params["floor"]
    ceiling = params["ceiling"]
    
    # 思路3：老客户（3单+）uplift最低保底1.0
    if orders is not None and orders >= 3:
        floor = max(floor, 1.0)  # 老客户保底1.0以上
    
    if extra_ceiling is not None:
        try:
            ceiling = min(ceiling, float(extra_ceiling))
        except (TypeError, ValueError):
            pass
    if ceiling < floor:
        ceiling = floor
    if days_since is None or threshold is None or threshold <= 0:
        return max(floor, min(ceiling, base))
    
    ratio = days_since / max(threshold, 1)
    uplift = base + max(0.0, ratio - 1.0)
    
    # VIP客户检测：高订单数 + 低退货率 + 高客单价
    # VIP客户价值极高，无论何时都应该优先跟进，不受时间衰减影响
    is_vip = False
    if orders is not None and orders >= 7:
        # 条件1：7单+ 且退货率低于20%
        if return_rate is not None and return_rate < 0.2:
            is_vip = True
        # 条件2：7单+ 且退货率为None（说明没退过货）
        elif return_rate is None:
            is_vip = True
    # 条件3：5单+ 且退货率极低(<10%) 且客单价高(>400)
    elif orders is not None and orders >= 5:
        if return_rate is not None and return_rate < 0.1 and avg_order_value > 400:
            is_vip = True
    # 条件4：客单价特别高(>500) 且订单数>=3 且低退货
    elif orders is not None and orders >= 3 and avg_order_value > 500:
        if return_rate is not None and return_rate < 0.15:
            is_vip = True
        elif return_rate is None:
            is_vip = True
    
    # VIP客户的uplift逻辑：时间影响极小，保持高优先级
    if is_vip:
        # VIP客户只做轻微的时间调整，最多降低20%
        if days_since <= 30:
            # 0-30天：VIP也稍微降低，但只降10%（普通客户降30%）
            uplift = uplift * 0.9
        elif days_since <= 90:
            # 30-90天：VIP保持高uplift，轻微提升10%
            uplift = uplift * 1.1
        elif days_since <= 180:
            # 90-180天：VIP只轻微降低到95%（普通客户降到80%）
            uplift = uplift * 0.95
        else:
            # 180天+：VIP降到80%（普通客户降到50%）
            uplift = uplift * 0.8
    else:
        # 普通客户的Uplift动态调整策略（根据促单黄金期设计）
        # 0-30天：刚买过，没必要促单 → 降低70%
        # 30-90天：黄金跟进期，最应该促单 → 提升30-50%
        # 90-180天：开始流失 → 缓慢降低到80%
        # 180天+：基本流失 → 大幅降低到50%
        
        if days_since <= 30:
            # 0-30天：刚买过，不需要促单，uplift降低到70%
            uplift = uplift * 0.7
        elif days_since <= 90:
            # 30-90天：黄金跟进期！线性提升30-50%
            # days_since=30 -> boost=1.3 (70%回升到130%)
            # days_since=60 -> boost=1.4
            # days_since=90 -> boost=1.5
            boost_factor = 1.3 + (days_since - 30) / 60 * 0.2
            uplift = uplift * boost_factor
        elif days_since <= 180:
            # 90-180天：开始流失，线性衰减从150%到80%
            # days_since=90 -> factor=1.5
            # days_since=180 -> factor=0.8
            decay_factor = 1.5 - (days_since - 90) / 90 * 0.7
            uplift = uplift * decay_factor
        else:
            # 180天+：基本流失，uplift降低到50%
            uplift = uplift * 0.5
    
    return max(floor, min(ceiling, uplift))


def calculate_timing_boost(
    days_since: Optional[int],
    personal_cycle_days: Optional[float],
    stats: 'CustomerStats',
    return_rate: Optional[float],
    config_model: ConfigModel
) -> float:
    """
    Calculate timing window boost for customers in optimal repurchase window.

    掐点推荐：给处于最佳购买时机的客户加分。

    Uses Gaussian distribution to provide peak boost at cycle center,
    with gradual taper to window edges.

    Special rule: 换货 counts as return for 2-order customers.
    - If customer has 2 orders and one is exchange (换), treat as 1 effective order
    - Only customers with all valid orders (no returns/exchanges) qualify

    Args:
        days_since: Days since last order
        personal_cycle_days: Customer's average repurchase cycle
        stats: Customer statistics (for order count check)
        return_rate: Customer's return rate
        config_model: Configuration model

    Returns:
        Boost points (0 to peak_boost), or 0 if customer doesn't qualify
    """
    timing_params = config_model.timing_boost_params()
    window_percentage = timing_params["window_percentage"]
    peak_boost = timing_params["peak_boost"]
    min_orders = timing_params["min_orders"]
    max_return_rate = timing_params["max_return_rate"]
    enabled = timing_params["enabled"]
    gaussian_k = timing_params["gaussian_k"]

    # Early exit if feature disabled
    if not enabled:
        return 0.0

    # Eligibility checks
    if days_since is None or personal_cycle_days is None:
        return 0.0

    if stats.orders < min_orders:
        return 0.0

    # Check for exchanges (换货) in order details
    # 换货 should be treated as returns, especially for 2-order customers
    exchange_count = 0
    if hasattr(stats, 'order_details') and stats.order_details:
        for detail in stats.order_details:
            refund_type = str(detail.get("退款类型", "") or "").strip()
            # Check if refund type contains "换" (exchange)
            if "换" in refund_type:
                exchange_count += 1

    # Calculate effective orders (excluding exchanges)
    effective_orders = stats.orders - exchange_count

    # If effective orders fall below minimum after excluding exchanges, reject
    if effective_orders < min_orders:
        return 0.0

    # For customers with exchange orders, reject them completely
    # (换货 is treated as a type of return, so having any exchanges disqualifies them)
    if exchange_count > 0:
        return 0.0

    # Normal return rate check for customers without exchanges
    if return_rate is not None and return_rate > max_return_rate:
        return 0.0

    # Calculate window boundaries
    window_radius = personal_cycle_days * window_percentage
    lower_bound = personal_cycle_days - window_radius
    upper_bound = personal_cycle_days + window_radius

    # Check if customer is within window
    if days_since < lower_bound or days_since > upper_bound:
        return 0.0

    # Calculate boost using Gaussian distribution
    center = personal_cycle_days
    distance_from_center = abs(days_since - center)
    normalized_distance = distance_from_center / window_radius if window_radius > 0 else 0

    # Gaussian formula: peak_boost × exp(-k × normalized_distance²)
    boost = peak_boost * math.exp(-gaussian_k * (normalized_distance ** 2))

    return round(boost, 2)


def bucket_priority_score(score: float) -> Tuple[str, str]:
    if score >= 80:
        return "高(≥80)", "priority-high"
    if score >= 50:
        return "中(50-79)", "priority-mid"
    if score >= 0:
        return "低(0-49)", "priority-low"
    return "负分", "priority-other"


def classify_customer_value(
    net_total: float,
    orders: int,
    avg_order_value: float,
) -> str:
    """
    客户价值分层:
    - 高价值: 累计消费>5000 或 (3单+且AOV>500)
    - 中价值: 累计消费贵2000-5000 或 2单+
    - 低价值: 其他
    """
    if net_total >= 5000 or (orders >= 3 and avg_order_value >= 500):
        return "高价值"
    elif net_total >= 2000 or orders >= 2:
        return "中价值"
    else:
        return "低价值"


def classify_customer_list(
    personal_cycle_days: Optional[float],
    avg_order_value: float,
    top_platform: Optional[str],
    orders: int,
    return_rate: Optional[float],
    potential_label: str,
    days_since: Optional[int],
    net_total: float,
    tags_text: str,
    priority_score: float,
    long_term_threshold: int,
) -> str:
    """
    分类客户到四个列表
    
    列表1: 超级VIP列表
    列表2: 活跃培养列表
    列表3: 濒临流失列表
    列表4: 高风险待排查列表
    """
    # 列表4: 高风险待排查（只放真正有退货问题的客户）
    if "退货激增" in tags_text:
        return "高风险待排查"
    if return_rate is not None and return_rate > 0.4 and priority_score >= 50:
        return "高风险待排查"
    # 移除消费骤降判断，因为优质客户也会有消费骤降，不代表风险
    
    # 列表1: 超级VIP列表（买的多 + 退的少 + 优先分高）
    # 条件1: 超多订单 + 低退货
    if orders >= 7 and (return_rate is None or return_rate < 0.2):
        return "超级VIP"
    # 条件2: 高累计消费 + 低退货 + 至少5单（避充2单客户被误判）
    if net_total >= 2000 and orders >= 5 and (return_rate is None or return_rate < 0.25):
        return "超级VIP"
    # 条件3: 优先分特别高 + 至少5单（说明综合质量好）
    if priority_score >= 200 and orders >= 5:
        return "超级VIP"
    # 条件4: 明星客户 + 至少5单
    if potential_label == "明星客户" and orders >= 5:
        return "超级VIP"
    
    # 列表2: 活跃培养列表
    if (days_since is not None and days_since < 90 and 
        orders >= 2 and 
        (return_rate is None or return_rate < 0.3)):
        return "活跃培养"
    
    # 列表3: 濒临流失列表
    if (days_since is not None and days_since >= 90 and 
        days_since <= long_term_threshold * 3 and
        (orders >= 3 or net_total >= 1000) and
        (return_rate is None or return_rate < 0.5)):
        return "濒临流失"
    
    # 默认归入活跃培养（兼容其他情况）
    return "活跃培养"


def compute_customer_lifecycle_value(
    stats: "CustomerStats",
    windows: Dict[str, float],
    today: date,
    category_profile: Dict[str, float],
) -> Tuple[float, str, str]:
    """
    计算客户生命周期价值(CLV)及分类
    
    Returns:
        (clv_score, growth_type, potential_label)
        - clv_score: 综合生命周期价值分数
        - growth_type: 成长类型（成长型/稳定型/下滑型/新客型）
        - potential_label: 潜力标签（明星客户/潜力客户/普通客户）
    """
    # 1. 历史价值分 (40%)
    historical_score = 0.0
    if stats.net_total > 0:
        # 累计消费分数，上限10000
        historical_score += min(100, (stats.net_total / 10000) * 100) * 0.25
        # 订单数分数，上限20单
        historical_score += min(100, (stats.orders / 20) * 100) * 0.15
    
    # 2. 当前活跃度分 (30%)
    activity_score = 0.0
    if stats.last_order:
        days_since = (today - stats.last_order).days
        # 近期活跃度，30天内满分
        recency_score = max(0, 100 - (days_since / 180) * 100)
        activity_score += recency_score * 0.15
        # 近90天消费占比
        if stats.net_total > 0:
            recent_ratio = windows.get("days_90", 0) / max(stats.net_total, 1)
            activity_score += min(100, recent_ratio * 200) * 0.15
    
    # 3. 增长潜力分 (30%)
    growth_score = 0.0
    growth_type = "新客型"
    
    # 计算增长趋势
    recent_90 = windows.get("days_90", 0)
    prev_90 = windows.get("prev_90", 0)
    recent_30 = windows.get("days_30", 0)
    
    if stats.orders >= 2:  # 老客户
        if prev_90 > 0:
            growth_rate = (recent_90 - prev_90) / prev_90
            if growth_rate > 0.2:
                growth_type = "成长型"
                growth_score += min(100, growth_rate * 100) * 0.2
            elif growth_rate < -0.2:
                growth_type = "下滑型"
                growth_score += max(0, 50 + growth_rate * 50) * 0.2
            else:
                growth_type = "稳定型"
                growth_score += 60 * 0.2
        else:
            growth_type = "稳定型" if recent_90 > 0 else "休眠型"
            growth_score += (60 if recent_90 > 0 else 20) * 0.2
            
        # 复购频率稳定性
        if stats.orders >= 3:
            # 高频稳定复购加分
            avg_cycle = (stats.last_order - stats.first_order).days / (stats.orders - 1) if stats.last_order and stats.first_order else 999
            if avg_cycle < 60:  # 高频复购
                growth_score += 10 * 0.1
        
        # AOV增长趋势
        avg_order_value = stats.net_total / stats.orders if stats.orders else 0
        if recent_90 > 0 and recent_30 > 0:
            recent_aov = recent_30  # 简化计算
            if recent_aov > avg_order_value * 1.2:
                growth_score += 10 * 0.1  # AOV增长加分
    else:  # 新客户
        if recent_30 > 200:  # 首单高价值
            growth_type = "高潜新客"
            growth_score += 80 * 0.3
        else:
            growth_type = "新客型"
            growth_score += 50 * 0.3
    
    # 总分
    clv_score = historical_score + activity_score + growth_score
    
    # 潜力标签判定
    if clv_score >= 70:
        potential_label = "明星客户"
    elif clv_score >= 40 or growth_type in ("成长型", "高潜新客"):
        potential_label = "潜力客户"
    else:
        potential_label = "普通客户"
    
    return clv_score, growth_type, potential_label


def explain_priority_score(
    priority_score: float,
    estimated_uplift: float,
    estimated_margin: float,
    estimated_return_rate: float,
    orders: int,
    days_since: Optional[int],
    threshold: int,
    total_spend: float,
) -> str:
    """
    生成优先分的解释说明，帮助销售人员理解促单理由
    
    Args:
        priority_score: 优先分
        estimated_uplift: 估算转化概率
        estimated_margin: 估算毛利
        estimated_return_rate: 估算退货率
        orders: 订单数
        days_since: 未复购天数
        threshold: 预警阈值
        total_spend: 累计消费金额
        
    Returns:
        解释文本
    """
    reasons = []
    
    # 1. 时机判断（最重要的触达理由）
    if days_since is not None:
        ratio = days_since / threshold if threshold > 0 else 0
        if ratio >= 1.2:
            reasons.append(f"已{days_since}天未买，超过预警线")
        elif ratio >= 1.0:
            reasons.append(f"{days_since}天未买，快要流失了")
        elif ratio >= 0.8:
            reasons.append(f"{days_since}天未买，快到复购期")
    
    # 2. 转化潜力（用白话说明）
    if estimated_uplift >= 1.3:
        reasons.append("现在联系成功率高")
    elif estimated_uplift >= 1.0:
        reasons.append("联系有机会成交")
    
    # 3. 利润价值（直接说能赚多少）
    if estimated_margin >= 500:
        reasons.append(f"预计能赚{estimated_margin:.0f}元")
    elif estimated_margin >= 200:
        reasons.append(f"预计能赚{estimated_margin:.0f}元")
    
    # 4. 退货风险（直白说明）
    if estimated_return_rate <= 0.1:
        reasons.append("基本不退货")
    elif estimated_return_rate >= 0.3:
        reasons.append(f"退货率高({estimated_return_rate:.0%})要小心")
    
    # 5. 客户价值（用累计消费说明）
    if total_spend >= 5000:
        reasons.append(f"买过{total_spend:.0f}元的好客户")
    elif total_spend >= 2000:
        reasons.append(f"累计买了{total_spend:.0f}元")
    
    # 6. 订单信心（用老客/新客说明）
    if orders >= 10:
        reasons.append(f"{orders}次老客户")
    elif orders >= 3:
        reasons.append(f"{orders}次老客户")
    elif orders == 2:
        reasons.append("买2次，可以培养")
    elif orders == 1:
        reasons.append("刚买1次的新客")
    
    # 综合解释
    if not reasons:
        return f"综合评分{priority_score:.0f}分"
    
    return "、".join(reasons)


def build_sop_recommendations(
    tags: List[str],
    favorite_item: Optional[str],
    platform: Optional[str],
    customer_value: str,
) -> str:
    """
    根据标签、平台和客户价值层级生成差异化推荐动作。
    
    策略框架:
    - 微信渠道: 高价值(1v1私聊) / 中价值(小福利唤醒) / 低价值(限时秒杀)
    - 闲鱼渠道: 统一促销策略(券/独家款/超低折扣)
    """
    favorite = favorite_item or "重点单品"
    platform_hint = platform or "主力平台"
    is_xianyu = platform and "闲鱼" in platform
    
    # 优先处理特殊标签（无论平台和价值层级）
    priority_mapping = {
        "退货激增": "【售后排查】(禁止促单，重点核实退款原因)",
        "退货率": "【质量复盘】(内部检视退货链路，暂缓促单)",
    }
    
    seen: Dict[str, str] = {}
    for tag in tags:
        if tag in priority_mapping:
            seen[tag] = priority_mapping[tag]
    
    # 如果有高优先级标签，直接返回
    if seen:
        return ", ".join(seen.values())
    
    # 闲鱼渠道：统一促销策略
    if is_xianyu:
        xianyu_mapping = {
            "高价值流失预警": f"【老用户回访】(话术：生日专属福利+{favorite}独家款)",
            "长期未复购": "【限时秒杀】(话术：超低折扣+朋友圈超低折扣)",
            "短期未复购": f"【直接送东西】(话术：无门槛券+{favorite}直送)",
            "消费骤降": "【独家款式】(话术：新品限量+看完下架)",
            "节点回访": "【生日惊喜】(话术：生日专属红包+独家福利)",
            "高价值活跃": f"【VIP专属】(话术：{favorite}独家渠道款+超值组合)",
        }
        for tag in tags:
            if tag in xianyu_mapping:
                seen[tag] = xianyu_mapping[tag]
        
        if not seen:
            return f"【促销策略】(话术：限时秒杀+{favorite}超低折扣+看完下架)"
        return ", ".join(seen.values())
    
    # 微信渠道：根据客户价值分层
    if customer_value == "高价值":
        wechat_high_mapping = {
            "高价值流失预警": f"【1v1专属私聊】(话术：{favorite}定制复购权益+专属折扣)",
            "长期未复购": f"【1v1暖场关怀】(话术：询问近期需求+{favorite}新品预览)",
            "短期未复购": f"【1v1新品推荐】(话术：{favorite}升级款OneOne模仿教程+买拍测评)",
            "消费骤降": "【1v1回访排查】(话术：询问体验反馈+物流服务改进)",
            "节点回访": "【1v1纪念日惊喜】(话术：专属纪念日券+定制福利)",
            "高价值活跃": f"【1v1增购建议】(话术：{favorite}同系列搭配+学习电商外卖平台2)",
        }
        for tag in tags:
            if tag in wechat_high_mapping:
                seen[tag] = wechat_high_mapping[tag]
        
        if not seen:
            return f"【1v1专属服务】(话术：{favorite}复购专享+OneOne模仿教程)"
    
    elif customer_value == "中价值":
        wechat_mid_mapping = {
            "高价值流失预警": f"【小福利唤醒】(话术：{favorite}限时优惠+复购红包)",
            "长期未复购": "【小福利唤醒】(话术：新品补货通知+限时折扣)",
            "短期未复购": f"【小福利唤醒】(话术：{favorite}热卖搭配+小红包)",
            "消费骤降": "【小福利唤醒】(话术：体验改进回馈+补偿券)",
            "节点回访": "【小福利唤醒】(话术：纪念日红包+专属优惠)",
            "高价值活跃": f"【小福利唤醒】(话术：{favorite}新品上架+会员福利)",
        }
        for tag in tags:
            if tag in wechat_mid_mapping:
                seen[tag] = wechat_mid_mapping[tag]
        
        if not seen:
            return f"【小福利唤醒】(话术：{favorite}限时活动+小额红包)"
    
    else:  # 低价值
        wechat_low_mapping = {
            "高价值流失预警": f"【限时秒杀】(话术：{favorite}特价清仓+最后机会)",
            "长期未复购": "【限时秒杀】(话术：断码清仓+超低价格)",
            "短期未复购": f"【特价清仓】(话术：{favorite}库存处理+限时抢购)",
            "消费骤降": "【特价清仓】(话术：品质保证+清仓优惠)",
            "节点回访": "【限时秒杀】(话术：纪念日特价+仅此一次)",
            "高价值活跃": f"【特价清仓】(话术：{favorite}组合优惠+批量折扣)",
        }
        for tag in tags:
            if tag in wechat_low_mapping:
                seen[tag] = wechat_low_mapping[tag]
        
        if not seen:
            return f"【限时秒杀+特价清仓】(话术：{favorite}断码处理+看完下架)"
    
    return ", ".join(seen.values())


class CustomerStats:
    __slots__ = (
        "key",
        "name",
        "phone",
        "address",
        "owners",
        "platforms",
        "items",
        "first_order",
        "last_order",
        "orders",
        "gross_total",
        "net_total",
        "cost_total",
        "profit_total",
        "refund_amount",
        "refund_count",
        "cancel_count",
        "order_history",
        "order_details",
    )

    def __init__(self, key: str):
        self.key = key
        self.name: Optional[str] = None
        self.phone: Optional[str] = None
        self.address: Optional[str] = None
        self.owners: Counter = Counter()
        self.platforms: Counter = Counter()
        self.items: Counter = Counter()
        self.first_order: Optional[date] = None
        self.last_order: Optional[date] = None
        self.orders: int = 0
        self.gross_total: float = 0.0
        self.net_total: float = 0.0
        self.cost_total: float = 0.0
        self.profit_total: float = 0.0
        self.refund_amount: float = 0.0
        self.refund_count: int = 0
        self.cancel_count: int = 0
        self.order_history: List[Tuple[date, float]] = []  # (order_date, net_amount)
        self.order_details: List[Dict[str, Any]] = []  # raw rows for HTML drilldown

    def register_valid_order(
        self,
        order_date: Optional[date],
        gross: float,
        net: float,
        cost: float,
        profit: float,
        item: Optional[str],
        owner: Optional[str],
        platform: Optional[str],
        address: Optional[str],
    ) -> None:
        self.orders += 1
        self.gross_total += gross
        self.net_total += net
        self.cost_total += cost
        self.profit_total += profit
        if order_date:
            self.order_history.append((order_date, net))
            if self.first_order is None or order_date < self.first_order:
                self.first_order = order_date
            if self.last_order is None or order_date > self.last_order:
                self.last_order = order_date
        if item:
            self.items[item.strip()] += 1
        if owner:
            self.owners[owner.strip()] += 1
        if platform:
            self.platforms[platform.strip()] += 1
        if address and not self.address:
            self.address = address.strip()

    def register_cancellation(self) -> None:
        self.cancel_count += 1

    def register_refund(self, amount: float, *, force: bool = False) -> None:
        """
        Register a refund event.
        - If amount > 0: always accumulate refund_amount and increment refund_count.
        - If amount <= 0: increment refund_count only when force=True (e.g., identified via 状态/类型字段)。
        """
        if amount is None:
            amount = 0.0
        if amount > 0:
            self.refund_amount += amount
            self.refund_count += 1
            return
        if force:
            self.refund_count += 1

    def append_order_detail(
        self,
        name: Optional[str],
        phone: Optional[str],
        platform: Optional[str],
        item: Optional[str],
        pay_amount: Optional[float],
        pay_cost: Optional[float],
        refund_type: Optional[str],
        refund_reason: Optional[str],
        order_date: Optional[date] = None,
        order_no: Optional[str] = None,
        return_no: Optional[str] = None,
        manufacturer: Optional[str] = None,
        notes: Optional[str] = None,
        color: Optional[str] = None,
        size: Optional[str] = None,
        data_source: Optional[str] = None,
        *,
        gross_amount: Optional[float] = None,
        net_amount: Optional[float] = None,
    ) -> None:
        gross_val = float(gross_amount) if isinstance(gross_amount, (int, float)) else (float(pay_amount) if isinstance(pay_amount, (int, float)) else 0.0)
        net_val = float(net_amount) if isinstance(net_amount, (int, float)) else 0.0
        self.order_details.append(
            {
                "姓名": (name or "").strip() if name else "",
                "手机号": (phone or "").strip() if phone else "",
                "下单平台": (platform or "").strip() if platform else "",
                "厂家": (manufacturer or "").strip() if manufacturer else "",
                "货品名": (item or "").strip() if item else "",
                "商品名称": (item or "").strip() if item else "",
                "颜色": (color or "").strip() if color else "",
                "尺码": (size or "").strip() if size else "",
                # 为下游模块统一提供：付款金额=收款额
                "付款金额": gross_val,
                # 同时保留原始口径，便于排查
                "收款额": gross_val,
                "净收款": net_val,
                "打款金额": float(pay_cost) if isinstance(pay_cost, (int, float)) else (float(pay_cost) if str(pay_cost or "").strip() else 0.0),
                "退款类型": (str(refund_type).strip() if refund_type is not None else ""),
                "退款原因": (str(refund_reason).strip() if refund_reason is not None else ""),
                "备注": (str(notes).strip() if notes is not None else ""),
                "下单时间": order_date.isoformat() if isinstance(order_date, date) else "",
                "订单号": str(order_no).strip() if order_no is not None else "",
                "退货单号": str(return_no).strip() if return_no is not None else "",
                "数据来源": (str(data_source).strip() if data_source is not None else ""),
            }
        )


 


def choose_primary(counter: Counter) -> Optional[str]:
    if not counter:
        return None
    item, _ = counter.most_common(1)[0]
    return item


def compute_time_windows(entries: List[Tuple[date, float]], today: date) -> Dict[str, float]:
    window_totals = {
        "days_30": 0.0,
        "days_90": 0.0,
        "prev_90": 0.0,
        "days_180": 0.0,
        "days_365": 0.0,
    }
    cutoff_30 = today - timedelta(days=30)
    cutoff_90 = today - timedelta(days=90)
    cutoff_prev_90 = today - timedelta(days=180)
    cutoff_180 = today - timedelta(days=180)
    cutoff_365 = today - timedelta(days=365)
    for order_date, net in entries:
        if order_date >= cutoff_30:
            window_totals["days_30"] += net
        if order_date >= cutoff_90:
            window_totals["days_90"] += net
        elif order_date >= cutoff_prev_90:
            window_totals["prev_90"] += net
        if order_date >= cutoff_180:
            window_totals["days_180"] += net
        if order_date >= cutoff_365:
            window_totals["days_365"] += net
    return window_totals


def build_customer_key(name: Optional[str], phone: Optional[str], address: Optional[str]) -> str:
    phone_part = phone or ""
    name_part = (name or "").strip()
    if phone_part:
        return phone_part
    if name_part and address:
        return f"{name_part}|{address.strip()}"
    if name_part:
        return name_part
    return address.strip() if address else "未知客户"


def load_customers(ws, today: date) -> Dict[str, CustomerStats]:
    iterator = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        raise ValueError("Worksheet is empty.")
    header_index = common_build_header_index(header_row)

    indices = {key: common_lookup_index(header_index, value) for key, value in COLUMNS.items()}

    customers: Dict[str, CustomerStats] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        name_raw = try_get(row, indices["name"])
        phone_raw = try_get(row, indices["phone"])
        address_raw = try_get(row, indices["address"])
        owner_raw = try_get(row, indices["owner"])
        platform_raw = try_get(row, indices["platform"])
        item_raw = try_get(row, indices["item"])
        manufacturer_raw = try_get(row, indices.get("manufacturer")) if indices.get("manufacturer") is not None else None
        status_raw = try_get(row, indices["status"])
        gross_raw = try_get(row, indices["gross"])
        net_raw = try_get(row, indices["net"])
        profit_raw = try_get(row, indices["profit"])
        cost_raw = try_get(row, indices["cost"]) if indices.get("cost") is not None else None
        refund_raw = try_get(row, indices["refund_amount"])
        refund_status_raw = try_get(row, indices["refund_status"])
        notes_raw = try_get(row, indices.get("notes")) if indices.get("notes") is not None else None
        refund_type_raw = try_get(row, indices.get("refund_type")) if indices.get("refund_type") is not None else None
        refund_reason_raw = try_get(row, indices.get("refund_reason")) if indices.get("refund_reason") is not None else None
        pay_raw = try_get(row, indices["pay_date"])
        order_no_raw = try_get(row, indices.get("order_no")) if indices.get("order_no") is not None else None
        return_no_raw = try_get(row, indices.get("return_no")) if indices.get("return_no") is not None else None
        data_source_raw = try_get(row, indices.get("data_source")) if indices.get("data_source") is not None else None
        color_raw = try_get(row, indices.get("color")) if indices.get("color") is not None else None
        size_raw = try_get(row, indices.get("size")) if indices.get("size") is not None else None

        phone = common_deduplicate_phone(phone_raw)
        key = build_customer_key(name_raw, phone, address_raw)

        stats = customers.get(key)
        if stats is None:
            stats = CustomerStats(key=key)
            customers[key] = stats

        if name_raw and not stats.name:
            stats.name = str(name_raw).strip()
        if phone and not stats.phone:
            stats.phone = phone
        if address_raw and not stats.address:
            stats.address = str(address_raw).strip()

        pay_date = common_parse_excel_date(pay_raw, today)
        status = str(status_raw).strip() if status_raw else ""
        refund_type_text = str(refund_type_raw).strip() if refund_type_raw is not None else ""
        # 取消识别：优先看行状态；若状态不含“取消”，但退款类型标记为“取消”，也视为取消单
        is_cancelled = ("取消" in status) or ("取消" in refund_type_text)

        gross = common_to_float(gross_raw)
        net = common_to_float(net_raw) if indices["net"] is not None else gross
        # 标记成本是否真实存在（避免把缺失当0参与“净收款-0”）
        cost_present = (indices.get("cost") is not None) and (cost_raw is not None) and (str(cost_raw).strip() != "")
        cost = common_to_float(cost_raw) if cost_present else 0.0
        # 忽略表中“毛利/利润估算”列，统一按口径计算；若无成本数据则记0，后续用毛利率估算
        profit = 0.0
        refund_amount = common_to_float(refund_raw)

        # Treat empty net as gross to avoid losing value on incomplete entries.
        if math.isclose(net, 0.0) and gross > 0:
            net = gross
        # 毛利口径：若存在打款金额/成本列，则优先按（收款额或净收款）- 打款金额
        if cost_present:
            base_amount = net if net > 0 else gross
            profit = base_amount - cost

        # Append order detail for drilldown (all rows)
        # Use computed net (fallbacks to gross when net is empty)
        # 在明细中统一使用"收款额"（gross）作为付款金额，避免净收款导致金额缩小的问题
        pay_amount_for_detail = gross
        stats.append_order_detail(
            name=str(name_raw).strip() if name_raw else (stats.name or ""),
            phone=phone,
            platform=str(platform_raw).strip() if platform_raw else None,
            item=str(item_raw).strip() if item_raw else None,
            pay_amount=pay_amount_for_detail,
            pay_cost=(cost if cost_present else None),
            refund_type=refund_type_raw if refund_type_raw is not None else refund_status_raw,
            refund_reason=refund_reason_raw,
            order_date=pay_date,
            order_no=str(order_no_raw).strip() if order_no_raw is not None else None,
            return_no=str(return_no_raw).strip() if return_no_raw is not None else None,
            manufacturer=str(manufacturer_raw).strip() if manufacturer_raw is not None else None,
            notes=str(notes_raw).strip() if notes_raw is not None else None,
            color=str(color_raw).strip() if color_raw is not None else None,
            size=str(size_raw).strip() if size_raw is not None else None,
            data_source=str(data_source_raw).strip() if data_source_raw is not None else None,
            gross_amount=gross,
            net_amount=net,
        )

        if is_cancelled or gross <= 0:
            stats.register_cancellation()
        else:
            stats.register_valid_order(
                order_date=pay_date,
                gross=gross,
                net=net,
                cost=cost,
                profit=profit,
                item=item_raw if item_raw else None,
                owner=owner_raw if owner_raw else None,
                platform=platform_raw if platform_raw else None,
                address=address_raw if address_raw else None,
            )

        # 退款识别：退款金额>0，或退货状态含“退”，或退款类型含“退” 均视为一笔退款
        if (refund_amount > 0) or (
            refund_status_raw and "退" in str(refund_status_raw)
        ) or (refund_type_text and "退" in refund_type_text):
            stats.register_refund(refund_amount, force=True)

    return customers


def determine_threshold(customers: Dict[str, CustomerStats], fraction: float) -> float:
    totals = sorted((stats.net_total for stats in customers.values()), reverse=True)
    if not totals:
        return 0.0
    cutoff_index = max(0, min(len(totals) - 1, int(len(totals) * fraction) - 1))
    return totals[cutoff_index]


def build_alert_rows(
    customers: Dict[str, CustomerStats],
    today: date,
    high_value_threshold: float,
    churn_days: int,
    churn_multiplier: float,
    drop_threshold: float,
    config_model: ConfigModel,
    anniversary_dates: Optional[List[date]] = None,
    anniversary_window: int = 0,
    anniversary_only: bool = False,
    contact_log: Optional[Dict[str, date]] = None,
    contact_info: Optional[Dict[str, Dict[str, Any]]] = None,
    cooldown_days: int = 0,
    cooldown_scope: str = "action",
    exclude_recent_days: int = 30,
    allow_high_return: bool = False,
) -> Tuple[List[List], List[Dict[str, Any]], int, Dict[str, Dict[str, Any]], Dict[str, date]]:
    overview_rows: List[List] = []
    action_rows: List[Dict[str, Any]] = []
    meta_map: Dict[str, Dict[str, Any]] = {}
    cooldown_customers: Dict[str, date] = {}  # 新增：存储冷却期客户及其联系日期
    cooldown_days = max(0, cooldown_days)
    contact_log = contact_log or {}
    contact_info = contact_info or {}
    snoozed_total = 0

    def classify_reply_status(status: Optional[str]) -> Optional[bool]:
        """
        根据联系记录中的“回复状态”等字段，粗略判断客户是否有回复。

        返回:
            True  -> 明确有回复/已沟通/成交
            False -> 明确无回复/拒绝/不需要
            None  -> 状态未知，不参与过滤
        """
        if not status:
            return None
        text = str(status).strip().lower()
        if not text:
            return None
        no_reply_keywords = (
            "无回复",
            "未回复",
            "没回复",
            "未回",
            "不回",
            "不理",
            "拒绝",
            "不需要",
            "不想要",
            "拉黑",
        )
        positive_keywords = (
            "已回复",
            "有回复",
            "已沟通",
            "已联系",
            "成交",
            "成单",
            "已加",
            "加微",
            "加微信",
            "已购买",
            "已下单",
        )
        if any(k in text for k in no_reply_keywords):
            return False
        if any(k in text for k in positive_keywords):
            return True
        return None

    for stats in customers.values():
        if not stats.orders:
            continue

        windows = compute_time_windows(stats.order_history, today)
        last_order = stats.last_order
        days_since = (today - last_order).days if last_order else None
        personal_cycle_days: Optional[float] = None
        if stats.orders >= 2 and stats.first_order and stats.last_order:
            order_dates = sorted({entry[0] for entry in stats.order_history if entry[0]})
            if len(order_dates) >= 2:
                intervals: List[int] = []
                prev_date = order_dates[0]
                for current_date in order_dates[1:]:
                    gap = (current_date - prev_date).days
                    intervals.append(gap if gap > 0 else 1)
                    prev_date = current_date
                if intervals:
                    personal_cycle_days = sum(intervals) / len(intervals)
            if personal_cycle_days is None and stats.orders > 1:
                span_days = (stats.last_order - stats.first_order).days
                if span_days < 0:
                    span_days = 0
                personal_cycle_days = span_days / (stats.orders - 1) if (stats.orders - 1) else None
                if personal_cycle_days is not None and personal_cycle_days <= 0:
                    personal_cycle_days = 1.0

        personalized_threshold = None
        if personal_cycle_days is not None:
            personalized_threshold = personal_cycle_days * churn_multiplier
            if personalized_threshold <= 0:
                personalized_threshold = None

        top_owner = choose_primary(stats.owners)
        top_platform = choose_primary(stats.platforms)
        top_item = choose_primary(stats.items)

        category_name = config_model.resolve_category_name(top_item)
        category_profile = config_model.category_profile(category_name)
        category_cycle = max(1.0, category_profile["category_cycle_days"])
        margin_cap = max(0.0, category_profile.get("max_estimated_margin", 10000.0))
        uplift_cap = category_profile.get("max_estimated_uplift")
        category_threshold = max(1, math.ceil(category_cycle * churn_multiplier))
        fallback_threshold = max(1, churn_days)

        threshold_candidates: List[int] = [fallback_threshold, category_threshold]
        if personalized_threshold is not None:
            threshold_candidates.append(max(1, math.ceil(personalized_threshold)))
        long_term_threshold = max(threshold_candidates) if threshold_candidates else fallback_threshold
        short_term_threshold = max(1, math.ceil(long_term_threshold / 2))
        if short_term_threshold >= long_term_threshold and long_term_threshold > 1:
            short_term_threshold = long_term_threshold - 1

        high_value = stats.net_total >= high_value_threshold and high_value_threshold > 0

        tags: List[str] = []
        anniversary_match = False
        if last_order and anniversary_dates:
            for target in anniversary_dates:
                if abs((last_order - target).days) <= anniversary_window:
                    anniversary_match = True
                    break

        if days_since is not None:
            if high_value and days_since >= long_term_threshold:
                tags.append("高价值流失预警")
            elif days_since >= long_term_threshold:
                tags.append("长期未复购")
            elif days_since >= short_term_threshold:
                tags.append("短期未复购")

        recent = windows["days_90"]
        previous = windows["prev_90"]
        drop_ratio = None
        if previous > 0:
            drop_ratio = recent / previous if previous else None
            if recent <= previous * drop_threshold:
                tags.append("消费骤降")

        if stats.refund_amount > 0 and stats.net_total > 0:
            refund_ratio = stats.refund_amount / max(stats.net_total, 1)
            if refund_ratio >= 0.3 or stats.refund_count >= 2:
                tags.append("退货激增")

        if anniversary_match:
            if "节点回访" not in tags:
                tags.append("节点回访")

        if not tags and stats.net_total >= high_value_threshold and days_since is not None:
            if days_since < max(30, short_term_threshold):
                tags.append("高价值活跃")

        return_rate = None
        if stats.orders > 0:
            return_rate = min(stats.refund_count / max(stats.orders, 1), 1.0)
        elif stats.refund_count > 0:
            return_rate = 1.0
        
        # 判断是否为优质客户，使用更高的Uplift上限（方案E）
        use_premium_uplift = False
        if stats.orders >= config_model.defaults.get("uplift_premium_min_orders", 3):
            if return_rate is not None and return_rate < config_model.defaults.get("uplift_premium_max_return", 0.3):
                # 检查时间跨度（首单到末单）
                if stats.first_order and last_order:
                    days_span = (last_order - stats.first_order).days
                    if days_span >= config_model.defaults.get("uplift_premium_min_days_span", 30):
                        use_premium_uplift = True
        
        # 如果是优质客户，使用更高的uplift上限
        effective_uplift_cap = uplift_cap
        if use_premium_uplift:
            premium_ceiling = config_model.defaults.get("uplift_ceiling_premium", 3.5)
            effective_uplift_cap = premium_ceiling if uplift_cap is None else min(premium_ceiling, uplift_cap)
        
        # 计算客单价（需要在estimate_uplift之前）
        avg_order_value = stats.net_total / stats.orders if stats.orders else 0.0
        
        # 传入订单数、退货率、客单价，让estimate_uplift可以判断是否为VIP客户
        estimated_uplift = estimate_uplift(
            days_since, 
            long_term_threshold, 
            config_model, 
            effective_uplift_cap, 
            orders=stats.orders,
            return_rate=return_rate,
            avg_order_value=avg_order_value
        )
        avg_profit_per_order = (
            stats.profit_total / stats.orders if stats.orders and stats.profit_total > 0 else 0.0
        )
        estimated_margin = avg_profit_per_order if avg_profit_per_order > 0 else avg_order_value * category_profile["gross_margin"]
        estimated_margin = max(0.0, estimated_margin)
        if margin_cap > 0:
            estimated_margin = min(estimated_margin, margin_cap)
        expected_return = category_profile["expected_return_rate"]
        observed_return = return_rate if isinstance(return_rate, (int, float)) else 0.0
        if stats.orders >= 3 and return_rate is not None:
            estimated_return_rate = min(0.95, 0.7 * observed_return + 0.3 * expected_return)
        elif return_rate is not None:
            estimated_return_rate = min(0.95, max(observed_return, expected_return))
        else:
            estimated_return_rate = min(0.95, expected_return)
        touch_cost = max(0.0, config_model.platform_cost(top_platform, category_profile["touch_cost"]))

        aov_value = stats.net_total / stats.orders if stats.orders else None
        aov_trend = drop_ratio

        # 计算CLV分数和分类
        clv_score, growth_type, potential_label = compute_customer_lifecycle_value(
            stats, windows, today, category_profile
        )
        
        # Base (pre-weight) score before confidence dampening
        priority_score_pre_weight = (estimated_uplift * estimated_margin * (1 - estimated_return_rate)) - touch_cost

        # 换货订单检查（在所有boost之前）
        # 检查客户订单中是否有换货，如果有换货则计算有效订单数
        exchange_count = 0
        if hasattr(stats, 'order_details') and stats.order_details:
            for detail in stats.order_details:
                refund_type = str(detail.get("退款类型", "") or "").strip()
                if "换" in refund_type:
                    exchange_count += 1

        # 计算有效订单数（排除换货）
        effective_orders = stats.orders - exchange_count

        # 如果有换货订单且有效订单数不足，应用严厉惩罚
        # 换货本身说明客户对商品不满意，视为负面信号
        exchange_penalty = 0
        if exchange_count > 0:
            # 有换货订单的客户，根据有效订单数决定惩罚力度
            if effective_orders < 2:
                # 有效订单少于2单，严厉惩罚（基本排除）
                exchange_penalty = -200  # 直接-200分，几乎排除出触达列表
            elif effective_orders == 2:
                # 有效订单2单，中度惩罚
                exchange_penalty = -100
            elif effective_orders <= 4:
                # 有效订单3-4单，轻度惩罚
                exchange_penalty = -50
            # 5单以上的换货客户，不额外惩罚（已经是老客户了）

        # CLV加权：给高CLV客户额外加分（优化订单数加成）
        clv_boost = 0
        if potential_label == "明星客户":
            clv_boost = 30  # 明星客户+30分
        elif potential_label == "潜力客户":
            clv_boost = 15  # 潜力客户+15分
        elif stats.orders >= 10:
            clv_boost = 40  # 10单以上老客+40分（提升自20）
        elif stats.orders >= 7:
            clv_boost = 30  # 7-9单+30分（新增）
        elif stats.orders >= 5:
            clv_boost = 20  # 5-6单+20分（提升自10）
        elif stats.orders >= 4:
            clv_boost = 10  # 4单+10分（新增）
        elif stats.orders >= 3:
            clv_boost = 5   # 3单+5分（新增，因为3单已是前4%）
        
        # 低退货率加成（方案B）
        return_rate_boost = 0
        # 强力惩罚高退货：≥49% 直接重扣，避免高退货客户顶到前面
        if return_rate is not None and return_rate >= 0.49:
            return_rate_boost = -80
        elif return_rate is not None and return_rate <= 0.0:
            return_rate_boost = 20  # 零退货
        elif return_rate is not None and return_rate < 0.1:
            return_rate_boost = 15  # 退货率<10%
        elif return_rate is not None and return_rate < 0.2:
            return_rate_boost = 10  # 退货率10-20%
        elif return_rate is not None and return_rate < 0.3:
            return_rate_boost = 5   # 退货率20-30%
        
        # 高客单价加成（方案C）
        aov_boost = 0
        avg_order_value_calc = stats.net_total / stats.orders if stats.orders else 0
        if avg_order_value_calc > 500:
            aov_boost = 25
        elif avg_order_value_calc > 400:
            aov_boost = 15
        elif avg_order_value_calc > 300:
            aov_boost = 10
        elif avg_order_value_calc > 200:
            aov_boost = 5
        
        # 活跃优质客户加成（思路1）
        # 短期未购 + 老客户 + 低退货 = 高潜力，需要一鼓作气发展成高频用户
        activity_boost = 0
        if days_since is not None and days_since < 45:  # 45天内未购（约1.5个月）
            if stats.orders >= 3 and return_rate is not None and return_rate < 0.2:  # 老客户+低退货
                activity_boost = 50  # +50分，重要！
            elif stats.orders >= 2 and return_rate is not None and return_rate < 0.3:  # 2单+中等退货
                activity_boost = 30  # +30分
        
        # 平台质量加成（第3点）
        # 根据平台的历史表现给予不同加成
        platform_boost = 0
        if top_platform:
            platform_quality = {
                "相册": 15,      # 高质量平台（平均109.8分）
                "三店": 15,      # 高质量（平均109.8分）
                "二店": 10,      # 中高质量（平均93.2分）
                "四店": 8,       # 中等质量（平均87.7分）
                "一店": 5,       # 中等（平均73.5分）
                "咸鱼": -5,      # 质量一般（平均62.3分）
                "咸鱼二": -5,   # 质量一般
                "代发": -10,     # 低质量（平均53分）
            }
            platform_boost = platform_quality.get(top_platform, 0)
        
        # 高频优质客户超级加成（第6点）
        # 这是你的优质资产，需要超级无敵跟进！
        high_freq_boost = 0
        if personal_cycle_days is not None and personal_cycle_days < 30 and top_platform != "代发":
            # 高频用户（复购周期<30天）
            if avg_order_value_calc > 300:  # 高频+高客单价 = 超级VIP
                high_freq_boost = 60  # 超级加成！
            else:
                high_freq_boost = 40
        
        # 超级VIP加成（核心逻辑！）
        # 高订单数 + 低退货率 = 你的最优质资产，应该始终排在最前面！
        super_vip_boost = 0
        if stats.orders >= 10 and (return_rate is None or return_rate < 0.2):
            # 10单+低退货 = 超级VIP
            super_vip_boost = 80  # 大幅加成！
        elif stats.orders >= 7 and (return_rate is None or return_rate < 0.15):
            # 7单+极低退货 = 高级VIP
            super_vip_boost = 60
        elif stats.orders >= 5 and (return_rate is None or return_rate < 0.1):
            # 5单+零退货 = VIP
            super_vip_boost = 40

        # 掐点推荐加成（新功能）
        # 客户处于个人复购周期的最佳时间窗口（默认±20%）
        timing_window_boost = calculate_timing_boost(
            days_since=days_since,
            personal_cycle_days=personal_cycle_days,
            stats=stats,
            return_rate=return_rate,
            config_model=config_model
        )

        # 汇总所有加成
        total_boost = clv_boost + return_rate_boost + aov_boost + activity_boost + platform_boost + high_freq_boost + super_vip_boost + timing_window_boost + exchange_penalty
        
        # 低订单数+长时间未购的额外惩罚（关键逻辑！）
        # 2-4单且180天+的客户，基本流失，不值得高优先级跟进
        long_dormant_penalty = 0
        if days_since is not None and days_since >= 180:
            if stats.orders == 2:
                # 2单+180天+ = 大幅降低优先级
                long_dormant_penalty = -100  # -100分！
            elif stats.orders == 3:
                # 3单+180天+ = 中度降低
                long_dormant_penalty = -60
            elif stats.orders == 4:
                # 4单+180天+ = 轻度降低
                long_dormant_penalty = -40
            # 5单+不惩罚，因为他们是VIP，已经有VIP保护逻辑
        
        # 将惩罚加入total_boost
        total_boost = total_boost + long_dormant_penalty
        
        # Apply order-based confidence weight (e.g., 1单=0.3, 2单=0.7, 3单=0.9, ≥4单=1.0)
        orders_weight = config_model.orders_weight(stats.orders)
        priority_score_unbounded = priority_score_pre_weight * orders_weight + total_boost
        lower_bound, upper_bound = config_model.priority_bounds()
        priority_score_clamped = min(upper_bound, max(lower_bound, priority_score_unbounded))
        priority_score = round(priority_score_clamped, 2)
        bucket_label, bucket_class = bucket_priority_score(priority_score_clamped)

        # 客户价值分层
        customer_value = classify_customer_value(stats.net_total, stats.orders, avg_order_value)

        sop_actions = build_sop_recommendations(tags, top_item, top_platform, customer_value)
        tags_text = ", ".join(tags)
        actions_text = sop_actions

        cycle_display = personal_cycle_days
        trigger_days_display = (
            round(personalized_threshold, 1) if personalized_threshold is not None else long_term_threshold
        )
        effective_threshold_days = long_term_threshold

        overview_rows.append(
            [
                stats.key,
                stats.name or "",
                stats.phone or "",
                top_owner or "",
                top_platform or "",
                stats.address or "",
                stats.first_order.isoformat() if stats.first_order else "",
                last_order.isoformat() if last_order else "",
                days_since if days_since is not None else "",
                round(cycle_display, 1) if cycle_display is not None else "",
                trigger_days_display,
                round(category_cycle, 1),
                round(estimated_uplift, 3),
                round(estimated_margin, 2),
                round(estimated_return_rate, 3),
                round(touch_cost, 2),
                priority_score,
                timing_window_boost,
                stats.orders,
                stats.cancel_count,
                round(stats.gross_total, 2),
                round(stats.net_total, 2),
                round(return_rate, 4) if return_rate is not None else "",
                round(windows["days_30"], 2),
                round(windows["days_90"], 2),
                round(windows["prev_90"], 2),
                round(windows["days_180"], 2),
                round(windows["days_365"], 2),
                round(drop_ratio, 3) if drop_ratio is not None else "",
                stats.refund_count,
                round(stats.refund_amount, 2),
                tags_text,
                actions_text,
                top_item or "",
                round(clv_score, 2),
                growth_type,
                potential_label,
            ]
        )

        # 全量元信息映射（供前端“全库命中”临时行填充真实分数/标签等）
        meta_map[stats.key] = {
            "priority_bucket": bucket_label,
            "priority_class": bucket_class,
            "priority_score": float(priority_score),
            "customer_value": customer_value,
            "favorite": top_item or "",
            "orders": int(stats.orders),
            "return_rate": float(return_rate) if isinstance(return_rate, (int, float)) else None,
            "days": int(days_since) if isinstance(days_since, int) else None,
            "aov": float(aov_value) if isinstance(aov_value, (int, float)) else None,
            "platform": top_platform or "",
            "last_order": last_order.isoformat() if last_order else "",
            "threshold_display": trigger_days_display,
            "category_cycle": float(category_cycle),
        }
        ci_meta = contact_info.get(stats.phone or "") or {}
        status_text = str(ci_meta.get("status") or "").strip()
        reply_flag = classify_reply_status(status_text)
        last_contact_date = contact_log.get(stats.phone) if (contact_log and stats.phone) else None
        try:
            meta_map[stats.key].update({
                "contact_employee": ci_meta.get("employee") or "",
                "contact_platform": ci_meta.get("platform") or "",
                "contact_status": ci_meta.get("status") or "",
                "contact_note": ci_meta.get("note") or "",
                "next_contact": (ci_meta.get("next_contact")).isoformat() if isinstance(ci_meta.get("next_contact"), date) else "",
                "happiness": float(ci_meta.get("happiness")) if isinstance(ci_meta.get("happiness"), (int,float)) else "",
            })
        except Exception:
            pass

        # 冷却范围为“全库”时：若手机号在冷却期内，则直接计入隐藏并跳过（不进入行动清单）
        if (
            cooldown_scope == "all"
            and cooldown_days > 0
            and last_contact_date
        ):
            delta_all = (today - last_contact_date).days
            if delta_all < 0:
                delta_all = 0
            # 明确标记为“无回复/不需要”等的客户，直接从触达名单排除（不再重复推送）
            if reply_flag is False:
                # 保留概览与 meta，但不进入行动清单
                continue
            if delta_all < cooldown_days:
                snoozed_total += 1
                cooldown_customers[stats.key] = last_contact_date
                # 即使不加入行动清单，仍然保留概览与meta映射
                continue

        # 放宽条件：不只看tags，也看潜力标签和优先分
        include_in_action = False
        if tags and not (anniversary_only and not anniversary_match):
            include_in_action = True
        elif potential_label in ("明星客户", "潜力客户"):
            include_in_action = True
        elif priority_score >= 50:
            include_in_action = True
        elif stats.orders >= 5:
            include_in_action = True
        elif growth_type in ("成长型", "高潜新客"):
            include_in_action = True
        
        if include_in_action:
            # 额外规则：若最近一次联系被明确标记为“无回复/不需要”，则不再进入触达名单
            if last_contact_date and reply_flag is False:
                continue

            # 检查是否在冷却期（仅对未被判定为“无回复”的客户）
            in_cooldown = False
            if cooldown_days > 0 and last_contact_date:
                delta_days = (today - last_contact_date).days
                if delta_days < 0:
                    delta_days = 0
                if delta_days < cooldown_days:
                    snoozed_total += 1
                    # 保存冷却期客户信息
                    cooldown_customers[stats.key] = last_contact_date
                    in_cooldown = True
                    # 不再 continue，而是将冷却期客户加入行动清单，标记为"冷却期"列表

            # 如果不在冷却期，应用其他过滤规则
            if not in_cooldown:
                # Exclude recent buyers within exclude_recent_days from the action list
                if isinstance(days_since, int) and days_since < max(0, exclude_recent_days):
                    continue
                # Exclude customers with a very high return rate from the action list
                if (not allow_high_return) and isinstance(return_rate, (int, float)) and return_rate >= 0.49:
                    continue
                # Enforce single-order inclusion policy for action list
                if stats.orders == 1 and not config_model.allow_single_order(last_order, today):
                    continue
                # 放宽流失阈值：原来是2倍，现在放宽到3倍，但明星客户不限制
                if potential_label != "明星客户" and days_since is not None and days_since > long_term_threshold * 3:
                    continue

            # 生成优先分解释
            priority_explanation = explain_priority_score(
                priority_score=priority_score,
                estimated_uplift=estimated_uplift,
                estimated_margin=estimated_margin,
                estimated_return_rate=estimated_return_rate,
                orders=stats.orders,
                days_since=days_since,
                threshold=effective_threshold_days,
                total_spend=stats.net_total,
            )

            # 判断客户属于哪个列表
            if in_cooldown:
                # 冷却期客户单独分类
                customer_list = "冷却期"
            else:
                customer_list = classify_customer_list(
                    personal_cycle_days=personal_cycle_days,
                    avg_order_value=avg_order_value_calc,
                    top_platform=top_platform,
                    orders=stats.orders,
                    return_rate=return_rate,
                    potential_label=potential_label,
                    days_since=days_since,
                    net_total=stats.net_total,
                    tags_text=tags_text,
                    priority_score=priority_score,
                    long_term_threshold=long_term_threshold,
                )
            
            ci = contact_info.get(stats.phone or "") or {}
            action_rows.append(
                {
                    "customer_list": customer_list,  # 新增：所属列表
                    "priority_bucket": bucket_label,
                    "priority_class": bucket_class,
                    "priority_score": priority_score,
                    # Raw value used for sorting (post-dampening, pre-rounding, post-clamp preserved via priority_score)
                    "priority_score_raw": priority_score,
                    # For diagnostics: before clamp but after dampening
                    "priority_score_unbounded": round(priority_score_unbounded, 2),
                    # For diagnostics: before dampening
                    "priority_score_pre_weight": round(priority_score_pre_weight, 2),
                    "orders_weight": round(orders_weight, 3),
                    "mark": "",
                    "name": stats.name or "",
                    "last_order": last_order.isoformat() if last_order else "",
                    "phone": stats.phone or "",
                    "tags": tags_text,
                    "actions": actions_text,
                    "priority_explanation": priority_explanation,
                    "customer_value": customer_value,
                    "clv_score": round(clv_score, 2),  # CLV分数
                    "growth_type": growth_type,  # 成长类型
                    "potential_label": potential_label,  # 潜力标签
                    "favorite": top_item or "",
                    "owner": top_owner or "",
                    "orders": stats.orders,
                    "return_rate": return_rate,
                    "days": days_since if days_since is not None else None,
                    "aov": aov_value,
                    "aov_trend": aov_trend,
                    "platform": top_platform or "",
                    "cycle": cycle_display,
                    "category_cycle": category_cycle,
                    "threshold_days": effective_threshold_days,
                    "threshold_display": trigger_days_display,
                    "estimated_uplift": estimated_uplift,
                    "estimated_margin": estimated_margin,
                    "estimated_return_rate": estimated_return_rate,
                    "touch_cost": touch_cost,
                    # detail mapping for HTML row click
                    "detail_key": stats.key,
                    "details": list(stats.order_details),
                    "contact_employee": ci.get("employee") or "",
                    "contact_platform": ci.get("platform") or "",
                    "contact_status": ci.get("status") or "",
                    "contact_note": ci.get("note") or "",
                    "next_contact": (ci.get("next_contact")).isoformat() if isinstance(ci.get("next_contact"), date) else "",
                    "happiness": float(ci.get("happiness")) if isinstance(ci.get("happiness"), (int,float)) else "",
                }
            )

    def action_sort_key(entry: Dict[str, Any]):
        score = entry.get("priority_score_raw", 0.0)
        orders_value = -(entry["orders"] if isinstance(entry["orders"], (int, float)) else 0)
        rate_value = entry["return_rate"] if isinstance(entry["return_rate"], (int, float)) else 1.1
        aov_value = entry["aov"] if isinstance(entry["aov"], (int, float)) else 0.0
        return (-score, orders_value, rate_value, -aov_value)

    action_rows.sort(key=action_sort_key)
    return overview_rows, action_rows, snoozed_total, meta_map, cooldown_customers


def autofit_columns(ws) -> None:
    return


def write_workbook(
    output_path: Path,
    overview_rows: List[List],
    action_rows: List[Dict[str, Any]],
    today: date,
    high_value_threshold: float,
    config: argparse.Namespace,
    anniversary_dates: Optional[List[date]] = None,
    contact_log_used: bool = False,
    cooldown_days: int = 0,
    snoozed_total: int = 0,
) -> None:
    wb = Workbook(write_only=True)
    ws_overview = wb.create_sheet(title="客户概览")
    ws_overview.append(
        [
            "客户ID",
            "姓名",
            "手机号",
            "主要负责人",
            "主要平台",
            "最近地址",
            "首次下单日",
            "最近下单日",
            "未复购天数",
            "平均复购周期(天)",
            "流失预警阈值(天)",
            "品类周期(天)",
            "估算Uplift",
            "估算毛利",
            "估算退货率",
            "触达成本",
            "优先分",
            "时间窗口加分",
            "有效订单数",
            "取消单数",
            "累计收款额",
            "累计净收款",
            "退货率",
            "近30天净收款",
            "近90天净收款",
            "前90天净收款",
            "近180天净收款",
            "近365天净收款",
            "近90天/前90天比",
            "退款单数",
            "退款金额",
            "风险标签",
            "推荐动作",
            "偏好单品",
            "CLV分数",
            "成长类型",
            "潜力标签",
        ]
    )
    for row in overview_rows:
        ws_overview.append(row)
    autofit_columns(ws_overview)

    ws_actions = wb.create_sheet("触达优先级")
    ws_actions.append(
        [
            "标记完成",
            "优先分",
            "姓名",
            "主要平台",
            "手机号",
            "最近下单日",
            "风险标签",
            "推荐动作",
            "促单理由",
            "价值层级",
            "CLV分数",
            "成长类型",
            "潜力标签",
            "偏好单品",
            "主要负责人",
            "有效订单数",
            "退货率",
            "未复购天数",
            "复购周期(天)",
            "品类周期(天)",
            "预警阈倿(天)",
            "估算Uplift",
            "估算毛利",
            "估算退货率",
            "触达成本",
            "AOV趋势",
            "平均客单价",
            "联系员工",
            "回复状态",
            "愉快值",
            "下一次联系日",
            "备注",
        ]
    )
    for entry in action_rows:
        ws_actions.append(
            [
                "",
                round(entry.get("priority_score"), 2),
                entry["name"],
                entry["platform"],
                entry["phone"],
                entry["last_order"],
                entry["tags"],
                entry["actions"],
                entry.get("priority_explanation", ""),
                entry.get("customer_value", ""),
                round(entry.get("clv_score", 0), 2),  # CLV分数
                entry.get("growth_type", ""),  # 成长类型
                entry.get("potential_label", ""),  # 潜力标签
                entry["favorite"],
                entry["owner"],
                entry["orders"],
                round(entry["return_rate"], 4)
                if isinstance(entry["return_rate"], (int, float))
                else entry["return_rate"] or "",
                entry["days"] if entry["days"] is not None else "",
                round(entry.get("cycle"), 1) if isinstance(entry.get("cycle"), (int, float)) else "",
                round(entry.get("category_cycle"), 1)
                if isinstance(entry.get("category_cycle"), (int, float))
                else "",
                round(entry.get("threshold_display"), 1)
                if isinstance(entry.get("threshold_display"), (int, float))
                else entry.get("threshold_display") or "",
                round(entry.get("estimated_uplift"), 3)
                if isinstance(entry.get("estimated_uplift"), (int, float))
                else "",
                round(entry.get("estimated_margin"), 2)
                if isinstance(entry.get("estimated_margin"), (int, float))
                else "",
                round(entry.get("estimated_return_rate"), 3)
                if isinstance(entry.get("estimated_return_rate"), (int, float))
                else "",
                round(entry.get("touch_cost"), 2)
                if isinstance(entry.get("touch_cost"), (int, float))
                else "",
                round(entry.get("aov_trend"), 3)
                if isinstance(entry.get("aov_trend"), (int, float))
                else "",
                round(entry["aov"], 2) if isinstance(entry["aov"], (int, float)) else "",
                entry.get("contact_employee", ""),
                entry.get("contact_status", ""),
                round(entry.get("happiness"), 1) if isinstance(entry.get("happiness"), (int, float)) else entry.get("happiness", ""),
                entry.get("next_contact", ""),
                entry.get("contact_note", ""),
            ]
        )
    autofit_columns(ws_actions)
    
    # 生成四个列表Sheet
    list_names = [("超级VIP", "超级VIP"), ("活跃培养", "活跃培养"), ("濒临流失", "濒临流失"), ("高风险待排查", "高风险待排查")]
    
    for list_title, list_key in list_names:
        list_rows = [entry for entry in action_rows if entry.get("customer_list") == list_key]
        
        ws_list = wb.create_sheet(list_title)
        ws_list.append(
            [
                "所属列表",
                "标记完成",
                "优先分",
                "姓名",
                "主要平台",
                "手机号",
                "最近下单日",
                "风险标签",
                "推荐动作",
                "促单理由",
                "价值层级",
                "CLV分数",
                "成长类型",
                "潜力标签",
                "偏好单品",
                "主要负责人",
                "有效订单数",
                "退货率",
                "未复购天数",
                "复购周期(天)",
                "平均客单价",
            ]
        )
        
        for entry in list_rows:
            ws_list.append(
                [
                    list_key,  # 所属列表
                    "",  # 标记完成
                    round(entry.get("priority_score"), 2),
                    entry["name"],
                    entry["platform"],
                    entry["phone"],
                    entry["last_order"],
                    entry["tags"],
                    entry["actions"],
                    entry.get("priority_explanation", ""),
                    entry.get("customer_value", ""),
                    round(entry.get("clv_score", 0), 2),
                    entry.get("growth_type", ""),
                    entry.get("potential_label", ""),
                    entry["favorite"],
                    entry["owner"],
                    entry["orders"],
                    round(entry["return_rate"], 4)
                    if isinstance(entry["return_rate"], (int, float))
                    else entry["return_rate"] or "",
                    entry["days"] if entry["days"] is not None else "",
                    round(entry.get("cycle"), 1) if isinstance(entry.get("cycle"), (int, float)) else "",
                    round(entry["aov"], 2) if isinstance(entry["aov"], (int, float)) else "",
                ]
            )
        autofit_columns(ws_list)

    ws_meta = wb.create_sheet("指标说明")
    ws_meta.append(["生成日期", today.isoformat()])
    default_long = max(1, config.churn_days)
    default_short_half = default_long // 2
    default_short = max(1, min(30, default_short_half if default_short_half else default_long))
    ws_meta.append(
        [
            "高价值阈值(净收款)",
            round(high_value_threshold, 2) if high_value_threshold else 0.0,
        ]
    )
    ws_meta.append(["高价值占比", config.value_top])
    if getattr(config, "config", None):
        try:
            ws_meta.append(["配置文件", Path(config.config).name])
        except Exception:
            ws_meta.append(["配置文件", str(config.config)])
    ws_meta.append(["复购周期倍数", f"{config.churn_multiplier:.2f}x"])
    ws_meta.append(["优先分公式", "（估算Uplift × 估算毛利 × (1-估算退货率)）-触达成本"])
    ws_meta.append(["默认流失预警天数", default_long])
    ws_meta.append(["消费骤降阈值", f"近90天 ÷ 前90天 ≤ {config.drop_threshold}"])
    if contact_log_used:
        ws_meta.append(["冷却期设置", f"{max(0, cooldown_days)} 天内联系过客户不再推送"])
        ws_meta.append(["今日冷却客户数", snoozed_total])
    if config.max_action:
        ws_meta.append(["触达名单每日上限", config.max_action])
    # Single-order policy meta
    try:
        cfg_model = load_config(Path(getattr(config, "config", "config.json")))
    except Exception:
        cfg_model = None
    if cfg_model and getattr(cfg_model, "single_order_enabled", False):
        mode_text = cfg_model.single_order_mode
        detail = (
            "仅展示上一个自然月的首单客户"
            if mode_text == "previous_month"
            else f"仅展示近 {max(0, cfg_model.single_order_days)} 天内的首单客户"
        )
        ws_meta.append(["首单客户策略", detail])
    if anniversary_dates:
        ws_meta.append(
            [
                "纪念日目标日期",
                ", ".join(d.isoformat() for d in anniversary_dates),
            ]
        )
        if config.anniversary_window:
            ws_meta.append(["纪念日匹配容差", f"±{config.anniversary_window} 天"])
    if getattr(config, "anniversary_only", False):
        ws_meta.append(["纪念日筛选", "仅保留纪念日客户"])
    ws_meta.append([])
    ws_meta.append(["数据说明", "字段来源于原始账单，取消单不计入营收与订单统计。"])
    ws_meta.append(["平均复购周期计算", "（最近下单日-首次下单日）/(有效订单数-1)，若结果≤0 按 1 天计。"])
    ws_meta.append(["估算Uplift说明", "综合个体复购阈值与品类周期得出超期倍数，受 config.json 限制。"])
    ws_meta.append(["毛利口径", "优先使用 收款额−打款金额；无打款金额时，优先用账面毛利；仍缺失时用 平均客单价×毛利率 估算。"])
    ws_meta.append(["触达成本参考", "默认取品类触达成本，若配置了平台成本则优先使用平台值。"])
    ws_meta.append(["风险标签", "触发条件"])
    value_percent = int(round(config.value_top * 100))
    multiplier_text = f"{config.churn_multiplier:.2f}"
    ws_meta.append(
        [
            "高价值流失预警",
            f"净收款进入前 {value_percent}% 且未复购天数 ≥ ceil(平均复购周期 * {multiplier_text})，"
            f"当缺少复购历史时使用默认阈值 {default_long} 天。",
        ]
    )
    ws_meta.append(
        [
            "长期未复购",
            f"未复购天数 ≥ ceil(平均复购周期 * {multiplier_text})，无复购历史时默认阈值 {default_long} 天。",
        ]
    )
    ws_meta.append(
        [
            "短期未复购",
            f"未复购天数 ≥ 上述阈值的一半，缺少历史时约为 {default_short} 天。",
        ]
    )
    ws_meta.append(["消费骤降", f"近90天净收款 ≤ 前90天 * {config.drop_threshold}"])
    ws_meta.append(["退货激增", "退款金额占累计净收款 ≥30% 或退款单数 ≥2"])
    ws_meta.append(["退货率", "退款单数 ÷ 有效订单数，若无订单但有退款视为 100%"])
    ws_meta.append(["过滤策略", "已排除退货率 ≥49% 的客户（触达名单）"])
    ws_meta.append(["近期购买过滤", f"排除近 {max(0, int(getattr(config, 'exclude_recent_days', 30)))} 天下单客户（触达名单）"])
    ws_meta.append(["高价值活跃", "净收款在高价值范围，未复购天数 <30 天"])
    autofit_columns(ws_meta)

    wb.save(output_path)


def build_product_search_index(global_details: Dict[str, List[Dict[str, Any]]]) -> dict:
    """
    构建货品名反向索引：规范化货品名 -> [手机号列表]

    Args:
        global_details: 全局订单明细字典，key为手机号，value为订单列表

    Returns:
        {"fz1103": ["13800138000", "13900139000"], ...}
    """
    import re
    from collections import defaultdict

    product_index = defaultdict(set)

    def normalize_product_name(name: str) -> str:
        """规范化：去除非字母数字字符，转小写"""
        if not name:
            return ''
        # 保留中文、英文字母和数字，移除其他字符
        normalized = re.sub(r'[^a-z0-9\u4e00-\u9fff]', '', str(name).lower())
        return normalized

    # 遍历所有客户的订单明细
    for phone, orders in global_details.items():
        for order in orders:
            product_name = order.get('货品名', '')
            normalized = normalize_product_name(product_name)
            if normalized:
                product_index[normalized].add(str(phone))

    # 转换为可序列化格式
    result = {k: list(v) for k, v in product_index.items()}

    print(f"[索引] 货品名索引构建完成: {len(result)}个关键词")
    return result


def write_html_dashboard(
    output_path: Path,
    today: date,
    action_rows: List[Dict[str, Any]],
    overview_rows: List[List],
    high_value_threshold: float,
    config: argparse.Namespace,
    anniversary_dates: Optional[List[date]] = None,
    contact_log_used: bool = False,
    cooldown_days: int = 0,
    snoozed_total: int = 0,
    global_details: Optional[Dict[str, List[Dict[str, Any]]]] = None,
    global_meta: Optional[Dict[str, Dict[str, Any]]] = None,
    cooldown_keys: Optional[List[str]] = None,
    cooldown_customers: Optional[Dict[str, date]] = None,
    product_search_index: Optional[Dict[str, List[str]]] = None,
) -> None:
    headers = [
        "标记完成",
        "优先分",
        "姓名",
        "主要平台",
        "手机号",
        "最近下单日",
        "风险标签",
        "推荐动作",
        "促单理由",
        "价值层级",
        "偏好单品",
        "有效订单数",
        "退货率",
        "未复购天数",
        "平均客单价",
    ]
    numeric_headers = {
        "优先分",
        "有效订单数",
        "退货率",
        "未复购天数",
        "复购周期(天)",
        "品类周期(天)",
        "预警阈值(天)",
        "估算Uplift",
        "估算毛利",
        "估算退货率",
        "触达成本",
        "平均客单价",
        "最近下单日",
    }
    header_cells: List[str] = []
    for header in headers:
        if header in numeric_headers:
            header_cells.append(f"<th data-sort-method='number'>{escape(header)}</th>")
        else:
            header_cells.append(f"<th>{escape(header)}</th>")

    summary_counts = Counter()
    tag_counter = Counter()
    platform_counter = Counter()

    priority_order = ["高(≥80)", "中(50-79)", "低(0-49)", "负分"]

    def to_sort_key(value: Optional[float]) -> Optional[str]:
        if isinstance(value, (int, float)):
            numeric = float(value)
            if math.isnan(numeric) or math.isinf(numeric):
                return None
            # Preserve precision for sorting without overwhelming decimals.
            return format(numeric, ".15g")
        return None

    def default_summary_table(counter: Counter, order: List[str]) -> str:
        if not counter:
            return "<p>暂无数据。</p>"
        rows = []
        total = sum(counter.values())
        for key in order:
            count = counter.get(key, 0)
            if not count:
                continue
            label = key
            pct_value = (count / total) * 100.0 if total else 0.0
            pct = f"{pct_value:.1f}%"
            rows.append(
                f"<tr><td>{escape(label)}</td>"
                f"<td data-sort-value='{count}'>{count}</td>"
                f"<td data-sort-value='{pct_value:.6f}'>{pct}</td></tr>"
            )
        return (
            "<table class='mini-table'><thead><tr><th>类别</th><th data-sort-method='number'>人数</th><th data-sort-method='number'>占比</th></tr></thead>"
            f"<tbody>{''.join(rows)}</tbody></table>"
        )

    def counter_table(counter: Counter, title: str, top_n: int = 10) -> str:
        if not counter:
            return f"<p>{escape(title)}暂无数据。</p>"
        rows = []
        for value, count in counter.most_common(top_n):
            rows.append(
                f"<tr><td>{escape(value)}</td><td data-sort-value='{count}'>{count}</td></tr>"
            )
        return (
            f"<table class='mini-table'><thead><tr><th>{escape(title)}</th><th data-sort-method='number'>人数</th></tr></thead>"
            f"<tbody>{''.join(rows)}</tbody></table>"
        )

    html_rows: List[str] = []
    detail_map: Dict[str, List[Dict[str, Any]]] = {}
    for idx, entry in enumerate(action_rows):
        score = entry.get("priority_score", 0.0)
        bucket_label = entry.get("priority_bucket") or bucket_priority_score(score)[0]
        if bucket_label:
            summary_counts[bucket_label] += 1
        tags_text = entry.get("tags") or ""
        tags = [t.strip() for t in tags_text.split(',') if t.strip()]
        tag_counter.update(tags)

        phone_val = entry.get("phone") or ""
        name_val = entry.get("name") or ""
        last_order_val = entry.get("last_order") or ""
        platform_val = entry.get("platform") or ""
        if platform_val:
            platform_counter[platform_val] += 1
        key_val = entry.get("detail_key") or phone_val or f"{name_val}-{idx}"
        details = entry.get("details") or []
        if isinstance(details, list) and key_val:
            detail_map[str(key_val)] = [
                {
                    "姓名": str(d.get("姓名", "")),
                    "手机号": str(d.get("手机号", "") or phone_val),
                    "下单时间": str(d.get("下单时间", "")),
                    "下单平台": str(d.get("下单平台", "")),
                    "厂家": str(d.get("厂家", "")),
                    "货品名": str(d.get("货品名", "")),
                    "商品名称": str(d.get("商品名称", "")),
                    "颜色": str(d.get("颜色", "")),
                    "尺码": str(d.get("尺码", "")),
                    "付款金额": float(d.get("付款金额", 0.0)) if isinstance(d.get("付款金额"), (int, float)) else common_to_float(d.get("付款金额")),
                    "打款金额": float(d.get("打款金额", 0.0)) if isinstance(d.get("打款金额"), (int, float)) else common_to_float(d.get("打款金额")),
                    "负责人": str(d.get("负责人", "")),
                    "订单号": str(d.get("订单号", "")),
                    "退货单号": str(d.get("退货单号", "")),
                    "退款类型": str(d.get("退款类型", "")),
                    "退款原因": str(d.get("退款原因", "")),
                    "备注": str(d.get("备注", "")),
                    "数据来源": str(d.get("数据来源", "")),
                }
                for d in details
            ]
        priority_class = entry.get("priority_class", "priority-other")
        # Collect possible单号/退货单号 for search匹配（包含原值与仅数字形式，便于粘贴查询）
        try:
            ids_set = set()
            for x in (detail_map.get(str(key_val)) or []):
                for k in ("订单号", "退货单号"):
                    val = str(x.get(k, "")).strip()
                    if not val:
                        continue
                    ids_set.add(val)
                    digits = "".join(ch for ch in val if ch.isdigit())
                    if len(digits) >= 6:
                        ids_set.add(digits)
            ids_join = "|".join(sorted(ids_set))
        except Exception:
            ids_join = ""
        customer_value_val = entry.get("customer_value") or ""
        customer_list_val = entry.get("customer_list") or "活跃培养"  # 默认值
        row_attrs = (
            f"class='{priority_class}' "
            f"data-key='{escape(key_val, quote=True)}' "
            f"data-phone='{escape(phone_val, quote=True)}' "
            f"data-name='{escape(name_val, quote=True)}' "
            f"data-score='{escape(f'{score:.2f}', quote=True)}' "
            f"data-bucket='{escape(bucket_label, quote=True)}' "
            f"data-value='{escape(customer_value_val, quote=True)}' "
            f"data-list='{escape(customer_list_val, quote=True)}' "
            f"data-platform='{escape(platform_val, quote=True)}' "
            f"data-owner='{escape(str(entry.get('owner') or ''), quote=True)}' "
            f"data-last-order='{escape(last_order_val, quote=True)}' "
            f"data-cycle='{escape(str(entry.get('cycle') or ''), quote=True)}' "
            f"data-threshold='{escape(str(entry.get('threshold_display') or ''), quote=True)}' "
            f"data-category-cycle='{escape(str(entry.get('category_cycle') or ''), quote=True)}' "
            f"data-ids='{escape(ids_join.lower(), quote=True)}'"
        )
        cells_html: List[str] = []
        score_raw = entry.get("priority_score_raw")
        score_display = int(round(score if isinstance(score, (int, float)) else 0))
        for header in headers:
            if header == "标记完成":
                checkbox_id = f"followup-{idx}"
                cells_html.append(
                    f"<td data-header='{escape(header)}'>"
                    f"<input type='checkbox' class='followup-checkbox' id='{checkbox_id}' "
                    f"data-key='{escape(key_val, quote=True)}' "
                    f"data-phone='{escape(phone_val, quote=True)}' "
                    f"data-name='{escape(name_val, quote=True)}'>"
                    f"</td>"
                )
            else:
                sort_value: Optional[float] = None
                if header == "优先分":
                    if isinstance(score_raw, (int, float)):
                        sort_value = float(score_raw)
                        value = str(int(round(score_raw)))
                    elif isinstance(score, (int, float)):
                        sort_value = float(score)
                        value = str(score_display)
                    else:
                        sort_value = None
                        value = ""
                elif header == "姓名":
                    value = name_val
                elif header == "主要平台":
                    value = entry.get("platform") or ""
                elif header == "手机号":
                    value = phone_val
                elif header == "最近下单日":
                    value = last_order_val
                    if last_order_val:
                        digits = "".join(ch for ch in last_order_val if ch.isdigit())
                        if len(digits) >= 8:
                            try:
                                sort_value = float(digits[:8])
                            except Exception:
                                sort_value = None
                elif header == "风险标签":
                    value = tags_text
                elif header == "推荐动作":
                    value = entry.get("actions") or ""
                elif header == "促单理由":
                    value = entry.get("priority_explanation") or ""
                elif header == "价值层级":
                    value = entry.get("customer_value") or ""
                elif header == "偏好单品":
                    value = entry.get("favorite") or ""
                elif header == "有效订单数":
                    orders = entry.get("orders")
                    if isinstance(orders, (int, float)):
                        value = str(int(orders))
                        sort_value = float(orders)
                    else:
                        value = ""
                elif header == "退货率":
                    rate = entry.get("return_rate")
                    if isinstance(rate, (int, float)):
                        value = f"{rate * 100:.1f}%"
                        sort_value = float(rate)
                    else:
                        value = rate or ""
                elif header == "未复购天数":
                    days = entry.get("days")
                    if isinstance(days, (int, float)):
                        value = str(int(days))
                        sort_value = float(days)
                    else:
                        value = ""
                elif header == "复购周期(天)":
                    cycle = entry.get("cycle")
                    if isinstance(cycle, (int, float)):
                        value = f"{cycle:.1f}".rstrip("0").rstrip(".")
                        sort_value = float(cycle)
                    else:
                        value = ""
                elif header == "品类周期(天)":
                    cat_cycle = entry.get("category_cycle")
                    if isinstance(cat_cycle, (int, float)):
                        value = f"{cat_cycle:.1f}".rstrip("0").rstrip(".")
                        sort_value = float(cat_cycle)
                    else:
                        value = ""
                elif header == "预警阈值(天)":
                    threshold = entry.get("threshold_display")
                    if isinstance(threshold, (int, float)):
                        value = f"{threshold:.1f}".rstrip("0").rstrip(".")
                        sort_value = float(threshold)
                    else:
                        value = threshold or ""
                        try:
                            sort_value = float(threshold)
                        except (TypeError, ValueError):
                            sort_value = None
                elif header == "估算Uplift":
                    uplift_val = entry.get("estimated_uplift")
                    if isinstance(uplift_val, (int, float)):
                        value = f"{uplift_val:.2f}"
                        sort_value = float(uplift_val)
                    else:
                        value = ""
                elif header == "估算毛利":
                    margin_val = entry.get("estimated_margin")
                    if isinstance(margin_val, (int, float)):
                        value = f"{margin_val:.2f}"
                        sort_value = float(margin_val)
                    else:
                        value = ""
                elif header == "估算退货率":
                    est_return = entry.get("estimated_return_rate")
                    if isinstance(est_return, (int, float)):
                        value = f"{est_return * 100:.1f}%"
                        sort_value = float(est_return)
                    else:
                        value = ""
                elif header == "触达成本":
                    cost_val = entry.get("touch_cost")
                    if isinstance(cost_val, (int, float)):
                        value = f"{cost_val:.2f}"
                        sort_value = float(cost_val)
                    else:
                        value = ""
                elif header == "AOV趋势":
                    trend = entry.get("aov_trend")
                    if isinstance(trend, (int, float)):
                        value = f"{trend:.2f}x"
                        sort_value = float(trend)
                    else:
                        value = ""
                elif header == "平均客单价":
                    aov = entry.get("aov")
                    if isinstance(aov, (int, float)):
                        value = f"{aov:.2f}"
                        sort_value = float(aov)
                    else:
                        value = ""
                else:
                    value = ""
                sort_key = to_sort_key(sort_value)
                sort_attr = (
                    f" data-sort-value='{escape(sort_key, quote=True)}'" if sort_key is not None else ""
                )
                cells_html.append(
                    f"<td data-header='{escape(header)}'{sort_attr}>{escape(value)}</td>"
                )
        html_rows.append(f"<tr {row_attrs}>{''.join(cells_html)}</tr>")
    
    # 统计各列表人数
    list_counts = Counter()
    for entry in action_rows:
        list_name = entry.get("customer_list") or "活跃培养"
        list_counts[list_name] += 1

    filters_html = f"""
    <div class="list-tabs" style="margin-bottom: 20px; border-bottom: 2px solid #e9ecf3;">
        <button class="list-tab active" data-list="全部" style="padding: 10px 20px; border: none; background: none; cursor: pointer; font-size: 15px; font-weight: 600; border-bottom: 3px solid transparent; transition: all 0.2s;">全部 ({len(action_rows)})</button>
        <button class="list-tab" data-list="超级VIP" style="padding: 10px 20px; border: none; background: none; cursor: pointer; font-size: 15px; font-weight: 600; border-bottom: 3px solid transparent; transition: all 0.2s;">⭐ 超级VIP ({list_counts.get("超级VIP", 0)})</button>
        <button class="list-tab" data-list="活跃培养" style="padding: 10px 20px; border: none; background: none; cursor: pointer; font-size: 15px; font-weight: 600; border-bottom: 3px solid transparent; transition: all 0.2s;">💡 活跃培养 ({list_counts.get("活跃培养", 0)})</button>
        <button class="list-tab" data-list="濒临流失" style="padding: 10px 20px; border: none; background: none; cursor: pointer; font-size: 15px; font-weight: 600; border-bottom: 3px solid transparent; transition: all 0.2s;">🔔 濒临流失 ({list_counts.get("濒临流失", 0)})</button>
        <button class="list-tab" data-list="高风险待排查" style="padding: 10px 20px; border: none; background: none; cursor: pointer; font-size: 15px; font-weight: 600; border-bottom: 3px solid transparent; transition: all 0.2s;">⚠️ 高风险待排查 ({list_counts.get("高风险待排查", 0)})</button>
        <button class="list-tab" data-list="冷却期" style="padding: 10px 20px; border: none; background: none; cursor: pointer; font-size: 15px; font-weight: 600; border-bottom: 3px solid transparent; transition: all 0.2s;">🕐 冷却期客户 ({list_counts.get("冷却期", 0)})</button>
    </div>
    <div class="filters">
        <label>关键词搜索：<input id="searchBox" type="search" placeholder="姓名 / 电话 / 标签 / 平台 / 单号 / 退货单号..."></label>
        <label>优先分：
            <select id="priorityFilter">
                <option value="">全部</option>
                <option value="高(≥80)">高(≥80)</option>
                <option value="中(50-79)">中(50-79)</option>
                <option value="低(0-49)">低(0-49)</option>
                <option value="负分">负分</option>
            </select>
        </label>
        <label>标签：
            <select id="tagFilter">
                <option value="">全部</option>
            </select>
        </label>
        <label>平台：
            <select id="platformFilter">
                <option value="">全部</option>
            </select>
        </label>
        <label>厂家：
            <select id="manufacturerFilter">
                <option value="">全部</option>
            </select>
        </label>
        <span class="counter" id="rowCounter"></span>
    </div>
    """

    tags_js = json.dumps([tag for tag, _ in tag_counter.most_common()], ensure_ascii=False, separators=(',', ':'))
    platforms_js = json.dumps(sorted(platform_counter.keys()), ensure_ascii=False, separators=(',', ':'))
    details_js = json.dumps(detail_map, ensure_ascii=False, separators=(',', ':'))
    # Global details for all客户（用于全库单号查询）- 使用压缩JSON格式减小文件大小
    global_details = global_details or {}
    global_details_js = json.dumps(global_details, ensure_ascii=False, separators=(',', ':'))
    # 货品名搜索索引（规范化货品名 -> 手机号列表）
    product_search_index = product_search_index or {}
    product_search_index_js = json.dumps(product_search_index, ensure_ascii=False, separators=(',', ':'))
    # 冷却名单（近N天 Excel/飞书 联系记录内的手机号集合）
    cooldown_keys = cooldown_keys or []
    cooldown_total = len(cooldown_keys)
    cooldown_keys_js = json.dumps(cooldown_keys, ensure_ascii=False, separators=(',', ':'))
    # 冷却期客户联系日期映射
    cooldown_customers = cooldown_customers or {}
    cooldown_customers_js = json.dumps({k: v.isoformat() if isinstance(v, date) else str(v) for k, v in cooldown_customers.items()}, ensure_ascii=False, separators=(',', ':'))
    # 跟进人建议与默认（来自环境变量）
    owners_raw = os.getenv('FOLLOWUP_OWNERS') or os.getenv('FEISHU_OWNER_LIST') or ''
    owner_suggestions = [s.strip() for s in owners_raw.split(',') if s and s.strip()]
    owner_suggestions_js = json.dumps(owner_suggestions, ensure_ascii=False, separators=(',', ':'))
    env_default_owner = os.getenv('DEFAULT_FOLLOWUP_OWNER') or os.getenv('FEISHU_CURRENT_USER') or os.getenv('FOLLOWUP_OWNER') or ''
    env_default_owner_js = json.dumps(env_default_owner, ensure_ascii=False, separators=(',', ':'))

    # 顶部冷却提示 HTML 片段，已调整：冷却期客户现在作为独立标签展示，不再隐藏
    if contact_log_used:
        meta_cooldown_html = (
            f'<div>冷却期过滤：近 {max(0, cooldown_days)} 天已联系客户已归入"冷却期客户"标签（'
            f'共 <span id="cooldownTotalSpan">{cooldown_total}</span> 人）'
            f'</div>'
        )
    elif max(0, cooldown_days) > 0:
        meta_cooldown_html = (
            f'<div>冷却期过滤：近 {max(0, cooldown_days)} 天已联系客户（未启用联系记录数据源）'
            f'</div>'
        )
    else:
        meta_cooldown_html = ''
    # Global meta for all customers（供临时行填充真实优先分/标签等）
    global_meta = global_meta or {}
    global_meta_js = json.dumps(global_meta, ensure_ascii=False, separators=(',', ':'))
    # AI Analysis Map (Empty default for now to prevent ReferenceError)
    ai_analysis_map_js = json.dumps({}, ensure_ascii=False, separators=(',', ':'))
    # 构建"加推SKU"候选（近45天 订单>2 且 退货率<30%）
    def build_sku_push_table(all_details: Dict[str, List[Dict[str, Any]]], today: date,
                             days_window: int = 45, min_orders: int = 3, max_return_rate: float = 0.30,
                             top_n: int = 50) -> str:
        if not all_details:
            return ""
        cutoff = today - timedelta(days=max(1, days_window))
        sku_stats: Dict[str, Dict[str, float]] = {}
        
        for _key, rows in all_details.items():
            if not isinstance(rows, list):
                continue
            for r in rows:
                try:
                    item = str(r.get("货品名", "")).strip()
                    if not item:
                        continue
                    # 排除代发和样品
                    platform = str(r.get("下单平台", "") or "").strip()
                    prod = str(r.get("商品名称", "") or "")
                    rmk = str(r.get("备注", "") or "")
                    combined = f"{platform} {prod} {item} {rmk}"
                    if "代发" in combined or "样品" in combined:
                        continue
                    # global_details中的字段名已经被标准化为"下单时间"
                    raw_date = r.get("下单时间")
                    od = common_parse_excel_date(raw_date, today)
                    if not isinstance(od, date) or od < cutoff:
                        continue
                    refund_type = str(r.get("退款类型", "") or "").strip()
                    return_no = str(r.get("退货单号", "") or "").strip()
                    order_no = str(r.get("订单号", "") or "").strip()
                    # global_details中的字段名已经被标准化为“付款金额”
                    pay_amount = r.get("付款金额", 0.0)
                    try:
                        pay_amount = float(pay_amount)
                    except Exception:
                        pay_amount = common_to_float(pay_amount)
                    is_cancel = ("取消" in refund_type) or ("取消" in order_no)
                    # 判定退货：退款类型中含“退/退货/退款”或存在退货单号，且非取消
                    rt_lower = refund_type.lower()
                    is_return = (("退" in refund_type) or ("退货" in refund_type) or ("退款" in refund_type) or (return_no and return_no not in {"/", "-"})) and not is_cancel
                    # 计入有效订单：金额>0 且 非取消（退货行通常金额为0，不计入订单数）
                    is_valid_order = (pay_amount or 0.0) > 0 and not is_cancel
                    stat = sku_stats.get(item)
                    if stat is None:
                        stat = {"orders": 0.0, "returns": 0.0, "revenue": 0.0}
                        sku_stats[item] = stat
                    if is_valid_order:
                        stat["orders"] += 1
                        stat["revenue"] += max(0.0, pay_amount)
                    if is_return:
                        stat["returns"] += 1
                except Exception:
                    continue
        
        # 构建候选集
        rows: List[Tuple[str, int, float, float]] = []  # (sku, orders, return_rate, revenue)
        for sku, stat in sku_stats.items():
            orders = int(stat.get("orders", 0) or 0)
            returns = int(stat.get("returns", 0) or 0)
            if orders <= 0:
                continue
            # 退货率控制在[0,1]
            rr = 0.0 if orders <= 0 else min(1.0, max(0.0, returns / max(1, orders)))
            if orders >= max(1, min_orders) and rr < max_return_rate:
                rows.append((sku, orders, rr, float(stat.get("revenue", 0.0) or 0.0)))
        if not rows:
            # 即使没有数据也显示卡片和提示
            return (
                "<div class=\"card\"><h3>加推SKU（近45天 订单>2 且 退货率<30%） "
                "<span class=\"text-xs font-normal text-slate-400 ml-2\">(0款)</span></h3>"
                "<div class=\"scroll-pane border border-slate-200 rounded-lg\"><div class=\"text-sm text-slate-400 text-center py-8\">近45天内暂无符合条件的SKU</div></div></div>"
            )
        # 排序：订单数降序，其次退货率升序，其次销售额降序
        rows.sort(key=lambda x: (-x[1], x[2], -x[3]))
        total_count = len(rows)
        rows = rows[:top_n]
        # 渲染表格
        body = []
        for sku, orders, rr, rev in rows:
            body.append(
                f"<tr data-sku=\"{escape(sku)}\">"
                f"<td class=\"font-medium text-slate-700\">{escape(sku)}</td>"
                f"<td data-sort-value='{orders}'>{orders}</td>"
                f"<td data-sort-value='{rr:.6f}'>{rr*100:.1f}%</td>"
                f"<td data-sort-value='{rev:.2f}'>¥{rev:.2f}</td>"
                f"</tr>"
            )
        # 构建完整表格
        tbl = (
            "<table class='mini-table w-full' id='skuPushTableStatic' style='display:table !important;'><thead><tr>"
            "<th>货品名</th><th data-sort-method='number'>订单数</th>"
            "<th data-sort-method='number'>退货率</th><th data-sort-method='number'>销售额</th>"
            "</tr></thead>"
            f"<tbody>{''.join(body)}</tbody></table>"
        )
        # 返回卡片内容（不包含 summary 包裹层）
        return (
            "<div class=\"card\"><h3>加推SKU（近45天 订单>2 且 退货率<30%） "
            f"<span class=\"text-xs font-normal text-slate-400 ml-2\">(共{total_count}款)</span></h3>"
            f"<div id=\"skuPushContainer\" class=\"scroll-pane border border-slate-200 rounded-lg\" style=\"min-height: 50px;\">{tbl}</div></div>"
        )

    sku_push_html = build_sku_push_table(global_details, today)
    def build_high_return_placeholder() -> str:
        return (
            "<div class=\"card\">"
            "<h3>高退货预警（明细>3，退货率>30%）</h3>"
            "<div id=\"skuReturnAlertTable\" class=\"scroll-pane border border-slate-200 rounded-lg\"></div>"
            "</div>"
        )
    sku_return_html = build_high_return_placeholder()

    sku_stats_margin = {}  # 非代发
    sku_stats_margin_proxy = {}  # 代发
    for _key, rows in (global_details or {}).items():
        if not isinstance(rows, list):
            continue
        for r in rows:
            try:
                item = str(r.get("货品名", "") or "").strip()
                if not item:
                    continue
                prod = str(r.get("商品名称", "") or "")
                rmk = str(r.get("备注", "") or "")
                platform = str(r.get("下单平台", "") or "")
                combined = f"{prod} {item} {rmk} {platform}"
                
                # 检查是否为样品，样品直接跳过
                if "样品" in combined:
                    continue
                
                # 判断是否为代发
                is_proxy = "代发" in combined
                
                # 过滤退货订单：检查状态和退款类型
                refund_type = str(r.get("退款类型", "") or "").strip()
                return_no = str(r.get("退货单号", "") or "").strip()
                order_no = str(r.get("订单号", "") or "").strip()
                is_cancel = "取消" in refund_type or "取消" in order_no
                is_return = (("退" in refund_type) or ("退货" in refund_type) or ("退款" in refund_type) or (return_no and return_no.lower() not in {"/", "-", "", "none", "null"})) and not is_cancel
                
                # 只统计有效订单，过滤退货和取消
                if is_return or is_cancel:
                    continue
                
                # global_details中的字段名已经被标准化
                rev_val = common_to_float(r.get("付款金额"))
                pay_val = common_to_float(r.get("打款金额"))
                # 放宽筛选条件：只要有收款即可，打款为0的订单也记录
                if rev_val <= 0:
                    continue
                
                # 获取下单日期
                order_date_raw = r.get("下单时间")
                order_date = common_parse_excel_date(order_date_raw, today)
                
                # 根据是否为代发分别统计
                target_dict = sku_stats_margin_proxy if is_proxy else sku_stats_margin
                stat = target_dict.get(item)
                if stat is None:
                    stat = {"rev": 0.0, "cost": 0.0, "cnt": 0, "last_date": None}
                    target_dict[item] = stat
                stat["rev"] += rev_val
                stat["cost"] += pay_val
                stat["cnt"] += 1
                # 记录最后一单日期
                if order_date:
                    if stat["last_date"] is None or order_date > stat["last_date"]:
                        stat["last_date"] = order_date
            except Exception:
                continue
    # 生成两个alert_list：非代发和代发
    def build_alert_list(stats_dict):
        result = []
        for sku, st in stats_dict.items():
            if st["cnt"] <= 1:
                continue
            avg_r = st["rev"] / st["cnt"]
            avg_c = st["cost"] / st["cnt"]
            if avg_r <= 0:
                continue
            mrg = (avg_r - avg_c) / avg_r
            if mrg < 0.35:
                last_date = st.get("last_date")
                last_date_str = last_date.strftime("%Y-%m-%d") if isinstance(last_date, date) else ""
                result.append((sku, st["cnt"], mrg, last_date_str, last_date))
        result.sort(key=lambda x: x[2])  # 按毛利率排序
        return result
    
    alert_list = build_alert_list(sku_stats_margin)  # 非代发
    alert_list_proxy = build_alert_list(sku_stats_margin_proxy)  # 代发
    
    if alert_list or alert_list_proxy:
        # 构建非代发表格
        tbody_parts = []
        for sku, cnt, mr, last_date_str, last_date_obj in alert_list:
            # clr = "#ff4d4f" if mr < 0.2 else "#faad14"
            color_class = "text-red-500" if mr < 0.2 else "text-amber-500"
            # 用日期对象作为排序值，如果有的话
            date_sort_value = last_date_obj.strftime("%Y%m%d") if last_date_obj else "00000000"
            tbody_parts.append(
                f"<tr data-sku=\"{escape(sku)}\" data-type='normal'><td class=\"font-medium text-slate-700\">{escape(sku)}</td>"
                f"<td data-sort-value='{cnt}'>{cnt}</td>"
                f"<td data-sort-value='{mr:.6f}' class='{color_class}'>{mr*100:.1f}%</td>"
                f"<td data-sort-value='{date_sort_value}'>{escape(last_date_str) if last_date_str else '-'}</td></tr>"
            )
        
        # 构建代发表格
        tbody_parts_proxy = []
        for sku, cnt, mr, last_date_str, last_date_obj in alert_list_proxy:
            # clr = "#ff4d4f" if mr < 0.2 else "#faad14"
            color_class = "text-red-500" if mr < 0.2 else "text-amber-500"
            date_sort_value = last_date_obj.strftime("%Y%m%d") if last_date_obj else "00000000"
            tbody_parts_proxy.append(
                f"<tr data-sku=\"{escape(sku)}\" data-type='proxy'><td class=\"font-medium text-slate-700\">{escape(sku)}</td>"
                f"<td data-sort-value='{cnt}'>{cnt}</td>"
                f"<td data-sort-value='{mr:.6f}' class='{color_class}'>{mr*100:.1f}%</td>"
                f"<td data-sort-value='{date_sort_value}'>{escape(last_date_str) if last_date_str else '-'}</td></tr>"
            )
        
        # 合并两个表格的tbody
        all_tbody = ''.join(tbody_parts) + ''.join(tbody_parts_proxy)
        
        tbl = (
            "<table class='mini-table w-full' id='lowProfitTable'><thead><tr>"
            "<th>货品名</th><th data-sort-method='number'>订单数</th>"
            "<th data-sort-method='number'>毛利率</th><th>末单日期</th></tr></thead>"
            f"<tbody>{all_tbody}</tbody></table>"
        )
        
        # 添加过滤器控制
        filter_html = (
            "<div class=\"flex items-center mb-3\">"
            "<label class=\"flex items-center mr-4 text-sm text-slate-600 cursor-pointer\">"
            "<input type='radio' name='lowProfitFilter' value='normal' checked class=\"mr-1.5 text-brand-600 focus:ring-brand-500\"/>不含代发"
            "</label>"
            "<label class=\"flex items-center text-sm text-slate-600 cursor-pointer\">"
            "<input type='radio' name='lowProfitFilter' value='proxy' class=\"mr-1.5 text-brand-600 focus:ring-brand-500\"/>仅代发"
            "</label>"
            f"<span class=\"ml-3 text-xs text-slate-400\" id='lowProfitCount'>(共{len(alert_list)}款)</span>"
            "</div>"
        )
        
        low_margin_html = (
            "<div class=\"card\"><h3>低毛利预警（毛利率<35%，明细>1）</h3>"
            f"{filter_html}"
            f"<div class=\"scroll-pane border border-slate-200 rounded-lg\">{tbl}</div></div>"
        )
    else:
        low_margin_html = ""
    # 构建全局单号索引（小写与纯数字形式均可命中）
    id_index: Dict[str, str] = {}
    # 新增：姓名索引（小写精确匹配 -> key），便于全库按姓名命中
    name_index: Dict[str, str] = {}
    for gkey, drows in global_details.items():
        if not isinstance(drows, list):
            continue
        for d in drows:
            if not isinstance(d, dict):
                continue
            for kk in ("订单号", "退货单号"):
                val = str(d.get(kk, "")).strip()
                if not val:
                    continue
                id_index[val.lower()] = gkey
                digits = "".join(ch for ch in val if ch.isdigit())
                if len(digits) >= 6:
                    id_index[digits] = gkey
        # 采集姓名索引（取第一条非空姓名）
        try:
            name_candidates = [str(x.get("姓名", "")).strip() for x in drows if isinstance(x, dict)]
            primary_name = next((n for n in name_candidates if n), "")
            if primary_name:
                lower_name = primary_name.lower()
                # 若同名多 key，仅保留先出现的一个，避免覆盖
                if lower_name not in name_index:
                    name_index[lower_name] = gkey
        except Exception:
            pass
    id_index_js = json.dumps(id_index, ensure_ascii=False, separators=(',', ':'))
    name_index_js = json.dumps(name_index, ensure_ascii=False, separators=(',', ':'))
    deepseek_key_js = json.dumps(config.deepseek_key or "", ensure_ascii=False)

    contact_server_port = int(os.getenv('CONTACT_SERVER_PORT') or '5005')
    contact_write_enabled = str(os.getenv('CONTACT_SERVER') or '0').strip().lower() in ('1','true','yes','on')
    html_template = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>客户触达仪表盘 - {today.isoformat()}</title>
    <style>
        body {{ font-family: 'Noto Sans SC', 'IBM Plex Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; margin: 24px; background: #f6f7fb; color: #333; }}
        h1 {{ margin-bottom: 8px; }}
        .meta {{ color: #666; margin-bottom: 16px; }}
        .filters {{ display: flex; flex-wrap: wrap; gap: 12px; align-items: center; margin-bottom: 16px; }}
        .filters label {{ font-size: 14px; }}
        .filters input, .filters select {{ padding: 4px 8px; font-size: 14px; }}
        .counter {{ font-weight: 600; }}
        .toolbar {{ display: flex; flex-wrap: wrap; gap: 12px; align-items: center; margin-bottom: 16px; }}
        .toolbar button {{ padding: 6px 14px; border: 1px solid #ccd5e3; background: #fff; border-radius: 6px; cursor: pointer; font-size: 14px; transition: background 0.2s; }}
        .toolbar button:hover {{ background: #f2f6ff; }}
        .toolbar span {{ font-size: 14px; color: #555; }}
        table {{ width: 100%; border-collapse: collapse; background: #fff; box-shadow: 0 2px 6px rgba(0,0,0,0.05); }}
        thead {{ background: #f0f2f5; }}
        th, td {{ padding: 8px 10px; border-bottom: 1px solid #e9ecf3; text-align: left; font-size: 14px; }}
        tr:hover {{ background: #f8fbff; }}
        .priority-high {{ border-left: 4px solid #ff4d4f; }}
        .priority-mid {{ border-left: 4px solid #faad14; }}
        .priority-low {{ border-left: 4px solid #1890ff; }}
        .priority-other {{ border-left: 4px solid #d9d9d9; }}
        tr.completed {{ background: #eefbf1 !important; }}
        tr.completed td {{ color: #6c6c6c; }}
        .followup-checkbox {{ width: 16px; height: 16px; }}
        .summary {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(max(320px, calc((100% - 32px) / 3)), 1fr)); gap: 16px; margin-bottom: 24px; }}
        @media (max-width: 1200px) {{ .summary {{ grid-template-columns: repeat(auto-fill, minmax(max(280px, calc((100% - 16px) / 2)), 1fr)); }} }}
        @media (max-width: 768px) {{ .summary {{ grid-template-columns: 1fr; }} }}
        .card {{ background: #fff; padding: 16px; border-radius: 8px; box-shadow: 0 2px 6px rgba(0,0,0,0.05); }}
        .mini-table {{ width: 100%; border-collapse: separate; border-spacing: 0; margin-top: 12px; font-size: 13px; color: #475569; }}
        .mini-table th {{ background: #f8fafc; font-weight: 600; color: #64748b; padding: 8px 12px; text-align: left; border-bottom: 1px solid #e2e8f0; white-space: nowrap; }}
        .mini-table td {{ padding: 8px 12px; border-bottom: 1px solid #f1f5f9; vertical-align: middle; }}
        .mini-table tr:last-child td {{ border-bottom: none; }}
        .mini-table tr:hover {{ background: #f8fafc; transition: background 0.1s; }}
        .mini-table td.num {{ font-variant-numeric: tabular-nums; }}
        .scroll-pane {{ max-height: 400px; overflow-y: auto; }}
        .sku-nav {{ display:flex; gap:8px; align-items:center; justify-content:flex-end; margin-bottom:6px; }}
        .sku-nav button {{ padding: 2px 8px; border: 1px solid #ccd5e3; background: #fff; border-radius: 6px; cursor: pointer; font-size: 12px; }}
        .sku-nav button:hover {{ background: #f2f6ff; }}
        th {{ cursor: pointer; user-select: none; }}
        th[data-sort-order="asc"]::after {{ content: " ▲"; color: #999; font-size: 12px; }}
        th[data-sort-order="desc"]::after {{ content: " ▼"; color: #999; font-size: 12px; }}
        .footer {{ margin-top: 32px; color: #888; font-size: 12px; }}
        /* Role Selector */
        .role-selector {{
            display: flex;
            gap: 12px;
            margin-bottom: 24px;
            padding: 16px;
            background:
                radial-gradient(circle at 0 0, #0f172a 0, #1e293b 40%),
                linear-gradient(135deg, #0ea5e9 0%, #22c55e 100%);
            border-radius: 12px;
            box-shadow: 0 4px 12px rgba(15, 23, 42, 0.35);
        }}
        .role-option {{
            flex: 1;
            position: relative;
        }}
        .role-option input[type="radio"] {{
            display: none;
        }}
        .role-option label {{
            display: block;
            padding: 12px 24px;
            text-align: center;
            background: rgba(255, 255, 255, 0.15);
            color: rgba(255, 255, 255, 0.85);
            border-radius: 8px;
            cursor: pointer;
            font-size: 16px;
            font-weight: 600;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            border: 2px solid transparent;
        }}
        .role-option label:hover {{
            background: rgba(255, 255, 255, 0.25);
            transform: translateY(-2px);
        }}
        .role-option input[type="radio"]:checked + label {{
            background: #fff;
            color: #667eea;
            border-color: #fff;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
        }}
        .role-option label::before {{
            content: '';
            display: inline-block;
            margin-right: 8px;
            font-size: 18px;
        }}
        .role-option.operations label::before {{
            content: '📊';
        }}
        .role-option.customer-service label::before {{
            content: '👥';
        }}
        /* Operations view search box styling */
        .operations-search-container {{
            display: none;
            max-width: 600px;
            margin: 0 auto 32px;
            padding: 20px;
            background: #fff;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }}
        .operations-search-container.active {{
            display: block;
        }}
        .operations-search-box {{
            width: 100%;
            padding: 12px 20px;
            font-size: 16px;
            border: 2px solid #e9ecf3;
            border-radius: 8px;
            transition: all 0.3s;
        }}
        .operations-search-box:focus {{
            outline: none;
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
        }}
        .operations-search-box::placeholder {{
            color: #999;
        }}
        /* Drilldown drawer */
        .hidden {{ display: none; }}
        .detail-backdrop {{ position: fixed; inset: 0; background: rgba(0,0,0,0.35); z-index: 999; }}
        .detail-panel {{ position: fixed; left: 0; right: 0; bottom: 0; height: 80vh; background: #fff; border-top-left-radius: 12px; border-top-right-radius: 12px; box-shadow: 0 -4px 14px rgba(0,0,0,0.08); z-index: 1000; display: flex; flex-direction: column; }}
        .detail-panel.hidden, .detail-backdrop.hidden {{ display: none !important; }}
        .detail-header {{ display: flex; align-items: center; justify-content: space-between; padding: 10px 14px; border-bottom: 1px solid #e9ecf3; }}
        .detail-header h3 {{ margin: 0; font-size: 16px; }}
        .detail-close {{ border: 1px solid #ccd5e3; background: #fff; border-radius: 6px; padding: 4px 10px; cursor: pointer; }}
        .detail-body {{ flex: 1; overflow: auto; padding: 12px; }}
        #detailTable {{ width: 100%; border-collapse: collapse; }}
        #detailTable th, #detailTable td {{ padding: 6px 8px; border-bottom: 1px solid #e9ecf3; font-size: 13px; }}
        
        /* AI Chatbox */
        .ai-chat-btn {{
            position: fixed;
            bottom: 24px;
            right: 24px;
            width: 56px;
            height: 56px;
            border-radius: 28px;
            background: linear-gradient(135deg, #6366f1 0%, #a855f7 100%);
            box-shadow: 0 4px 14px rgba(99, 102, 241, 0.4);
            color: #fff;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 24px;
            cursor: pointer;
            z-index: 2000;
            transition: transform 0.2s;
        }}
        .ai-chat-btn:hover {{ transform: scale(1.05); }}
        
        .ai-chat-panel {{
            position: fixed;
            bottom: 90px;
            right: 24px;
            width: 360px;
            height: 500px;
            background: #fff;
            border-radius: 12px;
            box-shadow: 0 8px 30px rgba(0,0,0,0.12);
            z-index: 2000;
            display: flex;
            flex-direction: column;
            border: 1px solid #e2e8f0;
        }}
        .ai-chat-panel.hidden {{ display: none; }}
        .ai-chat-header {{
            padding: 12px 16px;
            background: #f8fafc;
            border-bottom: 1px solid #e2e8f0;
            border-radius: 12px 12px 0 0;
            display: flex;
            justify-content: space-between;
            align-items: center;
            font-weight: 600;
            color: #334155;
        }}
        .ai-chat-body {{
            flex: 1;
            padding: 16px;
            overflow-y: auto;
            display: flex;
            flex-direction: column;
            gap: 12px;
            font-size: 13px;
        }}
        .chat-msg {{
            padding: 8px 12px;
            border-radius: 8px;
            max-width: 85%;
            line-height: 1.5;
        }}
        .chat-msg.user {{
            align-self: flex-end;
            background: #eff6ff;
            color: #1e40af;
            border: 1px solid #dbeafe;
        }}
        .chat-msg.ai {{
            align-self: flex-start;
            background: #f1f5f9;
            color: #334155;
            white-space: pre-wrap;
            font-family: monospace;
        }}
        .ai-chat-input-area {{
            padding: 12px;
            border-top: 1px solid #e2e8f0;
            display: flex;
            gap: 8px;
        }}
        .ai-chat-input {{
            flex: 1;
            padding: 8px;
            border: 1px solid #cbd5e1;
            border-radius: 6px;
            font-size: 13px;
            outline: none;
        }}
        .ai-chat-input:focus {{ border-color: #6366f1; }}
        .ai-chat-send {{
            padding: 6px 12px;
            background: #6366f1;
            color: #fff;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 13px;
        }}
        .ai-chat-send:disabled {{ background: #cbd5e1; cursor: not-allowed; }}
    </style>
    <script>
    (function () {{
        function getSortValue(cell, method) {{
            if (!cell) {{
                return method === 'number' ? -Infinity : '';
            }}
            const attr = cell.getAttribute('data-sort-value');
            if (method === 'number') {{
                const raw = attr !== null ? attr : cell.textContent || '';
                const parsed = parseFloat(raw);
                return Number.isFinite(parsed) ? parsed : -Infinity;
            }}
            return (attr !== null ? attr : cell.textContent || '').trim().toLowerCase();
        }}

        function Tablesort(table) {{
            if (!table) {{
                throw new Error('Tablesort requires a table element.');
            }}
            this.table = table;
            this.thead = table.querySelector('thead');
            this.headers = this.thead ? Array.from(this.thead.querySelectorAll('th')) : [];
            this.tbody = table.tBodies[0];
            this.init();
        }}

        Tablesort.prototype.init = function () {{
            const self = this;
            this.headers.forEach(function (header, index) {{
                if (header.classList.contains('no-sort')) {{
                    return;
                }}
                header.addEventListener('click', function () {{
                    self.sortBy(index, header);
                }});
            }});
        }};

        Tablesort.prototype.sortBy = function (index, header) {{
            if (!this.tbody) {{
                return;
            }}
            const method = header.getAttribute('data-sort-method') || undefined;
            // 只对可见行进行排序（排除display:none的行）
            const allRows = Array.from(this.tbody.rows);
            const visibleRows = allRows.filter(row => row.style.display !== 'none');
            const hiddenRows = allRows.filter(row => row.style.display === 'none');
            const currentOrder = header.getAttribute('data-sort-order') || 'asc';
            const newOrder = currentOrder === 'asc' ? 'desc' : 'asc';

            this.headers.forEach(function (head) {{
                if (head !== header) {{
                    head.removeAttribute('data-sort-order');
                }}
            }});

            visibleRows.sort(function (a, b) {{
                const aVal = getSortValue(a.cells[index], method);
                const bVal = getSortValue(b.cells[index], method);
                if (method === 'number') {{
                    return aVal - bVal;
                }}
                if (aVal === bVal) {{
                    return 0;
                }}
                return aVal > bVal ? 1 : -1;
            }});

            if (newOrder === 'desc') {{
                visibleRows.reverse();
            }}

            // 先添加可见行，再添加隐藏行（保持隐藏行在最后）
            const fragment = document.createDocumentFragment();
            visibleRows.forEach(function (row) {{
                fragment.appendChild(row);
            }});
            hiddenRows.forEach(function (row) {{
                fragment.appendChild(row);
            }});
            this.tbody.appendChild(fragment);
            header.setAttribute('data-sort-order', newOrder);
        }};

        window.Tablesort = Tablesort;
    }})();
    </script>
</head>
<body>
    <h1>客户触达仪表盘</h1>

    <!-- 角色选择器 -->
    <div class="role-selector">
        <div class="role-option operations">
            <input type="radio" id="roleOperations" name="userRole" value="operations">
            <label for="roleOperations">运营视角</label>
        </div>
        <div class="role-option customer-service">
            <input type="radio" id="roleCustomerService" name="userRole" value="customer-service" checked>
            <label for="roleCustomerService">客服视角</label>
        </div>
    </div>

    <!-- 运营视角专用搜索框 -->
    <div class="operations-search-container">
        <input type="search"
               class="operations-search-box"
               id="operationsSearchBox"
               placeholder="🔍 搜索 SKU 名称、订单号、退货单号...">
    </div>

    <div class="summary" id="skuSummary" style="display:none;">
        {sku_push_html}
        {sku_return_html}
        {low_margin_html}
        <div class="card" style="grid-column: 1 / -1;">
            <div style="display:flex; justify-content:space-between; align-items:center;">
                <h3>厂家分析（全库历史表现）</h3>
                <select id="manufacturerTimeSelect" style="font-size:12px; padding:4px 8px; border-radius:6px; border:1px solid #e2e8f0; margin-right:12px;">
                    <option value="all">全部时间</option>
                    <option value="30">近30天</option>
                    <option value="60" selected>近60天</option>
                    <option value="90">近90天</option>
                </select>
            </div>
            <p style="margin: 6px 0 10px 0; font-size: 12px; color: #64748b;">
                依据全库订单按厂家聚合，展示 Top 厂家的订单量、退货率与客单价概览，用于判断供应侧风险与潜力。
            </p>
            <div class="scroll-pane" style="max-height:none; border:none;">
                <div style="display:block;">
                    <!-- Chart removed as per request -->
                    <div style="display:none;">
                        <div id="manufacturerChart"></div>
                        <select id="manufacturerMetricSelect"></select>
                    </div>
                    
                    <div>
                        <table class="mini-table" style="width:100%;">
                            <thead>
                                <tr>
                                    <th>厂家</th>
                                    <th data-sort-method="number">订单数</th>
                                    <th>趋势</th>
                                    <th data-sort-method="number">总销售额</th>
                                    <th data-sort-method="number">退货率</th>
                                    <th data-sort-method="number">毛利率</th>
                                    <th>风险标签</th>
                                </tr>
                            </thead>
                            <tbody id="manufacturerSummaryBody">
                                <!-- 将由前端脚本注入 Top 厂家明细 -->
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <div class="toolbar">
        <button id="exportCsv">导出今日联系记录 (CSV)</button>
        <button id="clearMarks">清除今日标记</button>
        <label style="margin-left:12px;">跟进人（选择）：
            <select id="defaultOwnerSelect" style="width:140px;">
                <option value="">选择</option>
                <option value="黄蓉">黄蓉</option>
                <option value="云云">云云</option>
                <option value="铭慧">铭慧</option>
            </select>
        </label>
        <span id="followupSummary"></span>
        <span id="globalHitBox" style="margin-left:12px;color:#444"></span>
    </div>

    {filters_html}

    <table id="actionTable">
        <thead>
            <tr>{"".join(header_cells)}</tr>
        </thead>
        <tbody>
            {"".join(html_rows)}
        </tbody>
    </table>

    <div class="footer">
        自动生成脚本：generate_customer_alerts.py | 数据来源：账单汇总
    </div>

    <script>
    const searchBox = document.getElementById('searchBox');
    const priorityFilter = document.getElementById('priorityFilter');
    const tagFilter = document.getElementById('tagFilter');
    const platformFilter = document.getElementById('platformFilter');
    const manufacturerFilter = document.getElementById('manufacturerFilter');
    const table = document.getElementById('actionTable');
    const rows = table ? Array.from(table.querySelectorAll('tbody tr')) : [];
    // 启用全局表头点击排序（主表 + 摘要小表）
    try {{
        const allTables = Array.from(document.querySelectorAll('table'));
        allTables.forEach(t => {{ try {{ new Tablesort(t); }} catch (e) {{}} }});
        if (table) {{
            const priorityHeader = table.querySelector('th:nth-child(2)');
            if (priorityHeader) {{ priorityHeader.click(); }}
        }}
    }} catch (err) {{
        console.warn('Tablesort 初始化失败', err);
    }}

    const rowCounter = document.getElementById('rowCounter');
    const exportBtn = document.getElementById('exportCsv');
    const clearBtn = document.getElementById('clearMarks');
    const ownerInput = document.getElementById('defaultOwner');
    const ownerSelect = document.getElementById('defaultOwnerSelect');
    const ownerSuggestions = {owner_suggestions_js};
    const envDefaultOwner = {env_default_owner_js};
    const followupSummary = document.getElementById('followupSummary');
    // 默认跟进人：先读本地，其次用环境默认；提供下拉建议
    let defaultOwner = '';
    try {{ defaultOwner = localStorage.getItem('followupOwner') || ''; }} catch (e) {{}}
    if (!defaultOwner && envDefaultOwner) {{ defaultOwner = envDefaultOwner; }}
    if (ownerSelect) {{
        if (defaultOwner) {{ try {{ ownerSelect.value = defaultOwner; }} catch (e) {{}} }}
        ownerSelect.addEventListener('change', () => {{
            try {{ localStorage.setItem('followupOwner', ownerSelect.value || ''); }} catch (e) {{}}
        }});
    }} else if (ownerInput) {{
        ownerInput.value = defaultOwner;
        ownerInput.addEventListener('input', () => {{
            try {{ localStorage.setItem('followupOwner', ownerInput.value.trim()); }} catch (e) {{}}
        }});
    }}
    const todayStr = new Date().toISOString().split('T')[0];  // 实时获取当前日期
    const storageKey = 'followup-' + todayStr;

    const tags = {tags_js};
    const platforms = {platforms_js};
    const detailMap = {details_js};
    const globalDetails = {global_details_js};
    const productSearchIndex = {product_search_index_js};
    const idIndex = {id_index_js};
    const nameIndex = {name_index_js};
            const globalMeta = {global_meta_js};
            const aiAnalysisMap = {ai_analysis_map_js};
            const deepseekApiKey = {deepseek_key_js};
            const cooldownDays = {max(0, cooldown_days)};
        
     // 用于本地冷却期计算
    const cooldownKeys = {cooldown_keys_js};
    const cooldownCustomers = {cooldown_customers_js};  // 冷却期客户联系日期
    const writeEnabled = {str(contact_write_enabled).lower()};
    async function markCompleted(payload) {{
        try {{
            const resp = await fetch('http://127.0.0.1:{contact_server_port}/mark', {{
                method: 'POST',
                headers: {{ 'Content-Type': 'application/json' }},
                body: JSON.stringify(payload),
            }});
            if (!resp.ok) return false;
            const data = await resp.json();
            return !!data.ok;
        }} catch (e) {{
            return false;
        }}
    }}
    try {{ window.cooldownKeys = cooldownKeys; }} catch (e) {{}}
    try {{ window.cooldownCustomers = cooldownCustomers; }} catch (e) {{}}

    // 冷却期客户现已整合进主列表的"冷却期"标签，无需单独面板


    // 全库命中解析：优先匹配单号/退单号；其次手机号（key 即为手机号）；最后姓名精确匹配
    function resolveGlobalKey(searchRaw) {{
        if (!searchRaw) {{ return ''; }}
        const lower = String(searchRaw).trim().toLowerCase();
        const digits = lower.replace(/\D/g, '');
        // 1) 单号/退单号（小写/纯数字）
        let key = '';
        if (idIndex && typeof idIndex === 'object') {{
            key = idIndex[lower] || (digits.length >= 6 ? (idIndex[digits] || '') : '');
        }}
        // 2) 直接以手机号为 key（常见情况）
        if (!key && digits.length >= 7 && globalDetails && typeof globalDetails === 'object' && globalDetails[digits]) {{
            key = digits;
        }}
        // 3) 姓名精确小写匹配
        if (!key && nameIndex && typeof nameIndex === 'object') {{
            key = nameIndex[lower] || '';
        }}
        return key || '';
    }}

    // 规范化字符串：仅保留 a-z0-9，便于货品名"FZ-1103 / FZ 1103 / FZ1103"等形式统一匹配
    function normAlphaNum(s) {{
        return String(s || '').toLowerCase().replace(/[^a-z0-9]/g, '');
    }}

    // ==================== 搜索性能优化 ====================
    // LRU 缓存：最多保存 50 个搜索结果
    const searchCache = new Map();
    const MAX_CACHE_SIZE = 50;

    /**
     * 缓存搜索结果
     * @param {{string}} key - 规范化的搜索关键词
     * @param {{object}} value - 搜索结果（matchedCustomers）
     */
    function cacheSearchResult(key, value) {{
        // LRU 驱逐：删除最早的条目
        if (searchCache.size >= MAX_CACHE_SIZE) {{
            const firstKey = searchCache.keys().next().value;
            searchCache.delete(firstKey);
            console.log(`Cache evicted: ${{firstKey}}`);
        }}
        searchCache.set(key, value);
    }}

    /**
     * 获取缓存的搜索结果
     * @param {{string}} key - 规范化的搜索关键词
     * @returns {{object|null}} - 缓存的结果或 null
     */
    function getCachedSearchResult(key) {{
        if (searchCache.has(key)) {{
            console.log(`✅ Cache HIT for "${{key}}"`);
            return searchCache.get(key);
        }}
        return null;
    }}
    // =====================================================

    /**
     * 通过货品名搜索客户（使用预建索引）
     * @param {{string}} key - 客户key（手机号）
     * @param {{string}} search - 搜索词
     * @returns {{boolean}} 是否匹配
     */
    function matchesProductName(key, search) {{
        if (!search || search.length < 2) return false;
        if (!productSearchIndex || typeof productSearchIndex !== 'object') return false;

        const normalized = normAlphaNum(search);
        if (!normalized) return false;

        // 精确匹配索引
        if (productSearchIndex[normalized] && Array.isArray(productSearchIndex[normalized])) {{
            if (productSearchIndex[normalized].includes(key)) {{
                return true;
            }}
        }}

        // 模糊匹配索引（支持部分匹配，如"FZ"匹配"FZ1103"）
        for (const [productKey, customerKeys] of Object.entries(productSearchIndex)) {{
            if (productKey.includes(normalized) && Array.isArray(customerKeys) && customerKeys.includes(key)) {{
                return true;
            }}
        }}

        return false;
    }}

    function populateSelect(select, values) {{
        if (!select || !values) {{
            return;
        }}
        values.forEach(value => {{
            const option = document.createElement('option');
            option.value = value;
            option.textContent = value || '空';
            select.appendChild(option);
        }});
    }}

    populateSelect(tagFilter, tags);
    populateSelect(platformFilter, platforms);

    // 基于全库订单计算每个厂家的聚合数据
    var manufacturerList = [];
    var sortedManufacturers = [];

    function updateManufacturerData(daysFilter) {{
        var manufacturerStats = {{}};
        var nowMs = new Date(todayStr).getTime();
        var oneMonthMs = 30 * 24 * 60 * 60 * 1000;
        var twoMonthsMs = 60 * 24 * 60 * 60 * 1000;
        var filterMs = (daysFilter && daysFilter !== 'all') ? (parseInt(daysFilter) * 24 * 60 * 60 * 1000) : 0;

        if (typeof globalDetails === 'object' && globalDetails) {{
            Object.values(globalDetails).forEach(function(customerOrders) {{
                if (!Array.isArray(customerOrders)) return;
                customerOrders.forEach(function(order) {{
                    if (!order || typeof order !== 'object') return;
                    var mfr = (order['厂家'] || '').trim();
                    if (!mfr) return;
                    
                    // 日期逻辑
                    var dateStr = order['下单时间'] || order['下单日期'] || order['顾客付款日期'] || order['付款日期'] || '';
                    var orderMs = 0;
                    if (dateStr) {{
                        orderMs = new Date(dateStr).getTime();
                    }}
                    
                    // Time Filter
                    if (filterMs > 0 && orderMs > 0) {{
                        if ((nowMs - orderMs) > filterMs) return;
                    }}

                    var refundType = String(order['退款类型'] || '').trim();
                    var returnNo = String(order['退货单号'] || '').trim();
                    var orderNo = String(order['订单号'] || '').trim();
                    var payRaw = order['付款金额'];
                    var pay = parseFloat(payRaw);
                    var costRaw = order['打款金额'];
                    var cost = parseFloat(costRaw);
                    
                    // 优化判定逻辑：排除“取消”和“补”
                    var isCancelOrSupp = /取消|补/.test(refundType) || /取消|补/.test(orderNo);
                    var isReturn = !isCancelOrSupp && (/退/.test(refundType) || (returnNo && returnNo !== '/' && returnNo !== '-' && returnNo.toLowerCase() !== 'none'));
                    
                    var isValidOrder = isFinite(pay) && pay > 0 && !isCancelOrSupp;

                    if (!manufacturerStats[mfr]) {{
                        manufacturerStats[mfr] = {{ orders: 0, returns: 0, revenue: 0, cost: 0, validMarginRevenue: 0, validMarginCost: 0, orders_1m: 0, orders_2m: 0 }};
                    }}
                    if (isValidOrder) {{
                        manufacturerStats[mfr].orders += 1;
                        manufacturerStats[mfr].revenue += Math.max(0, pay);
                        if (isFinite(cost) && cost > 0) {{
                            manufacturerStats[mfr].cost += cost;
                            manufacturerStats[mfr].validMarginRevenue += Math.max(0, pay);
                            manufacturerStats[mfr].validMarginCost += cost;
                        }}
                        
                        if (orderMs > 0) {{
                            var diff = nowMs - orderMs;
                            if (diff <= oneMonthMs) {{
                                manufacturerStats[mfr].orders_1m += 1;
                            }}
                            if (diff <= twoMonthsMs) {{
                                manufacturerStats[mfr].orders_2m += 1;
                            }}
                        }}
                        
                        if (isReturn) {{
                            manufacturerStats[mfr].returns += 1;
                        }}
                    }}
                }});
            }});
        }}

        manufacturerList = Object.keys(manufacturerStats).map(function(name) {{
            var stat = manufacturerStats[name] || {{}};
            var orders = stat.orders || 0;
            var returns = stat.returns || 0;
            var revenue = stat.revenue || 0;
            var cost = stat.cost || 0;
            var returnRate = orders > 0 ? Math.min(1, Math.max(0, returns / orders)) : 0;
            var aov = orders > 0 ? revenue / orders : 0;
            var validMarginRevenue = stat.validMarginRevenue || 0;
            var validMarginCost = stat.validMarginCost || 0;
            var margin = validMarginRevenue > 0 ? (validMarginRevenue - validMarginCost) / validMarginRevenue : 0;
            var trend = stat.orders_1m - (stat.orders_2m - stat.orders_1m);

            return {{ 
                name: name, 
                orders: orders, 
                returns: returns, 
                revenue: revenue, 
                cost: cost,
                returnRate: returnRate, 
                aov: aov,
                margin: margin,
                trend: trend,
                orders_1m: stat.orders_1m,
                orders_2m: stat.orders_2m
            }};
        }}).filter(function(m) {{
            // Relaxed filter: Show any manufacturer with at least 1 total order if filtering by short time
            var threshold = (filterMs > 0 && filterMs <= oneMonthMs) ? 1 : 5;
            return m.orders >= threshold;
        }}).sort(function(a, b) {{
            return b.orders - a.orders;
        }});
        
        sortedManufacturers = manufacturerList.map(function(m) {{ return m.name; }});
    }}
    
    // Initial Load
    updateManufacturerData('60');
    populateSelect(manufacturerFilter, sortedManufacturers);

    (function renderManufacturerAnalytics() {{
        try {{
            // Use the new div container instead of canvas
            var chartContainer = document.getElementById('manufacturerChart');
            var tableBody = document.getElementById('manufacturerSummaryBody');
            var metricSelect = document.getElementById('manufacturerMetricSelect');
            var container = chartContainer ? chartContainer.parentElement : null;
            
            if (!chartContainer || !tableBody || !metricSelect || !container) {{
                return;
            }}
            
            // Create Details Container if not exists
            var detailsContainer = document.getElementById('manufacturerDetails');
            if (!detailsContainer) {{
                detailsContainer = document.createElement('div');
                detailsContainer.id = 'manufacturerDetails';
                detailsContainer.style.marginTop = '20px';
                detailsContainer.style.padding = '15px';
                detailsContainer.style.background = '#f8fafc';
                detailsContainer.style.borderRadius = '8px';
                detailsContainer.style.border = '1px solid #e2e8f0';
                detailsContainer.style.display = 'none'; // Hidden by default
                container.appendChild(detailsContainer);
            }}

            var MAX_BARS = 50; // Increased limit to show more manufacturers

            function getSortedByMetric(metric) {{
                var data = manufacturerList.slice();
                if (metric === 'return_rate') {{
                    data.sort(function(a, b) {{
                        return (b.returnRate - a.returnRate) || (b.orders - a.orders);
                    }});
                }} else if (metric === 'aov') {{
                    data.sort(function(a, b) {{
                        return (b.aov - a.aov) || (b.orders - a.orders);
                    }});
                }} else {{
                    data.sort(function(a, b) {{
                        return b.orders - a.orders;
                    }});
                }}
                return data.slice(0, MAX_BARS);
            }}
            
            var currentSkuFilter = null;
            
            window.toggleSkuFilter = function(sku, el) {{
                var table = document.getElementById('detailTable');
                if (!table) return;
                
                if (currentSkuFilter === sku) {{
                    currentSkuFilter = null;
                }} else {{
                    currentSkuFilter = sku;
                }}
                
                var diagDiv = document.getElementById('mfrDiagnosis');
                if (diagDiv) {{
                    var badges = diagDiv.querySelectorAll('span[onclick]');
                    badges.forEach(function(b) {{
                        b.style.boxShadow = 'none';
                        b.style.background = 'rgba(255,255,255,0.7)';
                    }});
                }}
                
                if (currentSkuFilter && el) {{
                    el.style.boxShadow = '0 0 0 2px #3b82f6';
                    el.style.background = '#fff';
                }}
                
                var tbody = table.querySelector('tbody');
                if (tbody) {{
                    var rows = Array.from(tbody.querySelectorAll('tr'));
                    rows.forEach(function(row) {{
                        if (!currentSkuFilter) {{
                            row.style.display = '';
                        }} else {{
                            var skuCell = row.children[2];
                            var rowSku = skuCell ? skuCell.textContent.trim() : '';
                            if (rowSku.indexOf(currentSkuFilter) !== -1) {{
                                row.style.display = '';
                            }} else {{
                                row.style.display = 'none';
                            }}
                        }}
                    }});
                }}
            }};

            window.toggleAiContent = function(header) {{
                var content = header.nextElementSibling;
                var icon = header.querySelector('.fa-chevron-up, .fa-chevron-down');
                if (content.style.display === 'none') {{
                    content.style.display = 'block';
                    if (icon) {{
                        icon.className = 'fa-solid fa-chevron-up';
                    }}
                }} else {{
                    content.style.display = 'none';
                    if (icon) {{
                        icon.className = 'fa-solid fa-chevron-down';
                    }}
                }}
            }};

            function showManufacturerDetails(mfrName) {{
                currentSkuFilter = null;
                const rows = [];
                // Statistics for diagnosis
                const skuStats = {{}};
                // Date Filter: Respect global selection
                var timeSelect = document.getElementById('manufacturerTimeSelect');
                var days = timeSelect ? timeSelect.value : '60';
                var filterMs = (days && days !== 'all') ? (parseInt(days) * 24 * 60 * 60 * 1000) : 0;
                var nowMs = new Date().getTime();
                const cutoffTime = filterMs > 0 ? (nowMs - filterMs) : 0;
                var timeLabel = (days === 'all') ? '全部时间' : ('近' + days + '天');

                if (typeof globalDetails === 'object' && globalDetails) {{
                    Object.values(globalDetails).forEach(function(customerOrders) {{
                        if (!Array.isArray(customerOrders)) return;
                        customerOrders.forEach(function(order) {{
                            if ((order['厂家'] || '').trim() === mfrName) {{
                                // Parse date to filter old orders
                                let odStr = String(order['下单时间'] || '');
                                let odTime = 0;
                                // Try standard YYYY-MM-DD match first for speed/accuracy
                                const match = odStr.match(/(\d{4})[-\/](\d{1,2})[-\/](\d{1,2})/);
                                if (match) {{
                                    odTime = new Date(parseInt(match[1]), parseInt(match[2])-1, parseInt(match[3])).getTime();
                                }} else {{
                                    const ts = Date.parse(odStr);
                                    if (!isNaN(ts)) odTime = ts;
                                }}
                                
                                // Skip if older than 60 days (and valid date found)
                                if (odTime > 0 && odTime < cutoffTime) return;

                                rows.push(order);
                                
                                // SKU Aggregation
                                const sku = (order['货品名'] || '未知').trim();
                                if (!skuStats[sku]) {{
                                    skuStats[sku] = {{ orders: 0, returns: 0, revenue: 0, cost: 0, validMarginOrders: 0 }};
                                }}
                                
                                const pay = parseFloat(order['付款金额'] || 0) || 0;
                                const cost = parseFloat(order['打款金额'] || 0) || 0;
                                const refundType = String(order['退款类型'] || order['状态'] || '').trim();
                                const orderNo = String(order['订单号'] || '').trim();
                                const returnNo = String(order['退货单号'] || '').trim();
                                
                                // 判定逻辑：排除“取消”和“补”
                                var isCancelOrSupp = /取消|补/.test(refundType) || /取消|补/.test(orderNo);
                                var isReturn = !isCancelOrSupp && (/退/.test(refundType) || (returnNo && returnNo !== '/' && returnNo !== '-'));
                                
                                skuStats[sku].orders += 1;
                                skuStats[sku].revenue += pay;
                                skuStats[sku].cost += cost;
                                
                                if (isReturn) {{
                                    skuStats[sku].returns += 1;
                                }}
                                
                                if (pay > 0 && cost > 0) {{
                                    skuStats[sku].validMarginOrders += 1;
                                }}
                            }}
                        }});
                    }});
                }}
                
                // Identify Issues
                const highReturnSkus = [];
                const lowMarginSkus = [];
                
                Object.keys(skuStats).forEach(function(sku) {{
                    const s = skuStats[sku];
                    if (s.orders >= 3) {{
                        const returnRate = s.returns / s.orders;
                        if (returnRate > 0.30) {{
                            highReturnSkus.push({{ sku: sku, rate: returnRate, orders: s.orders, returns: s.returns }});
                        }}
                        
                        // Margin check (only if we have cost data)
                        if (s.validMarginOrders >= 3) {{
                            const margin = (s.revenue - s.cost) / s.revenue;
                            if (margin < 0.15) {{
                                lowMarginSkus.push({{ sku: sku, margin: margin, orders: s.orders }});
                            }}
                        }}
                    }}
                }});
                
                // Sort by severity
                highReturnSkus.sort((a, b) => b.rate - a.rate);
                lowMarginSkus.sort((a, b) => a.margin - b.margin);

                // Sort orders by date descending
                rows.sort(function(a, b) {{
                    var da = new Date(a['下单时间'] || a['下单日期'] || a['顾客付款日期'] || '');
                    var db = new Date(b['下单时间'] || b['下单日期'] || b['顾客付款日期'] || '');
                    return db - da;
                }});

                // Reuse global detail panel elements
                if (detailTitle) detailTitle.innerHTML = mfrName + ' - 厂家订单明细 <span style="font-size:0.8em; color:#64748b; font-weight:normal;">(' + timeLabel + ': ' + rows.length + ' 条)</span>';
                
                // Inject Diagnosis Panel
                // Check if we already have a diagnosis container, if not create one before the table container
                let diagDiv = document.getElementById('mfrDiagnosis');
                if (!diagDiv) {{
                    diagDiv = document.createElement('div');
                    diagDiv.id = 'mfrDiagnosis';
                    diagDiv.style.marginBottom = '15px';
                    // Insert after title, before table container
                    const tableContainer = detailTable.parentElement; 
                    tableContainer.parentNode.insertBefore(diagDiv, tableContainer);
                }}
                
                let diagHtml = '';
                if (highReturnSkus.length > 0 || lowMarginSkus.length > 0) {{
                    diagHtml += '<div style="background:#fff1f2; border:1px solid #fda4af; border-radius:6px; padding:12px; font-size:13px; color:#881337;">';
                    diagHtml += '<div style="font-weight:bold; margin-bottom:8px; display:flex; align-items:center;">';
                    diagHtml += '<span style="margin-right:6px;">🕵️‍♂️ 智能诊断发现：</span>';
                    diagHtml += '</div>';
                    
                    if (highReturnSkus.length > 0) {{
                        diagHtml += '<div style="margin-bottom:10px;">';
                        diagHtml += '<div style="margin-bottom:6px;"><span style="background:#e11d48; color:white; padding:2px 6px; border-radius:4px; font-size:11px;">高退货风险</span></div>';
                        diagHtml += '<div style="display:flex; flex-wrap:wrap; gap:6px;">';
                        diagHtml += highReturnSkus.map(i => `<span onclick="toggleSkuFilter('${{i.sku}}', this)" style="cursor:pointer; background:rgba(255,255,255,0.7); border:1px solid #fb7185; padding:3px 8px; border-radius:4px; display:flex; align-items:center; gap:4px;" title="点击筛选/取消筛选该货品"><b>${{i.sku}}</b> <span style="font-size:11px; color:#9f1239;">${{(i.rate*100).toFixed(0)}}%</span></span>`).join('');
                        diagHtml += '</div></div>';
                    }}
                    
                    if (lowMarginSkus.length > 0) {{
                        diagHtml += '<div>';
                        diagHtml += '<div style="margin-bottom:6px;"><span style="background:#f59e0b; color:white; padding:2px 6px; border-radius:4px; font-size:11px;">低毛利预警</span></div>';
                        diagHtml += '<div style="display:flex; flex-wrap:wrap; gap:6px;">';
                        diagHtml += lowMarginSkus.map(i => `<span onclick="toggleSkuFilter('${{i.sku}}', this)" style="cursor:pointer; background:rgba(255,255,255,0.7); border:1px solid #fbbf24; padding:3px 8px; border-radius:4px; display:flex; align-items:center; gap:4px;" title="点击筛选/取消筛选该货品"><b>${{i.sku}}</b> <span style="font-size:11px; color:#92400e;">${{(i.margin*100).toFixed(0)}}%</span></span>`).join('');
                        diagHtml += '</div></div>';
                    }}
                    diagHtml += '</div>';
                }} else if (rows.length > 0) {{
                    diagHtml = '<div style="background:#f0fdf4; border:1px solid #bbf7d0; border-radius:6px; padding:8px 12px; font-size:13px; color:#166534;">✅ 该厂家近期表现良好，暂未发现高风险 SKU。</div>';
                }}
                diagDiv.innerHTML = diagHtml;

                // Add AI Analysis Section if available
                let aiDiv = document.getElementById('mfrAiAnalysis');
                if (!aiDiv) {{
                    aiDiv = document.createElement('div');
                    aiDiv.id = 'mfrAiAnalysis';
                    aiDiv.style.marginBottom = '15px';
                    diagDiv.parentNode.insertBefore(aiDiv, diagDiv.nextSibling);
                }}
                
                let aiContent = aiAnalysisMap[mfrName] || '';
                
                // Helper to render AI HTML
                const renderAiHtml = (content) => {{
                    return '<div style="background:linear-gradient(135deg, #f5f3ff 0%, #ede9fe 100%); border:1px solid #c4b5fd; border-radius:8px; padding:15px; position:relative; overflow:hidden;">' +
                        '<div style="position:absolute; right:-10px; top:-10px; opacity:0.1;"><i class="fa-solid fa-brain" style="font-size:60px; color:#7c3aed;"></i></div>' +
                        '<div onclick="toggleAiContent(this)" style="cursor:pointer; font-weight:bold; color:#6d28d9; margin-bottom:10px; display:flex; align-items:center; justify-content:space-between;">' +
                        '<div style="display:flex; align-items:center;">' +
                        '<span style="background:#7c3aed; color:white; width:24px; height:24px; border-radius:50%; display:flex; align-items:center; justify-content:center; margin-right:8px; font-size:12px;">AI</span>' +
                        'DeepSeek 深度运营策略：</div><i class="fa-solid fa-chevron-up"></i></div>' +
                        '<div style="font-size:13.5px; line-height:1.6; color:#4c1d95; white-space:pre-wrap;">' + content + '</div>' +
                        '</div>';
                }};

                if (aiContent) {{
                    aiDiv.innerHTML = renderAiHtml(aiContent);
                    aiDiv.style.display = 'block';
                }} else if (deepseekApiKey) {{
                    // Show "Generate Analysis" button
                    aiDiv.innerHTML = '<button id="btnGenAi" style="background:linear-gradient(135deg, #7c3aed, #6d28d9); color:white; border:none; padding:8px 16px; border-radius:6px; cursor:pointer; font-size:13px; display:flex; align-items:center; gap:6px; box-shadow:0 2px 4px rgba(124,58,237,0.3); transition:all 0.2s;"><i class="fa-solid fa-wand-magic-sparkles"></i> 生成 DeepSeek 运营分析</button>';
                    aiDiv.style.display = 'block';
                    
                    const btn = document.getElementById('btnGenAi');
                    btn.onclick = async function() {{
                        btn.disabled = true;
                        btn.innerHTML = '<i class="fa-solid fa-circle-notch fa-spin"></i> 正在分析中...';
                        
                        try {{
                            // Prepare SKU stats for this manufacturer
                            const skuStats = {{}};
                            rows.forEach(r => {{
                                const item = r['商品名称'] || r['货品名'] || '未知';
                                if (!skuStats[item]) skuStats[item] = {{ orders: 0, returns: 0, revenue: 0 }};
                                const amt = parseFloat(r['付款金额']) || 0;
                                const refundType = String(r['退款类型'] || '');
                                const returnNo = String(r['退货单号'] || '');
                                const isReturn = refundType.includes('退') || (returnNo && returnNo !== '/' && returnNo !== '-');
                                
                                if (amt > 0) {{ skuStats[item].orders++; skuStats[item].revenue += amt; }}
                                if (isReturn) skuStats[item].returns++;
                            }});
                            
                            const statsList = Object.entries(skuStats).map(([k, v]) => ({{
                                sku: k, ...v,
                                return_rate: v.orders > 0 ? (v.returns / v.orders) : 0
                            }})).sort((a,b) => b.revenue - a.revenue).slice(0, 20);

                            const resp = await fetch('http://127.0.0.1:{contact_server_port}/analyze_manufacturer', {{
                                method: 'POST',
                                headers: {{ 'Content-Type': 'application/json' }},
                                body: JSON.stringify({{
                                    api_key: deepseekApiKey,
                                    mfr_name: mfrName,
                                    sku_stats: statsList
                                }})
                            }});
                            
                            if (!resp.ok) throw new Error('API Error');
                            const data = await resp.json();
                            if (data.ok && data.analysis) {{
                                aiAnalysisMap[mfrName] = data.analysis;
                                aiDiv.innerHTML = renderAiHtml(data.analysis);
                            }} else {{
                                throw new Error(data.error || 'Unknown error');
                            }}
                        }} catch (err) {{
                            console.error(err);
                            btn.innerHTML = '<i class="fa-solid fa-triangle-exclamation"></i> 分析失败: ' + (err.message || '未知错误');
                            btn.style.background = '#ef4444';
                            setTimeout(() => {{ 
                                btn.disabled = false; 
                                btn.innerHTML = '<i class="fa-solid fa-wand-magic-sparkles"></i> 重试 DeepSeek 分析';
                                btn.style.background = 'linear-gradient(135deg, #7c3aed, #6d28d9)';
                            }}, 3000);
                        }}
                    }};
                }} else {{
                    aiDiv.style.display = 'none';
                }}

                if (detailTbody) detailTbody.innerHTML = '';
                
                if (!rows.length) {{
                    const tr = document.createElement('tr');
                    const td = document.createElement('td');
                    td.colSpan = 15;
                    td.textContent = '暂无数据';
                    tr.appendChild(td);
                    if (detailTbody) detailTbody.appendChild(tr);
                }} else {{
                    const frag = document.createDocumentFragment();
                    rows.forEach(function(r) {{
                        const tr = document.createElement('tr');
                        const paymentAmt = r['付款金额'];
                        const costAmt = r['打款金额'];
                        
                        // Helper function for margin calculation
                        const calcM = (typeof calcMargin === 'function') ? calcMargin : function(p, c) {{
                            const pv = parseFloat(p) || 0;
                            const cv = parseFloat(c) || 0;
                            if (pv <= 0) return '-';
                            return ((pv - cv) / pv * 100).toFixed(1) + '%';
                        }};

                        const formatA = (typeof formatAmount === 'function') ? formatAmount : function(v) {{
                            const n = parseFloat(v);
                            return isFinite(n) ? n.toFixed(2) : '';
                        }};

                        const cols = [
                            (r['下单时间'] || ''),
                            (r['姓名'] || '') + '<br><span style="color:#64748b;font-size:12px">' + (r['手机号'] || '') + '</span>',
                            (r['商品名称'] || r['货品名'] || '') + '<br><span style="color:#64748b;font-size:12px">' + (r['颜色'] || '') + ' ' + (r['尺码'] || '') + '</span>',
                            '<div style="display:flex; flex-direction:column; align-items:flex-end;"><span>' + formatA(paymentAmt) + ' / ' + formatA(costAmt) + '</span><span style="font-weight:bold; color:' + (parseFloat(calcM(paymentAmt, costAmt)) < 15 ? '#f59e0b' : '#10b981') + '">' + calcM(paymentAmt, costAmt) + '</span></div>',
                            (r['订单号'] || '') + (r['退货单号'] ? '<br><span style="color:#ef4444; font-size:12px">退: ' + r['退货单号'] + '</span>' : ''),
                            (r['退款类型'] || '') + (r['退款原因'] ? '<br><span style="color:#64748b; font-size:12px">' + r['退款原因'] + '</span>' : '')
                        ];
                        
                        cols.forEach(function(html, i) {{
                            const td = document.createElement('td');
                            if (i === 0) {{
                                const digits = String(html).replace(/\D/g, '');
                                if (digits.length >= 8) {{ td.setAttribute('data-sort-value', digits.slice(0,8)); }}
                            }} else if (i === 3) {{
                                const mNum = parseFloat(calcM(paymentAmt, costAmt));
                                if (isFinite(mNum)) td.setAttribute('data-sort-value', String(mNum));
                                td.style.textAlign = 'right';
                            }}
                            td.innerHTML = html;
                            tr.appendChild(td);
                        }});
                        frag.appendChild(tr);
                    }});
                    if (detailTbody) detailTbody.appendChild(frag);
                }}

                // Show Panel
                if (detailBackdrop) detailBackdrop.classList.remove('hidden');
                if (panel) panel.classList.remove('hidden');
                panelOpen = true;

                // Re-init tablesort
                if (typeof Tablesort !== 'undefined' && detailTable) {{
                    try {{ new Tablesort(detailTable); }} catch (e) {{}}
                }}
            }}

            function renderChart(metric) {{
                var data = getSortedByMetric(metric);
                chartContainer.innerHTML = '';
                
                if (!data.length) {{
                    chartContainer.innerHTML = '<div style="color:#94a3b8; font-size:12px; text-align:center; padding-top:40px;">暂无足够的厂家订单数据</div>';
                    tableBody.innerHTML = '';
                    return;
                }}

                var maxValue = 0;
                if (metric === 'return_rate') {{
                    // For return rate, stick to a fixed max or slightly above max
                    maxValue = Math.max.apply(null, data.map(function(d){{ return d.returnRate; }})); 
                    maxValue = Math.max(0.5, maxValue); // Minimum 50% scale
                }} else if (metric === 'aov') {{
                    maxValue = Math.max.apply(null, data.map(function(d){{ return d.aov; }}));
                }} else {{
                    maxValue = Math.max.apply(null, data.map(function(d){{ return d.orders; }}));
                }}

                var fragChart = document.createDocumentFragment();

                data.forEach(function(d, idx) {{
                    var val = 0;
                    var label = '';
                    var barColor = '#3b82f6'; // Default Blue
                    
                    if (metric === 'return_rate') {{
                        val = d.returnRate;
                        label = (d.returnRate * 100).toFixed(1) + '%';
                        // Red > Orange > Green logic
                        barColor = val > 0.35 ? '#ef4444' : (val > 0.20 ? '#f59e0b' : '#10b981');
                    }} else if (metric === 'aov') {{
                        val = d.aov;
                        label = '¥' + d.aov.toFixed(0);
                        barColor = '#8b5cf6'; // Violet
                    }} else {{
                        val = d.orders;
                        label = d.orders + ' 单';
                        barColor = '#0ea5e9'; // Sky
                    }}

                    var pct = maxValue ? (val / maxValue) * 100 : 0;
                    
                    // Row Container (Clickable)
                    var rowDiv = document.createElement('div');
                    rowDiv.style.marginBottom = '12px';
                    rowDiv.style.cursor = 'pointer';
                    rowDiv.title = '点击查看详情';
                    rowDiv.onclick = function() {{ showManufacturerDetails(d.name); }};
                    
                    // Label Row
                    var labelDiv = document.createElement('div');
                    labelDiv.style.display = 'flex';
                    labelDiv.style.justifyContent = 'space-between';
                    labelDiv.style.fontSize = '12px';
                    labelDiv.style.marginBottom = '4px';
                    labelDiv.style.color = '#64748b';
                    
                    var nameSpan = document.createElement('span');
                    nameSpan.textContent = (idx + 1) + '. ' + d.name;
                    nameSpan.style.fontWeight = '600';
                    nameSpan.style.color = '#334155';
                    
                    var valSpan = document.createElement('span');
                    valSpan.textContent = label;
                    valSpan.style.fontFamily = 'ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, monospace';
                    
                    labelDiv.appendChild(nameSpan);
                    labelDiv.appendChild(valSpan);
                    
                    // Bar Track
                    var trackDiv = document.createElement('div');
                    trackDiv.style.height = '6px';
                    trackDiv.style.background = '#f1f5f9';
                    trackDiv.style.borderRadius = '3px';
                    trackDiv.style.overflow = 'hidden';
                    
                    // Bar Fill
                    var fillDiv = document.createElement('div');
                    fillDiv.style.width = pct + '%';
                    fillDiv.style.height = '100%';
                    fillDiv.style.background = barColor;
                    fillDiv.style.borderRadius = '3px';
                    fillDiv.style.transition = 'width 0.5s ease-out';
                    
                    trackDiv.appendChild(fillDiv);
                    rowDiv.appendChild(labelDiv);
                    rowDiv.appendChild(trackDiv);
                    
                    fragChart.appendChild(rowDiv);
                }});
                
                chartContainer.appendChild(fragChart);

                // Render Table (re-use the same data sorted by metric)
                tableBody.innerHTML = '';
                var fragTable = document.createDocumentFragment();
                data.forEach(function(d) {{
                    var tr = document.createElement('tr');
                    tr.style.cursor = 'pointer';
                    tr.title = '点击查看详情';
                    tr.onclick = function() {{ showManufacturerDetails(d.name); }};
                    tr.onmouseover = function() {{ this.style.backgroundColor = '#f1f5f9'; }};
                    tr.onmouseout = function() {{ this.style.backgroundColor = ''; }};

                    var rrPct = (d.returnRate * 100).toFixed(1) + '%';
                    var aovVal = d.aov.toFixed(2);
                    var revVal = '¥' + d.revenue.toLocaleString('en-US', {{minimumFractionDigits: 0, maximumFractionDigits: 0}});
                    var marginPct = (d.margin * 100).toFixed(1) + '%';
                    
                    var cols = [
                        d.name,
                        String(d.orders),
                        String(d.trend),
                        revVal,
                        rrPct,
                        marginPct,
                        ''
                    ];
                    
                    cols.forEach(function(text, idx) {{
                        var td = document.createElement('td');
                        if (idx === 1) {{ // Orders
                            td.setAttribute('data-sort-value', String(d.orders));
                            td.innerHTML = '<span style="font-weight:600; color:#334155;">' + text + '</span>';
                        }} else if (idx === 2) {{ // Trend
                            td.setAttribute('data-sort-value', String(d.trend));
                            var color = d.trend > 0 ? '#ef4444' : (d.trend < 0 ? '#10b981' : '#94a3b8');
                            var arrow = d.trend > 0 ? '⬆️' : (d.trend < 0 ? '⬇️' : '➖');
                            td.innerHTML = '<span style="color:' + color + '; font-weight:bold;">' + arrow + '</span>';
                        }} else if (idx === 3) {{ // Revenue
                            td.setAttribute('data-sort-value', String(d.revenue));
                            td.style.fontFamily = 'ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, monospace';
                            td.textContent = text;
                        }} else if (idx === 4) {{ // Return Rate
                            td.setAttribute('data-sort-value', String(d.returnRate));
                            var rrVal = d.returnRate;
                            var color = rrVal > 0.35 ? '#ef4444' : (rrVal > 0.20 ? '#f59e0b' : '#10b981');
                            td.innerHTML = '<span style="color:' + color + '; font-weight:600;">' + text + '</span>';
                        }} else if (idx === 5) {{ // Margin
                            td.setAttribute('data-sort-value', String(d.margin));
                            var mVal = d.margin;
                            var color = mVal < 0.15 ? '#f59e0b' : '#334155';
                            td.innerHTML = '<span style="color:' + color + ';">' + text + '</span>';
                        }} else if (idx === 6) {{ // Risk Label
                             var tags = [];
                             if (d.returnRate > 0.30) tags.push('<span style="color:#ef4444; background:#fee2e2; padding:2px 6px; border-radius:4px; font-size:11px; margin-right:4px;">高退货</span>');
                             if (d.margin < 0.15) tags.push('<span style="color:#f59e0b; background:#fef3c7; padding:2px 6px; border-radius:4px; font-size:11px; margin-right:4px;">低毛利</span>');
                             if (d.returnRate < 0.20 && d.margin > 0.25) tags.push('<span style="color:#10b981; background:#d1fae5; padding:2px 6px; border-radius:4px; font-size:11px;">优质</span>');
                             td.innerHTML = tags.join('');
                        }} else {{ // Name
                            td.textContent = text;
                            td.style.fontWeight = '500';
                            td.style.color = '#1e293b';
                        }}
                        tr.appendChild(td);
                    }});
                    fragTable.appendChild(tr);
                }});
                tableBody.appendChild(fragTable);
            }}

            var initialMetric = metricSelect.value || 'return_rate';
            renderChart(initialMetric);

            var timeSelect = document.getElementById('manufacturerTimeSelect');
            if (timeSelect) {{
                timeSelect.addEventListener('change', function() {{
                    updateManufacturerData(this.value);
                    renderChart(metricSelect.value || 'return_rate');
                }});
            }}

            metricSelect.addEventListener('change', function() {{
                renderChart(metricSelect.value || 'return_rate');
            }});
        }} catch (e) {{
            console.warn('厂家分析渲染失败:', e);
        }}
    }})();

    let followupMap = {{}};
    try {{
        const stored = localStorage.getItem(storageKey);
        if (stored) {{
            followupMap = JSON.parse(stored) || {{}};
        }}
    }} catch (err) {{
        followupMap = {{}};
    }}

    function sanitizeFollowupMap() {{
        const cleaned = {{}};
        Object.entries(followupMap).forEach(([key, entry]) => {{
            if (entry && entry.date === todayStr) {{
                cleaned[key] = entry;
            }}
        }});
        followupMap = cleaned;
    }}

    function updateSummary() {{
        if (!followupSummary) {{
            return;
        }}
        const completed = Object.keys(followupMap).length;
        const total = rows.length;
        followupSummary.textContent = total
            ? '已标记 ' + completed + ' / ' + total + ' 人'
            : '';
    }}

    function persistFollowupMap() {{
        if (Object.keys(followupMap).length) {{
            localStorage.setItem(storageKey, JSON.stringify(followupMap));
        }} else {{
            localStorage.removeItem(storageKey);
        }}
        updateSummary();
    }}

    function setRowState(row, checked) {{
        if (!row) {{
            return;
        }}
        row.classList.toggle('completed', checked);
    }}

    sanitizeFollowupMap();
    persistFollowupMap();

    const checkboxes = Array.from(document.querySelectorAll('.followup-checkbox'));
    checkboxes.forEach(checkbox => {{
        const row = checkbox.closest('tr');
        const key = checkbox.dataset.key;
        if (followupMap[key]) {{
            checkbox.checked = true;
            setRowState(row, true);
        }}
        checkbox.addEventListener('change', () => {{
            if (checkbox.checked) {{
                const ownerToUse = (ownerSelect && ownerSelect.value)
                    || (ownerInput && ownerInput.value && ownerInput.value.trim())
                    || (envDefaultOwner || '')
                    || (row && row.getAttribute('data-owner'))
                    || '';
                followupMap[key] = {{
                    phone: checkbox.dataset.phone,
                    name: checkbox.dataset.name,
                    date: todayStr,
                    platform: (row && row.getAttribute('data-platform')) || '',
                    owner: ownerToUse,
                    note: '已在仪表盘标记完成',
                    timestamp: new Date().toISOString(),
                }};
                setRowState(row, true);
                if (writeEnabled) {{
                    const payload = {{
                        phone: checkbox.dataset.phone || '',
                        name: checkbox.dataset.name || '',
                        owner: ownerToUse,
                        platform: (row && row.getAttribute('data-platform')) || '',
                        note: '已在仪表盘标记完成',
                    }};
                    markCompleted(payload).then(ok => {{
                        if (!ok) {{
                            alert('写入飞书失败，请稍后重试');
                            checkbox.checked = false;
                            delete followupMap[key];
                            setRowState(row, false);
                            persistFollowupMap();
                        }}
                    }});
                }}
            }} else {{
                delete followupMap[key];
                setRowState(row, false);
            }}
            persistFollowupMap();
        }});
    }});

    // 当前选中的列表
    let currentList = '全部';

    // 检查客户的订单中是否包含指定厂家
    function matchesManufacturerForCustomer(customerKey, targetManufacturer) {{
        if (!targetManufacturer || !globalDetails || !customerKey) return true;
        const customerOrders = globalDetails[customerKey];
        if (!Array.isArray(customerOrders)) return false;
        return customerOrders.some(order => {{
            const mfr = (order['厂家'] || '').trim();
            return mfr === targetManufacturer;
        }});
    }}

    function applyFilters() {{
        const opsSearch = document.getElementById('operationsSearchBox');
        const opsContainer = document.querySelector('.operations-search-container');
        let searchRaw = '';
        let isOpsMode = false;
        
        if (opsContainer && opsContainer.classList.contains('active') && opsSearch) {{
            searchRaw = opsSearch.value.trim();
            isOpsMode = true;
        }} else if (searchBox) {{
            searchRaw = searchBox.value.trim();
        }}
        const search = searchRaw.toLowerCase();

        // Ops Mode: Toggle Summary/Table based on search
        if (isOpsMode) {{
            const summary = document.querySelector('.summary');
            if (!search) {{
                if (summary) summary.style.display = 'grid';
                if (table) table.style.display = 'none';
                return;
            }} else {{
                if (summary) summary.style.display = 'none';
                if (table) table.style.display = 'table';
            }}
        }}

        const priority = priorityFilter ? priorityFilter.value : '';
        const tag = tagFilter ? tagFilter.value : '';
        const platform = platformFilter ? platformFilter.value : '';
        const manufacturer = manufacturerFilter ? manufacturerFilter.value : '';

        let visible = 0;
        rows.forEach(row => {{
            const rowPriority = row.dataset.bucket || '';
            const rowPlatform = row.dataset.platform || '';
            const rowList = row.dataset.list || '活跃培养';  // 新增：获取所属列表
            const tagCell = row.querySelector('td[data-header="风险标签"]');
            const tagText = tagCell ? tagCell.textContent.trim() : '';
            const idsAttr = (row.getAttribute('data-ids') || '').toLowerCase();
            const key = row.getAttribute('data-key') || '';

            // 增强搜索逻辑：支持全文、订单号、货品名搜索
            const matchesSearch = !search ||
                row.textContent.toLowerCase().includes(search) ||
                idsAttr.includes(search) ||
                matchesProductName(key, search);  // 新增：货品名搜索

            const matchesPriority = !priority || rowPriority === priority;
            const matchesTag = !tag || tagText.split(',').map(t => t.trim()).includes(tag);
            const matchesPlatform = !platform || rowPlatform === platform;
            const matchesManufacturer = !manufacturer || matchesManufacturerForCustomer(key, manufacturer);
            // 列表过滤：默认“全部”不显示冷却期客户，避免重复触达
            const matchesList = (currentList === '全部'
                ? (rowList !== '冷却期')
                : (rowList === currentList));

            if (matchesSearch && matchesPriority && matchesTag && matchesPlatform && matchesManufacturer && matchesList) {{
                row.style.display = '';
                visible++;
            }} else {{
                row.style.display = 'none';
            }}
        }});
        // 在无匹配时尝试插入“全库命中”的临时行（支持单号/退单号/手机号/姓名），以便下方表格也可查看
        const tbody = table ? table.querySelector('tbody') : null;
        // 清理旧的临时行
        if (tbody) {{
            Array.from(tbody.querySelectorAll('tr.synthetic-row')).forEach(tr => tr.remove());
        }}
        let addedSynthetic = false;
        if (tbody) {{
            const hasVisible = visible > 0;
            // Use searchRaw from outer scope
            if (!hasVisible && searchRaw) {{
                const _key = resolveGlobalKey(searchRaw);
                if (_key && globalDetails && globalDetails[_key]) {{
                    const det = globalDetails[_key] || [];
                    // 基础字段
                    const name = det.length ? (det[0]['姓名'] || '') : '';
                    const phone = (/^\d{7,}$/.test(_key) ? _key : '');
                    // 最近下单日
                    let lastDigits = 0, lastDateStr = '';
                    det.forEach(r => {{
                        const d = String(r['下单时间'] || '').replace(/\D/g, '').slice(0,8);
                        if (d && d.length === 8) {{
                            const n = parseInt(d, 10);
                            if (n > lastDigits) {{ lastDigits = n; lastDateStr = (r['下单时间'] || ''); }}
                        }}
                    }});
                    // 平台/偏好
                    const platCount = {{}};
                    const itemCount = {{}};
                    det.forEach(r => {{
                        const p = (r['下单平台'] || '').trim();
                        if (p) platCount[p] = (platCount[p] || 0) + 1;
                        const it = (r['货品名'] || '').trim();
                        if (it) itemCount[it] = (itemCount[it] || 0) + 1;
                    }});
                    function topKey(map) {{
                        let best = '', cnt = -1;
                        Object.keys(map).forEach(k => {{ if (map[k] > cnt) {{ cnt = map[k]; best = k; }} }});
                        return best;
                    }}
                    const topPlat = topKey(platCount);
                    const favItem = topKey(itemCount);
                    // data-ids
                    const idsSet = new Set();
                    det.forEach(r => {{
                        ['订单号','退货单号'].forEach(kf => {{
                            const v = (r[kf] || '').toString().trim();
                            if (v) {{ idsSet.add(v.toLowerCase()); const dg = v.replace(/\D/g,''); if (dg.length>=6) idsSet.add(dg); }}
                        }});
                    }});
                    const idsAttr = Array.from(idsSet).join('|');

                    const tr = document.createElement('tr');
                    const meta = (globalMeta && globalMeta[_key]) ? globalMeta[_key] : null;
                    const prClass = meta && meta.priority_class ? meta.priority_class : 'priority-other';
                    tr.className = prClass + ' synthetic-row';
                    tr.setAttribute('data-key', _key);
                    tr.setAttribute('data-phone', phone);
                    tr.setAttribute('data-name', name);
                    tr.setAttribute('data-score', meta && meta.priority_score !== undefined ? String(meta.priority_score) : '');
                    tr.setAttribute('data-bucket', meta && meta.priority_bucket ? meta.priority_bucket : '负分');
                    tr.setAttribute('data-platform', topPlat);
                    tr.setAttribute('data-last-order', lastDateStr);
                    tr.setAttribute('data-cycle', '');
                    tr.setAttribute('data-threshold', '');
                    tr.setAttribute('data-category-cycle', '');
                    tr.setAttribute('data-ids', idsAttr);

                    const headers = ['标记完成','优先分','姓名','主要平台','手机号','最近下单日','风险标签','推荐动作','偏好单品','有效订单数','退货率','未复购天数','平均客单价'];
                    const scoreDisp = meta && typeof meta.priority_score === 'number' ? String(Math.round(meta.priority_score)) : '';
                    const valueLevel = meta && meta.customer_value ? meta.customer_value : '';
                    const ords = meta && typeof meta.orders === 'number' ? String(meta.orders) : String(det.length || '');
                    const rr = meta && typeof meta.return_rate === 'number' ? (meta.return_rate*100).toFixed(1)+'%' : '';
                    const ds = meta && typeof meta.days === 'number' ? String(meta.days) : '';
                    const aov = meta && typeof meta.aov === 'number' ? meta.aov.toFixed(2) : '';
                    const fav = meta && meta.favorite ? meta.favorite : (favItem || '');
                    const values = [
                        scoreDisp,        // 优先分
                        name,             // 姓名
                        topPlat || (meta && meta.platform ? meta.platform : ''),    // 主要平台
                        phone || '',      // 手机号
                        lastDateStr || (meta && meta.last_order ? meta.last_order : ''),
                        '全库命中(未纳入触达)', // 风险标签
                        '',               // 推荐动作
                        fav,              // 偏好单品
                        ords,             // 有效订单数
                        rr,               // 退货率
                        ds,               // 未复购天数
                        aov,              // 平均客单价
                    ];
                    // 复用打勾逻辑
                    const cb = document.createElement('input');
                    cb.type = 'checkbox';
                    cb.className = 'followup-checkbox';
                    cb.setAttribute('data-key', _key);
                    cb.setAttribute('data-phone', phone);
                    cb.setAttribute('data-name', name);
                    headers.forEach((h, idx) => {{
                        const td = document.createElement('td');
                        td.setAttribute('data-header', h);
                        if (idx === 0) {{
                            td.appendChild(cb);
                        }} else {{
                            td.textContent = values[idx-1] || '';
                        }}
                        tr.appendChild(td);
                    }});
                    cb.addEventListener('change', () => {{
                        if (cb.checked) {{
                            followupMap[_key] = {{ phone: phone, name: name, date: todayStr, timestamp: new Date().toISOString() }};
                            setRowState(tr, true);
                        }} else {{
                            delete followupMap[_key];
                            setRowState(tr, false);
                        }}
                        persistFollowupMap();
                    }});
                    tr.addEventListener('click', (e) => {{
                        if (panelOpen) return;
                        const isCheckbox = e.target && (e.target.tagName === 'INPUT' || e.target.closest('input'));
                        if (isCheckbox) return;
                        openDetailForKey(_key, name);
                    }});
                    tbody.appendChild(tr);
                    addedSynthetic = true;
                }} else {{
                    // 若未命中客户 key，则尝试在全库中寻找“包含关键字”的最相关客户（遍历整行文本）
                    const qn = normAlphaNum(searchRaw);
                    let bestKey = '';
                    let bestCount = 0;
                    try {{
                        Object.entries(globalDetails || {{}}).forEach(([k, list]) => {{
                            if (!Array.isArray(list)) return;
                            let c = 0;
                            for (let i=0;i<list.length;i++) {{
                                const r = list[i];
                                if (!r || typeof r !== 'object') continue;
                                let blob = '';
                                try {{
                                    for (const kk in r) {{
                                        if (!Object.prototype.hasOwnProperty.call(r, kk)) continue;
                                        const vv = r[kk];
                                        if (vv === null || vv === undefined) continue;
                                        if (typeof vv === 'string' || typeof vv === 'number') blob += ' ' + String(vv);
                                    }}
                                }} catch (e) {{}}
                                const txt = normAlphaNum(blob);
                                if (txt && qn && txt.indexOf(qn) !== -1) c += 1;
                            }}
                            if (c > bestCount) {{ bestCount = c; bestKey = k; }}
                        }});
                    }} catch (e) {{}}
                    if (bestKey && globalDetails[bestKey]) {{
                        const det = globalDetails[bestKey] || [];
                        const name = det.length ? (det[0]['姓名'] || '') : '';
                        const phone = (/^\d{7,}$/.test(bestKey) ? bestKey : '');
                        let lastDigits = 0, lastDateStr = '';
                        det.forEach(r => {{
                            const d = String(r['下单时间'] || '').replace(/\D/g, '').slice(0,8);
                            if (d && d.length === 8) {{
                                const n = parseInt(d, 10);
                                if (n > lastDigits) {{ lastDigits = n; lastDateStr = (r['下单时间'] || ''); }}
                            }}
                        }});
                        const platCount = {{}};
                        const itemCount = {{}};
                        det.forEach(r => {{
                            const p = (r['下单平台'] || '').trim();
                            if (p) platCount[p] = (platCount[p] || 0) + 1;
                            const it = (r['货品名'] || '').trim();
                            if (it) itemCount[it] = (itemCount[it] || 0) + 1;
                        }});
                        function topKey(map) {{ let best = '', cnt = -1; Object.keys(map).forEach(k => {{ if (map[k] > cnt) {{ cnt = map[k]; best = k; }} }}); return best; }}
                        const topPlat = topKey(platCount);
                        const favItem = topKey(itemCount);
                        const idsSet = new Set();
                        det.forEach(r => {{
                            ['订单号','退货单号'].forEach(kf => {{
                                const v = (r[kf] || '').toString().trim();
                                if (v) {{ idsSet.add(v.toLowerCase()); const dg = v.replace(/\D/g,''); if (dg.length>=6) idsSet.add(dg); }}
                            }});
                        }});
                        const idsAttr = Array.from(idsSet).join('|');
                        const tr = document.createElement('tr');
                    const meta = (globalMeta && globalMeta[bestKey]) ? globalMeta[bestKey] : null;
                    const prClass = meta && meta.priority_class ? meta.priority_class : 'priority-other';
                    tr.className = prClass + ' synthetic-row';
                        tr.setAttribute('data-key', bestKey);
                        tr.setAttribute('data-phone', phone);
                        tr.setAttribute('data-name', name);
                        tr.setAttribute('data-score', meta && meta.priority_score !== undefined ? String(meta.priority_score) : '');
                        tr.setAttribute('data-bucket', meta && meta.priority_bucket ? meta.priority_bucket : '负分');
                        tr.setAttribute('data-platform', topPlat);
                        tr.setAttribute('data-last-order', lastDateStr);
                        tr.setAttribute('data-cycle', '');
                        tr.setAttribute('data-threshold', '');
                        tr.setAttribute('data-category-cycle', '');
                        tr.setAttribute('data-ids', idsAttr);
                        const headers = ['标记完成','优先分','姓名','主要平台','手机号','最近下单日','风险标签','推荐动作','偏好单品','有效订单数','退货率','未复购天数','平均客单价'];
                        const scoreDisp = meta && typeof meta.priority_score === 'number' ? String(Math.round(meta.priority_score)) : '';
                        const ords = meta && typeof meta.orders === 'number' ? String(meta.orders) : String(det.length || '');
                        const rr = meta && typeof meta.return_rate === 'number' ? (meta.return_rate*100).toFixed(1)+'%' : '';
                        const ds = meta && typeof meta.days === 'number' ? String(meta.days) : '';
                        const aov = meta && typeof meta.aov === 'number' ? meta.aov.toFixed(2) : '';
                        const fav = meta && meta.favorite ? meta.favorite : (favItem || '');
                        const values = [scoreDisp, name, topPlat || (meta && meta.platform ? meta.platform : ''), phone || '', lastDateStr || (meta && meta.last_order ? meta.last_order : ''), '全库命中(未纳入触达)', '', fav, ords, rr, ds, aov];
                        const cb = document.createElement('input');
                        cb.type = 'checkbox';
                        cb.className = 'followup-checkbox';
                        cb.setAttribute('data-key', bestKey);
                        cb.setAttribute('data-phone', phone);
                        cb.setAttribute('data-name', name);
                        headers.forEach((h, idx) => {{
                            const td = document.createElement('td');
                            td.setAttribute('data-header', h);
                            if (idx === 0) {{ td.appendChild(cb); }} else {{ td.textContent = values[idx-1] || ''; }}
                            tr.appendChild(td);
                        }});
                        cb.addEventListener('change', () => {{
                            if (cb.checked) {{ followupMap[bestKey] = {{ phone: phone, name: name, date: todayStr, timestamp: new Date().toISOString() }}; setRowState(tr, true); }}
                            else {{ delete followupMap[bestKey]; setRowState(tr, false); }}
                            persistFollowupMap();
                        }});
                        tr.addEventListener('click', (e) => {{ if (panelOpen) return; const isCheckbox = e.target && (e.target.tagName === 'INPUT' || e.target.closest('input')); if (isCheckbox) return; openDetailForKey(bestKey, name); }});
                        tbody.appendChild(tr);
                        addedSynthetic = true;
                    }}
                }}
            }}
        }}

        const totalVisible = visible + (addedSynthetic ? 1 : 0);
        rowCounter.textContent = totalVisible ? ('当前显示 ' + totalVisible + ' 人') : '无匹配客户';

        // 全库命中提示（即使该客户未出现在触达列表）
        const hitBox = document.getElementById('globalHitBox');
        if (hitBox) {{
            hitBox.innerHTML = '';
            if (search) {{
                const key = resolveGlobalKey(search);
                if (key) {{
                    const name = (globalDetails[key] && globalDetails[key].length) ? (globalDetails[key][0]['姓名'] || '') : '';
                    const btn = document.createElement('button');
                    btn.textContent = '查看明细';
                    btn.style.marginLeft = '6px';
                    btn.style.border = '1px solid #ccd5e3';
                    btn.style.background = '#fff';
                    btn.style.borderRadius = '6px';
                    btn.style.padding = '3px 10px';
                    btn.style.cursor = 'pointer';
                    btn.addEventListener('click', () => openDetailForKey(key, name));
                    const span = document.createElement('span');
                    span.textContent = '全库命中：' + (name || key);
                    hitBox.appendChild(span);
                    hitBox.appendChild(btn);
                }} else {{
                    // 全字段模糊匹配（遍历整行文本，规格化后比较，不限于单一列）
                    const qn = normAlphaNum(search);
                    let count = 0;
                    try {{
                        Object.values(globalDetails || {{}}).forEach(list => {{
                            if (!Array.isArray(list)) return;
                            list.forEach(r => {{
                                if (!r || typeof r !== 'object') return;
                                let blob = '';
                                try {{
                                    for (const k in r) {{
                                        if (!Object.prototype.hasOwnProperty.call(r, k)) continue;
                                        const v = r[k];
                                        if (v === null || v === undefined) continue;
                                        if (typeof v === 'string' || typeof v === 'number') blob += ' ' + String(v);
                                    }}
                                }} catch (e) {{}}
                                const txt = normAlphaNum(blob);
                                if (txt && qn && txt.indexOf(qn) !== -1) count += 1;
                            }});
                        }});
                    }} catch (e) {{}}
                    if (count > 0) {{
                        const btn = document.createElement('button');
                        btn.textContent = '查看明细';
                        btn.style.marginLeft = '6px';
                        btn.style.border = '1px solid #ccd5e3';
                        btn.style.background = '#fff';
                        btn.style.borderRadius = '6px';
                        btn.style.padding = '3px 10px';
                        btn.style.cursor = 'pointer';
                        btn.addEventListener('click', () => openDetailForSkuQuery(search));
                        const span = document.createElement('span');
                        span.textContent = `全库货品命中：含"${{search}}"的订单 ${{count}} 条`;
                        hitBox.appendChild(span);
                        hitBox.appendChild(btn);
                    }}
                }}
            }}

            // 全局厂家/平台命中（当没有搜索关键词，但选择了厂家或平台时）
            if (!search && (manufacturer || platform)) {{
                const filterType = manufacturer ? '厂家' : '平台';
                const filterValue = manufacturer || platform;
                let matchCount = 0;
                try {{
                    Object.values(globalDetails || {{}}).forEach(list => {{
                        if (!Array.isArray(list)) return;
                        list.forEach(r => {{
                            if (!r || typeof r !== 'object') return;
                            const val = manufacturer ? (r['厂家'] || '').trim() : (r['下单平台'] || '').trim();
                            if (val === filterValue) matchCount += 1;
                        }});
                    }});
                }} catch (e) {{}}
                if (matchCount > 0) {{
                    const btn = document.createElement('button');
                    btn.textContent = '查看明细';
                    btn.style.marginLeft = '6px';
                    btn.style.border = '1px solid #ccd5e3';
                    btn.style.background = '#fff';
                    btn.style.borderRadius = '6px';
                    btn.style.padding = '3px 10px';
                    btn.style.cursor = 'pointer';
                    btn.addEventListener('click', () => {{
                        if (manufacturer) {{
                            openDetailForManufacturer(filterValue);
                        }} else {{
                            openDetailForPlatform(filterValue);
                        }}
                    }});
                    const span = document.createElement('span');
                    span.textContent = `全库${{filterType}}命中：含"${{filterValue}}"的订单 ${{matchCount}} 条`;
                    hitBox.appendChild(span);
                    hitBox.appendChild(btn);
                }}
            }}
        }}

        // 货品搜索：直接在主表格下方显示订单明细表格
        let productDetailContainer = document.getElementById('productDetailContainer');
        if (!productDetailContainer) {{
            productDetailContainer = document.createElement('div');
            productDetailContainer.id = 'productDetailContainer';
            productDetailContainer.style.marginTop = '24px';
            productDetailContainer.style.display = 'none';

            // Insert after main table
            const tableParent = table ? table.closest('.card, .container, div') : null;
            if (tableParent && tableParent.parentElement) {{
                tableParent.parentElement.insertBefore(productDetailContainer, tableParent.nextSibling);
                console.log('Created productDetailContainer after table parent');
            }} else {{
                if (table && table.parentElement) {{
                    table.parentElement.insertAdjacentElement('afterend', productDetailContainer);
                    console.log('Created productDetailContainer after table');
                }} else {{
                    document.body.appendChild(productDetailContainer);
                    console.log('Created productDetailContainer in body');
                }}
            }}
        }}

        // Use global searchRaw
        if (searchRaw && searchRaw.length >= 2 && productSearchIndex && typeof productSearchIndex === 'object') {{
            const normalized = normAlphaNum(searchRaw);
            let isProductSearch = false;

            // Check if matches product name
            if (normalized && productSearchIndex[normalized]) {{
                isProductSearch = true;
                console.log('Exact match found in productSearchIndex');
            }} else if (normalized) {{
                for (const productKey of Object.keys(productSearchIndex)) {{
                    if (productKey.includes(normalized)) {{
                        isProductSearch = true;
                        console.log('Fuzzy match found:', productKey);
                        break;
                    }}
                }}
            }}

            if (isProductSearch) {{
                const qn = normalized;

                // ✅ 优化 2: 检查缓存
                let matchedCustomers = getCachedSearchResult(normalized);

                if (matchedCustomers) {{
                    // 从缓存读取，跳过搜索逻辑
                    console.log(`Using cached results for "${{normalized}}"`);
                }} else {{
                    // 缓存未命中，执行搜索
                    matchedCustomers = {{}};

                    // ✅ 优化 1: 使用索引快速查找
                    const searchStart = performance.now();
                    const customerPhones = new Set();

                    try {{
                        // 精确匹配（O(1) 查找）
                        if (productSearchIndex[normalized]) {{
                            productSearchIndex[normalized].forEach(p => customerPhones.add(p));
                            console.log(`Exact match found ${{customerPhones.size}} customers for "${{normalized}}"`);
                        }}

                        // 模糊匹配（仅在精确匹配为空时）
                        if (customerPhones.size === 0) {{
                            console.log(`No exact match, trying fuzzy search for "${{normalized}}"`);
                            for (const [productKey, phones] of Object.entries(productSearchIndex)) {{
                                if (productKey.includes(normalized)) {{
                                    phones.forEach(p => customerPhones.add(p));
                                }}
                            }}
                            console.log(`Fuzzy search found ${{customerPhones.size}} customers`);
                        }}

                        // 只处理匹配的客户（O(n) where n = matched customers）
                        customerPhones.forEach(phone => {{
                            const orders = globalDetails[phone];
                            if (!orders || !Array.isArray(orders)) return;

                            const matchingOrders = orders.filter(r => {{
                                if (!r || typeof r !== 'object') return false;
                                const productName = normAlphaNum(r['货品名'] || r['商品名称'] || '');
                                return productName && productName.includes(qn);
                            }});

                            if (matchingOrders.length > 0) {{
                                matchedCustomers[phone] = {{
                                    phone,
                                    allOrders: orders,
                                    matchingOrders,
                                    matchCount: matchingOrders.length
                                }};
                            }}
                        }});

                        const searchTime = performance.now() - searchStart;
                        console.log(`Search completed in ${{searchTime.toFixed(2)}}ms`);
                        console.log(`Processed ${{customerPhones.size}} customers, found ${{Object.keys(matchedCustomers).length}} matches`);

                    }} catch (e) {{
                        console.error('Error collecting customers:', e);
                        // 降级方案：回退到原始遍历逻辑
                        console.warn('Falling back to full scan...');
                        Object.entries(globalDetails || {{}}).forEach(([phone, orders]) => {{
                            if (!Array.isArray(orders)) return;
                            const matchingOrders = orders.filter(r => {{
                                if (!r || typeof r !== 'object') return false;
                                const productName = normAlphaNum(r['货品名'] || r['商品名称'] || '');
                                return productName && productName.includes(qn);
                            }});
                            if (matchingOrders.length > 0) {{
                                matchedCustomers[phone] = {{
                                    phone,
                                    allOrders: orders,
                                    matchingOrders,
                                    matchCount: matchingOrders.length
                                }};
                            }}
                        }});
                    }}

                    // 搜索完成后缓存结果
                    cacheSearchResult(normalized, matchedCustomers);
                }}

                console.log('Matched customers count:', Object.keys(matchedCustomers).length);

                // 将客户数据转为数组并聚合信息
                const customerList = Object.values(matchedCustomers).map(({{ phone, allOrders, matchingOrders, matchCount }}) => {{
                    // 从第一条订单提取客户信息
                    const firstOrder = allOrders[0] || {{}};
                    const name = firstOrder['姓名'] || '';

                    // 最近下单时间（该产品的最新订单）
                    const sortedMatchingOrders = matchingOrders.slice().sort((a, b) => {{
                        const ad = String(a['下单时间'] || '').replace(/\D/g, '').slice(0,8);
                        const bd = String(b['下单时间'] || '').replace(/\D/g, '').slice(0,8);
                        return parseInt(bd, 10) - parseInt(ad, 10);
                    }});
                    const lastOrderDate = sortedMatchingOrders[0]?.['下单时间'] || '';

                    // 主要平台统计
                    const platformCount = {{}};
                    matchingOrders.forEach(r => {{
                        const p = (r['下单平台'] || '').trim();
                        if (p) platformCount[p] = (platformCount[p] || 0) + 1;
                    }});
                    const mainPlatform = Object.keys(platformCount).sort((a, b) => platformCount[b] - platformCount[a])[0] || '';

                    // 手机号脱敏
                    const maskedPhone = phone.length > 7
                        ? phone.slice(0, 3) + '****' + phone.slice(-4)
                        : phone;

                    return {{
                        phone,           // 完整手机号（用于 data-key）
                        name,
                        maskedPhone,     // 脱敏手机号（显示用）
                        matchCount,      // 购买该产品次数
                        totalOrders: allOrders.length,  // 总订单数
                        lastOrderDate,
                        mainPlatform
                    }};
                }});

                // 默认按总订单数降序排序（高价值客户排在前面）
                customerList.sort((a, b) => b.totalOrders - a.totalOrders);

                const totalMatchedOrders = customerList.reduce((sum, c) => sum + c.matchCount, 0);

                // 排序状态跟踪
                let currentSort = {{ field: 'totalOrders', ascending: false }};

                // 渲染表格函数（支持动态排序）
                function renderTable() {{
                    if (customerList.length > 0) {{
                        productDetailContainer.innerHTML = `
                            <div style="background: white; border: 1px solid #e2e8f0; border-radius: 12px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.05);">
                                <h3 style="margin: 0 0 12px 0; font-size: 16px; font-weight: 600; color: #1e293b;">
                                    购买"${{escapeHtml(searchRaw)}}"的客户（${{customerList.length}} 人，共${{totalMatchedOrders}}件）
                                </h3>
                                <p style="margin: 0 0 16px 0; font-size: 13px; color: #64748b;">
                                    💡 点击客户行查看完整订单历史（含其他商品）· 点击表头可排序
                                </p>
                                <div style="overflow-x: auto; border: 1px solid #e2e8f0; border-radius: 8px;">
                                    <table class="customer-table" style="width: 100%; border-collapse: collapse;">
                                        <thead>
                                            <tr style="background: #f8fafc;">
                                                <th style="padding: 10px; text-align: left; border-bottom: 1px solid #e2e8f0; font-weight: 600; color: #475569;">姓名</th>
                                                <th style="padding: 10px; text-align: left; border-bottom: 1px solid #e2e8f0; font-weight: 600; color: #475569;">手机号</th>
                                                <th data-sort="totalOrders" style="padding: 10px; text-align: center; border-bottom: 1px solid #e2e8f0; font-weight: 600; color: #475569; cursor: pointer; user-select: none; transition: background-color 0.15s;" onmouseover="this.style.backgroundColor='#e2e8f0'" onmouseout="this.style.backgroundColor='transparent'">
                                                    总订单数 ${{currentSort.field === 'totalOrders' ? (currentSort.ascending ? '↑' : '↓') : '⇅'}}
                                                </th>
                                                <th style="padding: 10px; text-align: left; border-bottom: 1px solid #e2e8f0; font-weight: 600; color: #475569;">最近下单</th>
                                                <th style="padding: 10px; text-align: left; border-bottom: 1px solid #e2e8f0; font-weight: 600; color: #475569;">主要平台</th>
                                            </tr>
                                        </thead>
                                        <tbody>
                                            ${{customerList.map(c => `
                                                <tr
                                                    class="customer-row"
                                                    data-key="${{escapeHtml(c.phone)}}"
                                                    data-name="${{escapeHtml(c.name)}}"
                                                    data-total-orders="${{c.totalOrders}}"
                                                    style="
                                                        border-bottom: 1px solid #f1f5f9;
                                                        cursor: pointer;
                                                        transition: background-color 0.15s;
                                                    "
                                                    onmouseover="this.style.backgroundColor='#f8fafc'"
                                                    onmouseout="this.style.backgroundColor='white'"
                                                >
                                                    <td style="padding: 8px; color: #334155; font-weight: 500;">${{escapeHtml(c.name)}}</td>
                                                    <td style="padding: 8px; color: #64748b; font-family: monospace; font-size: 12px;">${{escapeHtml(c.maskedPhone)}}</td>
                                                    <td style="padding: 8px; text-align: center; color: #0ea5e9; font-weight: 600;">${{c.totalOrders}}单</td>
                                                    <td style="padding: 8px; color: #334155;">${{escapeHtml(c.lastOrderDate)}}</td>
                                                    <td style="padding: 8px; color: #334155;">${{escapeHtml(c.mainPlatform)}}</td>
                                                </tr>
                                            `).join('')}}
                                        </tbody>
                                    </table>
                                </div>
                            </div>
                        `;

                        productDetailContainer.style.display = 'block';

                        // ✅ 绑定点击事件：调用现有的 openDetailForKey 函数
                        productDetailContainer.querySelectorAll('.customer-row').forEach(row => {{
                            row.addEventListener('click', () => {{
                                const phone = row.getAttribute('data-key');
                                const name = row.getAttribute('data-name');
                                if (phone && name) {{
                                    openDetailForKey(phone, name);
                                }}
                            }});
                        }});

                        // ✅ 绑定表头排序事件
                        productDetailContainer.querySelectorAll('th[data-sort]').forEach(th => {{
                            th.addEventListener('click', () => {{
                                const field = th.getAttribute('data-sort');

                                // 如果点击同一列，切换升序/降序；否则默认降序
                                if (currentSort.field === field) {{
                                    currentSort.ascending = !currentSort.ascending;
                                }} else {{
                                    currentSort.field = field;
                                    currentSort.ascending = false;
                                }}

                                // 执行排序
                                customerList.sort((a, b) => {{
                                    const aVal = a[field];
                                    const bVal = b[field];
                                    return currentSort.ascending ? (aVal - bVal) : (bVal - aVal);
                                }});

                                // 重新渲染
                                renderTable();
                            }});
                        }});

                        console.log('Customer table rendered with', customerList.length, 'customers');
                    }} else {{
                        productDetailContainer.style.display = 'none';
                    }}
                }}

                // 初始渲染
                renderTable();
            }} else {{
                productDetailContainer.style.display = 'none';
            }}
        }} else {{
            productDetailContainer.style.display = 'none';
        }}
    }}

    // ✅ 优化 3: 输入防抖（300ms）
    let searchTimeout;
    const DEBOUNCE_DELAY = 300;

    function debouncedApplyFilters() {{
        clearTimeout(searchTimeout);
        searchTimeout = setTimeout(() => {{
            applyFilters();
        }}, DEBOUNCE_DELAY);
    }}

    [searchBox, priorityFilter, tagFilter, platformFilter, manufacturerFilter].forEach(el => {{
        if (!el) {{
            return;
        }}
        // 对搜索框使用防抖
        if (el === searchBox) {{
            el.addEventListener('input', debouncedApplyFilters);
        }} else {{
            el.addEventListener('input', applyFilters);
        }}
        el.addEventListener('change', applyFilters);
    }});
    
    // 列表tab切换
    const listTabs = document.querySelectorAll('.list-tab');
    listTabs.forEach(tab => {{
        tab.addEventListener('click', () => {{
            // 移除所有active
            listTabs.forEach(t => {{
                t.classList.remove('active');
                t.style.borderBottom = '3px solid transparent';
                t.style.color = '#666';
            }});
            // 添加当前active
            tab.classList.add('active');
            tab.style.borderBottom = '3px solid #1890ff';
            tab.style.color = '#1890ff';
            // 更新当前列表
            currentList = tab.dataset.list;
            // 重新过滤
            applyFilters();
        }});
    }});

    function exportCsv() {{
        // 只收集当天的联系记录
        const todayEntries = {{}};

        // 读取当天的 localStorage 记录
        try {{
            const todayKey = `followup-${{todayStr}}`;
            const storedData = localStorage.getItem(todayKey);
            if (storedData) {{
                const dayMap = JSON.parse(storedData) || {{}};
                Object.entries(dayMap).forEach(([phoneKey, entry]) => {{
                    if (entry && entry.phone) {{
                        todayEntries[entry.phone] = {{
                            phone: entry.phone,
                            date: entry.date,
                            name: entry.name,
                            platform: entry.platform || '',
                            owner: entry.owner || '',
                            note: entry.note || ''
                        }};
                    }}
                }});
            }}
        }} catch (err) {{
            console.warn('读取今日联系记录失败:', err);
        }}

        // 确保当天内存中的记录也被包含
        Object.values(followupMap).forEach(entry => {{
            if (entry && entry.phone) {{
                todayEntries[entry.phone] = {{
                    phone: entry.phone,
                    date: entry.date,
                    name: entry.name,
                    platform: entry.platform || '',
                    owner: entry.owner || '',
                    note: entry.note || ''
                }};
            }}
        }});

        const entries = Object.values(todayEntries);
        if (!entries.length) {{
            alert('当前没有可导出的联系记录。');
            return;
        }}

        // 按日期降序排序
        entries.sort((a, b) => {{
            if (a.date > b.date) return -1;
            if (a.date < b.date) return 1;
            return 0;
        }});

        const header = '手机号,姓名,联系平台,最近联系日期,跟进人';
        const body = entries
            .map(entry => `${{entry.phone}},${{entry.name||''}},${{entry.platform||''}},${{entry.date}},${{entry.owner||''}}`)
            .join('\\n');
        const csvContent = `${{header}}\\n${{body}}`;
        const blob = new Blob([csvContent], {{ type: 'text/csv;charset=utf-8;' }});
        const url = URL.createObjectURL(blob);
        const link = document.createElement('a');
        link.href = url;
        link.download = `contact_log_${{todayStr}}.csv`;
        document.body.appendChild(link);
        link.click();
        document.body.removeChild(link);
        URL.revokeObjectURL(url);

        // 显示导出统计
        alert(`成功导出 ${{entries.length}} 条今日联系记录`);
    }}

    function clearMarks() {{
        if (!Object.keys(followupMap).length) {{
            return;
        }}
        if (!confirm('确定要清除今日的跟进标记吗？')) {{
            return;
        }}
        followupMap = {{}};
        checkboxes.forEach(checkbox => {{
            checkbox.checked = false;
            setRowState(checkbox.closest('tr'), false);
        }});
        persistFollowupMap();
    }}

    if (exportBtn) {{
        exportBtn.addEventListener('click', exportCsv);
    }}
    if (clearBtn) {{
        clearBtn.addEventListener('click', clearMarks);
    }}

    applyFilters();
    updateSummary();

    // Drilldown drawer: list all orders for the clicked customer
    const detailBackdrop = document.createElement('div');
    detailBackdrop.id = 'detailBackdrop';
    detailBackdrop.className = 'detail-backdrop hidden';
    document.body.appendChild(detailBackdrop);

    const panel = document.createElement('div');
    panel.id = 'detailPanel';
    panel.className = 'detail-panel hidden';
    panel.innerHTML = `
      <div class="detail-header">
        <h3 id="detailTitle">订单明细</h3>
        <button id="detailClose" class="detail-close">关闭</button>
      </div>
      <div class="detail-body">
        <table id="detailTable">
          <thead>
            <tr>
              <th data-sort-method="number">下单时间</th>
              <th>客户信息</th>
              <th>货品信息</th>
              <th data-sort-method="number">金额/毛利</th>
              <th>订单/退货号</th>
              <th>售后状态</th>
            </tr>
          </thead>
          <tbody id="detailTbody"></tbody>
        </table>
      </div>
    `;
    document.body.appendChild(panel);

    const detailTitle = panel.querySelector('#detailTitle');
    const detailClose = panel.querySelector('#detailClose');
    const detailTbody = panel.querySelector('#detailTbody');
    const detailTable = panel.querySelector('#detailTable');
    let panelOpen = false;

    function formatAmount(n) {{
      const v = (typeof n === 'number') ? n : parseFloat(n);
      if (!isFinite(v)) return '0.00';
      return v.toFixed(2);
    }}
    
    function calcMargin(payment, cost) {{
      const pay = (typeof payment === 'number') ? payment : parseFloat(payment);
      const cst = (typeof cost === 'number') ? cost : parseFloat(cost);
      if (!isFinite(pay) || pay <= 0) return '-';
      if (!isFinite(cst)) return '-';
      const margin = ((pay - cst) / pay) * 100;
      return margin.toFixed(1) + '%';
    }}

    // HTML escape function to prevent HTML breaking
    function escapeHtml(str) {{
      if (!str) return '';
      return String(str)
        .replace(/&/g, '&amp;')
        .replace(/</g, '&lt;')
        .replace(/>/g, '&gt;')
        .replace(/"/g, '&quot;')
        .replace(/'/g, '&#039;');
    }}

    function openDetailForKey(key, name) {{
      // Hide Manufacturer specific panels
      var diagPanel = document.getElementById('mfrDiagnosis');
      if (diagPanel) diagPanel.remove();
      var aiPanel = document.getElementById('mfrAiAnalysis');
      if (aiPanel) aiPanel.remove();

      let rows = (detailMap && detailMap[key]) ? detailMap[key] : [];
      if (!rows.length && globalDetails && globalDetails[key]) {{
        rows = globalDetails[key];
      }}
      // 默认按下单时间降序（最近在上）
      const orderedRows = rows.slice().sort((a, b) => {{
        const ad = String(a['下单时间'] || '').replace(/\D/g, '').slice(0,8);
        const bd = String(b['下单时间'] || '').replace(/\D/g, '').slice(0,8);
        const an = ad ? parseInt(ad, 10) : 0;
        const bn = bd ? parseInt(bd, 10) : 0;
        return bn - an;
      }});
      detailTitle.textContent = (name ? name + ' - ' : '') + '订单明细（' + rows.length + ' 条）';
      detailTbody.innerHTML = '';
      if (!rows.length) {{
        const tr = document.createElement('tr');
        const td = document.createElement('td');
        td.colSpan = 15;
        td.textContent = '暂无数据';
        tr.appendChild(td);
        detailTbody.appendChild(tr);
      }} else {{
        const frag = document.createDocumentFragment();
        orderedRows.forEach(r => {{
          const tr = document.createElement('tr');
          const paymentAmt = r['付款金额'];
          const costAmt = r['打款金额'];
          const cols = [
            (r['下单时间'] || ''),
            (r['姓名'] || '') + '<br><span style="color:#64748b;font-size:12px">' + (r['手机号'] || '') + '</span>',
            (r['商品名称'] || r['货品名'] || '') + '<br><span style="color:#64748b;font-size:12px">' + (r['颜色'] || '') + ' ' + (r['尺码'] || '') + '</span>',
            '<div style="display:flex; flex-direction:column; align-items:flex-end;"><span>' + formatAmount(paymentAmt) + ' / ' + formatAmount(costAmt) + '</span><span style="font-weight:bold; color:' + (parseFloat(calcMargin(paymentAmt, costAmt)) < 15 ? '#f59e0b' : '#10b981') + '">' + calcMargin(paymentAmt, costAmt) + '</span></div>',
            (r['订单号'] || '') + (r['退货单号'] ? '<br><span style="color:#ef4444; font-size:12px">退: ' + r['退货单号'] + '</span>' : ''),
            (r['退款类型'] || '') + (r['退款原因'] ? '<br><span style="color:#64748b; font-size:12px">' + r['退款原因'] + '</span>' : '')
          ];
          cols.forEach((html, i) => {{
            const td = document.createElement('td');
            if (i === 0) {{
              const digits = String(html).replace(/\D/g, '');
              if (digits.length >= 8) {{ td.setAttribute('data-sort-value', digits.slice(0,8)); }}
            }} else if (i === 3) {{
              const mNum = parseFloat(calcMargin(paymentAmt, costAmt));
              if (isFinite(mNum)) td.setAttribute('data-sort-value', String(mNum));
              td.style.textAlign = 'right';
            }}
            td.innerHTML = html;
            tr.appendChild(td);
          }});
          frag.appendChild(tr);
        }});
        detailTbody.appendChild(frag);
      }}
      // Initialize or refresh sorter for detail table
      try {{ new Tablesort(detailTable); }} catch (e) {{}}
      detailBackdrop.classList.remove('hidden');
      panel.classList.remove('hidden');
      panelOpen = true;
    }}

    function closeDetail() {{
      panel.classList.add('hidden');
      detailBackdrop.classList.add('hidden');
      panelOpen = false;
    }}

    detailBackdrop.addEventListener('click', closeDetail);
    detailClose.addEventListener('click', closeDetail);

    // Row click binding (ignore direct checkbox clicks)
    if (table) {{
      table.querySelectorAll('tbody tr').forEach(tr => {{
        tr.addEventListener('click', (e) => {{
          if (panelOpen) return;
          const isCheckbox = e.target && (e.target.tagName === 'INPUT' || e.target.closest('input'));
          if (isCheckbox) return;
          const key = tr.getAttribute('data-key');
          const name = tr.getAttribute('data-name');
          if (!key) return;
          openDetailForKey(key, name);
        }});
      }});
    }}
    // ESC key closes drawer
    document.addEventListener('keydown', (e) => {{
      if (e.key === 'Escape' && panelOpen) {{
        e.preventDefault();
        closeDetail();
      }}
    }});

    // SKU 明细抽屉（按货品名汇总展示全量订单）
    // filterType: 'all' | 'normal' | 'proxy'
    function openDetailForSku(sku, filterType) {{
      // Hide Manufacturer specific panels
      var diagPanel = document.getElementById('mfrDiagnosis');
      if (diagPanel) diagPanel.remove();
      var aiPanel = document.getElementById('mfrAiAnalysis');
      if (aiPanel) aiPanel.remove();

      const rows = [];
      try {{
        Object.values(globalDetails || {{}}).forEach(list => {{
          if (!Array.isArray(list)) return;
          list.forEach(r => {{
            if ((r && typeof r === 'object') && String(r['货品名']||'').trim() === sku) {{
              // 根据filterType过滤订单
              if (filterType && filterType !== 'all') {{
                const platform = String(r['下单平台'] || '');
                const prod = String(r['商品名称'] || '');
                const rmk = String(r['备注'] || '');
                const combined = platform + ' ' + prod + ' ' + sku + ' ' + rmk;
                const isProxy = combined.includes('代发');
                
                if (filterType === 'normal' && isProxy) return;  // 只要非代发，跳过代发订单
                if (filterType === 'proxy' && !isProxy) return;  // 只要代发，跳过非代发订单
              }}
              rows.push(r);
            }}
          }});
        }});
      }} catch (e) {{}}
      const orderedRows = rows.slice().sort((a, b) => {{
        const ad = String(a['下单时间'] || '').replace(/\D/g, '').slice(0,8);
        const bd = String(b['下单时间'] || '').replace(/\D/g, '').slice(0,8);
        const an = ad ? parseInt(ad, 10) : 0;
        const bn = bd ? parseInt(bd, 10) : 0;
        return bn - an;
      }});
      const filterLabel = filterType === 'proxy' ? '（仅代发）' : filterType === 'normal' ? '（不含代发）' : '';
      detailTitle.textContent = sku + ' - 订单明细' + filterLabel + '（' + rows.length + ' 条）';
      detailTbody.innerHTML = '';
      if (!rows.length) {{
        const tr = document.createElement('tr');
        const td = document.createElement('td');
        td.colSpan = 15;
        td.textContent = '暂无数据';
        tr.appendChild(td);
        detailTbody.appendChild(tr);
      }} else {{
        const frag = document.createDocumentFragment();
        orderedRows.forEach(r => {{
          const tr = document.createElement('tr');
          const paymentAmt = r['付款金额'];
          const costAmt = r['打款金额'];
          const cols = [
            (r['下单时间'] || ''),
            (r['姓名'] || '') + '<br><span style="color:#64748b;font-size:12px">' + (r['手机号'] || '') + '</span>',
            (r['商品名称'] || r['货品名'] || '') + '<br><span style="color:#64748b;font-size:12px">' + (r['颜色'] || '') + ' ' + (r['尺码'] || '') + '</span>',
            '<div style="display:flex; flex-direction:column; align-items:flex-end;"><span>' + formatAmount(paymentAmt) + ' / ' + formatAmount(costAmt) + '</span><span style="font-weight:bold; color:' + (parseFloat(calcMargin(paymentAmt, costAmt)) < 15 ? '#f59e0b' : '#10b981') + '">' + calcMargin(paymentAmt, costAmt) + '</span></div>',
            (r['订单号'] || '') + (r['退货单号'] ? '<br><span style="color:#ef4444; font-size:12px">退: ' + r['退货单号'] + '</span>' : ''),
            (r['退款类型'] || '') + (r['退款原因'] ? '<br><span style="color:#64748b; font-size:12px">' + r['退款原因'] + '</span>' : '')
          ];
          cols.forEach((html, i) => {{
            const td = document.createElement('td');
            if (i === 0) {{
              const digits = String(html).replace(/\D/g, '');
              if (digits.length >= 8) {{ td.setAttribute('data-sort-value', digits.slice(0,8)); }}
            }} else if (i === 3) {{
              const mNum = parseFloat(calcMargin(paymentAmt, costAmt));
              if (isFinite(mNum)) td.setAttribute('data-sort-value', String(mNum));
              td.style.textAlign = 'right';
            }}
            td.innerHTML = html;
            tr.appendChild(td);
          }});
          frag.appendChild(tr);
        }});
        detailTbody.appendChild(frag);
      }}
      try {{ new Tablesort(detailTable); }} catch (e) {{}}
      detailBackdrop.classList.remove('hidden');
      panel.classList.remove('hidden');
      panelOpen = true;
    }}

    // 货品名“包含”查询（全库模糊匹配）
    function openDetailForSkuQuery(queryRaw) {{
      // Hide Manufacturer specific panels
      var diagPanel = document.getElementById('mfrDiagnosis');
      if (diagPanel) diagPanel.remove();
      var aiPanel = document.getElementById('mfrAiAnalysis');
      if (aiPanel) aiPanel.remove();

      const qn = normAlphaNum(queryRaw || '');
      if (!qn) return;
      const rows = [];
      try {{
        Object.values(globalDetails || {{}}).forEach(list => {{
          if (!Array.isArray(list)) return;
          list.forEach(r => {{
            if (!r || typeof r !== 'object') return;
            let blob = '';
            try {{
              for (const k in r) {{
                if (!Object.prototype.hasOwnProperty.call(r, k)) continue;
                const v = r[k];
                if (v === null || v === undefined) continue;
                if (typeof v === 'string' || typeof v === 'number') blob += ' ' + String(v);
              }}
            }} catch (e) {{}}
            const txt = normAlphaNum(blob);
            if (txt && qn && txt.indexOf(qn) !== -1) rows.push(r);
          }});
        }});
      }} catch (e) {{}}
      const orderedRows = rows.slice().sort((a, b) => {{
        const ad = String(a['下单时间'] || '').replace(/\D/g, '').slice(0,8);
        const bd = String(b['下单时间'] || '').replace(/\D/g, '').slice(0,8);
        const an = ad ? parseInt(ad, 10) : 0;
        const bn = bd ? parseInt(bd, 10) : 0;
        return bn - an;
      }});
      detailTitle.textContent = '包含“' + (queryRaw || '') + '”的货品 - 订单明细（' + rows.length + ' 条）';
      detailTbody.innerHTML = '';
      if (!rows.length) {{
        const tr = document.createElement('tr');
        const td = document.createElement('td');
        td.colSpan = 15;
        td.textContent = '暂无数据';
        tr.appendChild(td);
        detailTbody.appendChild(tr);
      }} else {{
        const frag = document.createDocumentFragment();
        orderedRows.forEach(r => {{
          const tr = document.createElement('tr');
          const paymentAmt = r['付款金额'];
          const costAmt = r['打款金额'];
          const cols = [
            r['姓名'] || '',
            r['手机号'] || '',
            (r['下单时间'] || ''),
            r['下单平台'] || '',
            r['厂家'] || '',
            (r['商品名称'] || r['货品名'] || ''),
            r['颜色'] || r['色号'] || '',
            r['尺码'] || r['规格'] || r['码数'] || '',
            formatAmount(paymentAmt),
            formatAmount(costAmt),
            calcMargin(paymentAmt, costAmt),
            r['订单号'] || '',
            r['退货单号'] || '',
            r['退款类型'] || '',
            r['退款原因'] || ''
          ];
          cols.forEach((text, i) => {{
            const td = document.createElement('td');
            if (i === 2) {{
              const digits = String(text).replace(/\D/g, '');
              if (digits.length >= 8) {{ td.setAttribute('data-sort-value', digits.slice(0,8)); }}
            }} else if (i === 8 || i === 9) {{
              const val = parseFloat(text);
              if (isFinite(val)) td.setAttribute('data-sort-value', String(val));
            }} else if (i === 10) {{ // 毛利率 numeric sort key
              const marginText = String(text).replace('%', '');
              const val = parseFloat(marginText);
              if (isFinite(val)) td.setAttribute('data-sort-value', String(val));
            }}
            td.textContent = text;
            tr.appendChild(td);
          }});
          frag.appendChild(tr);
        }});
        detailTbody.appendChild(frag);
      }}
      try {{ new Tablesort(detailTable); }} catch (e) {{}}
      detailBackdrop.classList.remove('hidden');
      panel.classList.remove('hidden');
      panelOpen = true;
    }}

    function openDetailForManufacturer(manufacturer) {{
      // Hide Manufacturer specific panels
      var diagPanel = document.getElementById('mfrDiagnosis');
      if (diagPanel) diagPanel.remove();
      var aiPanel = document.getElementById('mfrAiAnalysis');
      if (aiPanel) aiPanel.remove();

      if (!manufacturer) return;
      const rows = [];
      try {{
        Object.values(globalDetails || {{}}).forEach(list => {{
          if (!Array.isArray(list)) return;
          list.forEach(r => {{
            if (!r || typeof r !== 'object') return;
            const mfr = (r['厂家'] || '').trim();
            if (mfr === manufacturer) rows.push(r);
          }});
        }});
      }} catch (e) {{}}
      const orderedRows = rows.slice().sort((a, b) => {{
        const ad = String(a['下单时间'] || '').replace(/\D/g, '').slice(0,8);
        const bd = String(b['下单时间'] || '').replace(/\D/g, '').slice(0,8);
        const an = ad ? parseInt(ad, 10) : 0;
        const bn = bd ? parseInt(bd, 10) : 0;
        return bn - an;
      }});
      detailTitle.textContent = '厂家"' + manufacturer + '" - 订单明细（' + rows.length + ' 条）';
      detailTbody.innerHTML = '';
      if (!rows.length) {{
        const tr = document.createElement('tr');
        const td = document.createElement('td');
        td.colSpan = 15;
        td.textContent = '暂无数据';
        tr.appendChild(td);
        detailTbody.appendChild(tr);
      }} else {{
        const frag = document.createDocumentFragment();
        orderedRows.forEach(r => {{
          const tr = document.createElement('tr');
          const paymentAmt = r['付款金额'];
          const costAmt = r['打款金额'];
          const cols = [
            r['姓名'] || '',
            r['手机号'] || '',
            (r['下单时间'] || ''),
            r['下单平台'] || '',
            r['厂家'] || '',
            (r['商品名称'] || r['货品名'] || ''),
            r['颜色'] || r['色号'] || '',
            r['尺码'] || r['规格'] || r['码数'] || '',
            formatAmount(paymentAmt),
            formatAmount(costAmt),
            calcMargin(paymentAmt, costAmt),
            r['订单号'] || '',
            r['退货单号'] || '',
            r['退款类型'] || '',
            r['退款原因'] || ''
          ];
          cols.forEach((text, i) => {{
            const td = document.createElement('td');
            if (i === 2) {{
              const digits = String(text).replace(/\D/g, '');
              if (digits.length >= 8) {{ td.setAttribute('data-sort-value', digits.slice(0,8)); }}
            }} else if (i === 8 || i === 9) {{
              const val = parseFloat(text);
              if (isFinite(val)) td.setAttribute('data-sort-value', String(val));
            }} else if (i === 10) {{
              const marginText = String(text).replace('%', '');
              const val = parseFloat(marginText);
              if (isFinite(val)) td.setAttribute('data-sort-value', String(val));
            }}
            td.textContent = text;
            tr.appendChild(td);
          }});
          frag.appendChild(tr);
        }});
        detailTbody.appendChild(frag);
      }}
      try {{ new Tablesort(detailTable); }} catch (e) {{}}
      detailBackdrop.classList.remove('hidden');
      panel.classList.remove('hidden');
      panelOpen = true;
    }}

    function openDetailForPlatform(platform) {{
      // Hide Manufacturer specific panels
      var diagPanel = document.getElementById('mfrDiagnosis');
      if (diagPanel) diagPanel.remove();
      var aiPanel = document.getElementById('mfrAiAnalysis');
      if (aiPanel) aiPanel.remove();

      if (!platform) return;
      const rows = [];
      try {{
        Object.values(globalDetails || {{}}).forEach(list => {{
          if (!Array.isArray(list)) return;
          list.forEach(r => {{
            if (!r || typeof r !== 'object') return;
            const plat = (r['下单平台'] || '').trim();
            if (plat === platform) rows.push(r);
          }});
        }});
      }} catch (e) {{}}
      const orderedRows = rows.slice().sort((a, b) => {{
        const ad = String(a['下单时间'] || '').replace(/\D/g, '').slice(0,8);
        const bd = String(b['下单时间'] || '').replace(/\D/g, '').slice(0,8);
        const an = ad ? parseInt(ad, 10) : 0;
        const bn = bd ? parseInt(bd, 10) : 0;
        return bn - an;
      }});
      detailTitle.textContent = '平台"' + platform + '" - 订单明细（' + rows.length + ' 条）';
      detailTbody.innerHTML = '';
      if (!rows.length) {{
        const tr = document.createElement('tr');
        const td = document.createElement('td');
        td.colSpan = 15;
        td.textContent = '暂无数据';
        tr.appendChild(td);
        detailTbody.appendChild(tr);
      }} else {{
        const frag = document.createDocumentFragment();
        orderedRows.forEach(r => {{
          const tr = document.createElement('tr');
          const paymentAmt = r['付款金额'];
          const costAmt = r['打款金额'];
          const cols = [
            r['姓名'] || '',
            r['手机号'] || '',
            (r['下单时间'] || ''),
            r['下单平台'] || '',
            r['厂家'] || '',
            (r['商品名称'] || r['货品名'] || ''),
            r['颜色'] || r['色号'] || '',
            r['尺码'] || r['规格'] || r['码数'] || '',
            formatAmount(paymentAmt),
            formatAmount(costAmt),
            calcMargin(paymentAmt, costAmt),
            r['订单号'] || '',
            r['退货单号'] || '',
            r['退款类型'] || '',
            r['退款原因'] || ''
          ];
          cols.forEach((text, i) => {{
            const td = document.createElement('td');
            if (i === 2) {{
              const digits = String(text).replace(/\D/g, '');
              if (digits.length >= 8) {{ td.setAttribute('data-sort-value', digits.slice(0,8)); }}
            }} else if (i === 8 || i === 9) {{
              const val = parseFloat(text);
              if (isFinite(val)) td.setAttribute('data-sort-value', String(val));
            }} else if (i === 10) {{
              const marginText = String(text).replace('%', '');
              const val = parseFloat(marginText);
              if (isFinite(val)) td.setAttribute('data-sort-value', String(val));
            }}
            td.textContent = text;
            tr.appendChild(td);
          }});
          frag.appendChild(tr);
        }});
        detailTbody.appendChild(frag);
      }}
      try {{ new Tablesort(detailTable); }} catch (e) {{}}
      detailBackdrop.classList.remove('hidden');
      panel.classList.remove('hidden');
      panelOpen = true;
    }}

    // 绑定"加推SKU"表格事件
    try {{
        var tableStatic = document.getElementById('skuPushTableStatic');
        if (tableStatic) {{
            if (typeof Tablesort !== 'undefined') {{
                try {{ new Tablesort(tableStatic); }} catch(e) {{}}
            }}
            var rowsStatic = tableStatic.querySelectorAll('tbody tr');
            for (var i = 0; i < rowsStatic.length; i++) {{
                (function(row) {{
                    row.addEventListener('click', function() {{
                        var sku = row.getAttribute('data-sku');
                        if (sku && typeof openDetailForSku === 'function') {{
                            openDetailForSku(sku, 'all');
                        }}
                    }});
                    row.style.cursor = 'pointer';
                }})(rowsStatic[i]);
            }}
        }}
    }} catch (e) {{ console.error('Push SKU Bind Error:', e); }}

    // 渲染“高退货预警”（明细>3，退货率>30%）
    try {{
      const container = document.getElementById('skuReturnAlertTable');
      if (container && typeof globalDetails === 'object' && globalDetails) {{
        const now = new Date(todayStr.replace(/\./g,'-').replace(/\//g,'-'));
        const cutoff = new Date(now.getTime() - 45 * 24 * 3600 * 1000);
        // 45天内统计（用于筛选）
        const stats45 = {{}};
        const add45 = (sku) => {{ if (!stats45[sku]) stats45[sku] = {{details:0, orders:0, returns:0, revenue:0, typeCount: {{}}, reasonCount: {{}}}}; return stats45[sku]; }};
        // 历史全部统计（用于展示）
        const statsAll = {{}};
        const addAll = (sku) => {{ if (!statsAll[sku]) statsAll[sku] = {{details:0, orders:0, returns:0, revenue:0, typeCount: {{}}, reasonCount: {{}}}}; return statsAll[sku]; }};
        const parseDate = (s) => {{
          if (!s) return null;
          const t = new Date(String(s).replace(/\./g,'-').replace(/\//g,'-'));
          return isNaN(t) ? null : t;
        }};
        Object.values(globalDetails).forEach(list => {{
          if (!Array.isArray(list)) return;
          list.forEach(r => {{
            const sku = (r['货品名'] || '').trim();
            if (!sku) return;
            // 排除代发和样品
            const platform = String(r['下单平台'] || '').trim();
            const prod = String(r['商品名称'] || '');
            const rmk = String(r['备注'] || '');
            const combined = platform + ' ' + prod + ' ' + sku + ' ' + rmk;
            if (combined.includes('代发') || combined.includes('样品')) return;
            const d = parseDate(r['下单时间']);
            if (!d) return; // 跳过无效日期
            const refundType = String(r['退款类型'] || '').trim();
            const returnNo = String(r['退货单号'] || '').trim();
            const orderNo = String(r['订单号'] || '').trim();
            const pay = Number(r['付款金额'] || 0) || 0;
            const isCancel = refundType.includes('取消') || orderNo.includes('取消');
            const isReturn = !isCancel && ((/退|退货|退款/.test(refundType)) || (returnNo && returnNo !== '/' && returnNo !== '-'));

            // 计算45天内统计（用于筛选）
            const is45Day = d >= cutoff;
            if (is45Day) {{
              const stat45 = add45(sku);
              if (!isCancel && pay > 0) {{ stat45.orders += 1; stat45.revenue += Math.max(0, pay); stat45.details += 1; }}
              if (isReturn) {{
                stat45.returns += 1;
                if (refundType) {{ stat45.typeCount[refundType] = (stat45.typeCount[refundType]||0) + 1; }}
                const rr = String(r['退款原因']||'').trim();
                if (rr) {{ stat45.reasonCount[rr] = (stat45.reasonCount[rr]||0) + 1; }}
              }}
            }}

            // 计算历史全部统计（用于展示）
            const statAll = addAll(sku);
            if (!isCancel && pay > 0) {{ statAll.orders += 1; statAll.revenue += Math.max(0, pay); statAll.details += 1; }}
            if (isReturn) {{
              statAll.returns += 1;
              if (refundType) {{ statAll.typeCount[refundType] = (statAll.typeCount[refundType]||0) + 1; }}
              const rr = String(r['退款原因']||'').trim();
              if (rr) {{ statAll.reasonCount[rr] = (statAll.reasonCount[rr]||0) + 1; }}
            }}
          }});
        }});
        function topKey(obj) {{
          let k = '', m = 0;
          Object.entries(obj||{{}}).forEach(([kk, vv]) => {{ if ((vv|0) > m) {{ m = vv|0; k = kk; }} }});
          return k;
        }}
        function topKeys(obj, n) {{
          const arr = Object.entries(obj||{{}}).map(([k,v]) => [k, v|0]);
          arr.sort((a,b) => (b[1]-a[1]) || (a[0] > b[0] ? 1 : -1));
          return arr.slice(0, Math.max(0, n|0)).map(x => x[0]);
        }}
        function truncate(text, maxLen) {{
          const s = String(text || '');
          if (s.length <= maxLen) return s;
          return s.slice(0, Math.max(0, maxLen)).trim() + '…';
        }}
        const rows = Object.entries(stats45).map(([sku, s45]) => {{
          // 使用45天数据进行筛选
          const orders45 = s45.orders|0;
          if (!orders45) return null;
          const rr45 = Math.min(1, Math.max(0, (s45.returns||0) / Math.max(1, orders45)));

          // 筛选条件：45天内明细>3 且 退货率>30%
          if (s45.details <= 3 || rr45 <= 0.30) return null;

          // 获取历史全部数据用于展示
          const sAll = statsAll[sku] || {{details:0, orders:0, returns:0, revenue:0, typeCount: {{}}, reasonCount: {{}}}};
          const ordersAll = sAll.orders|0;
          const rrAll = ordersAll ? Math.min(1, Math.max(0, (sAll.returns||0) / Math.max(1, ordersAll))) : 0;
          const reasonsTop3 = topKeys(sAll.reasonCount, 3).map(x => truncate(x, 10)).join('、');

          // 返回历史全部数据用于展示
          return {{ sku, details: sAll.details|0, orders: ordersAll, rev: Number(sAll.revenue||0), rr: rrAll, rtype: topKey(sAll.typeCount), rreasons: reasonsTop3 }};
        }}).filter(x => x)
          .sort((a,b) => b.rr - a.rr || b.details - a.details || b.rev - a.rev);
        if (rows.length) {{
          const head = "<thead><tr><th>货品名</th><th data-sort-method='number'>明细笔数</th><th data-sort-method='number'>销售额</th><th data-sort-method='number'>退货率</th><th>退款类型</th></tr></thead>";
          const body = rows.map(r => `<tr data-sku=\"${{r.sku}}\"><td>${{r.sku}}</td><td data-sort-value='${{r.details}}'>${{r.details}}</td><td data-sort-value='${{r.rev.toFixed(2)}}'>${{r.rev.toFixed(2)}}</td><td data-sort-value='${{r.rr.toFixed(6)}}'>${{(r.rr*100).toFixed(1)}}%</td><td>${{r.rtype||''}}</td></tr>`).join('');
          container.innerHTML = `<table class='mini-table'>${{head}}<tbody>${{body}}</tbody></table>`;
          container.querySelectorAll('tr[data-sku]').forEach(tr => {{
            tr.addEventListener('click', (e) => {{
              // 忽略点击表头的情况
              if (e.target.tagName === 'TH' || e.target.closest('th')) return;
              const sku = tr.getAttribute('data-sku') || '';
              if (sku) openDetailForSku(sku, 'all');
            }});
          }});
          try {{ new Tablesort(container.querySelector('table')); }} catch (e) {{}}
        }} else {{
          container.innerHTML = '<p>暂无符合条件的SKU。</p>';
        }}
      }}
    }} catch (e) {{ /* no-op */ }}

    // 为低毛利预警表格添加点击事件和排序
    try {{
      const lowProfitTable = document.getElementById('lowProfitTable');
      if (lowProfitTable) {{
        // 初始化排序
        try {{ new Tablesort(lowProfitTable); }} catch (e) {{}}
        // 绑定点击事件（传递当前过滤器类型）
        function bindLowProfitClicks() {{
          lowProfitTable.querySelectorAll('tbody tr[data-sku]').forEach(tr => {{
            tr.addEventListener('click', (e) => {{
              // 忽略点击表头的情况
              if (e.target.tagName === 'TH' || e.target.closest('th')) return;
              const sku = tr.getAttribute('data-sku') || '';
              if (!sku) return;
              // 获取当前选择的过滤器类型
              const currentFilter = document.querySelector('input[name="lowProfitFilter"]:checked')?.value || 'normal';
              openDetailForSku(sku, currentFilter);
            }});
          }});
        }}
        bindLowProfitClicks();
        
        // 添加过滤器逻辑
        const filterRadios = document.querySelectorAll('input[name="lowProfitFilter"]');
        const countSpan = document.getElementById('lowProfitCount');
        
        function applyLowProfitFilter() {{
          const selectedValue = document.querySelector('input[name="lowProfitFilter"]:checked')?.value || 'normal';
          const rows = lowProfitTable.querySelectorAll('tbody tr[data-type]');
          let visibleCount = 0;
          
          rows.forEach(row => {{
            const rowType = row.getAttribute('data-type');
            if (selectedValue === 'normal' && rowType === 'normal') {{
              row.style.display = '';
              visibleCount++;
            }} else if (selectedValue === 'proxy' && rowType === 'proxy') {{
              row.style.display = '';
              visibleCount++;
            }} else {{
              row.style.display = 'none';
            }}
          }});
          
          if (countSpan) {{
            countSpan.textContent = `(共${{visibleCount}}款)`;
          }}
        }}
        
        filterRadios.forEach(radio => {{
          radio.addEventListener('change', applyLowProfitFilter);
        }});
        
        // 初始应用过滤
        applyLowProfitFilter();
      }}
    }} catch (e) {{ /* no-op */ }}

    // ============ 角色切换功能 ============
    (function() {{
        const roleOperations = document.getElementById('roleOperations');
        const roleCustomerService = document.getElementById('roleCustomerService');

        // 定义需要控制的元素
        const operationsElements = {{
            summary: document.querySelector('.summary'), // SKU分析卡片
            searchContainer: document.querySelector('.operations-search-container') // 运营搜索框
        }};
        const customerServiceElements = {{
            toolbar: document.querySelector('.toolbar'),
            listTabs: document.querySelector('.list-tabs'),
            filters: document.querySelector('.filters'),
            table: document.querySelector('table')
        }};

        // 角色切换函数
        function switchRole(role) {{
            // Sync Top Buttons & Stats
            const btnService = document.getElementById('topRoleService');
            const btnOps = document.getElementById('topRoleOps');
            const dashboardStats = document.getElementById('dashboardStats');
            
            if (btnService && btnOps) {{
                const baseClass = "px-3 py-1.5 rounded-md text-xs font-semibold transition-all ";
                const active = "bg-white text-brand-600 shadow-sm";
                const inactive = "text-slate-500 hover:text-slate-700 bg-transparent shadow-none";
                
                if (role === 'customer-service') {{
                    btnService.className = baseClass + active;
                    btnOps.className = baseClass + inactive;
                    if (dashboardStats) dashboardStats.style.display = '';
                }} else {{
                    btnService.className = baseClass + inactive;
                    btnOps.className = baseClass + active;
                    if (dashboardStats) dashboardStats.style.display = 'none';
                }}
            }}

            if (role === 'operations') {{
                // Check if search has content
                const opsSearch = document.getElementById('operationsSearchBox');
                const hasSearch = opsSearch && opsSearch.value.trim();
                
                // 运营视角：显示SKU分析、搜索框和客户列表，隐藏工具栏和列表标签
                if (operationsElements.summary) operationsElements.summary.style.display = hasSearch ? 'none' : 'grid';
                if (operationsElements.searchContainer) {{
                    operationsElements.searchContainer.classList.add('active');
                }}
                if (customerServiceElements.toolbar) customerServiceElements.toolbar.style.display = 'none';
                if (customerServiceElements.listTabs) customerServiceElements.listTabs.style.display = 'none';
                if (customerServiceElements.filters) customerServiceElements.filters.style.display = 'none';
                // 保留客户列表表格，用于显示搜索结果
                if (customerServiceElements.table) customerServiceElements.table.style.display = hasSearch ? 'table' : 'none';
            }} else {{
                // 客服视角：隐藏SKU分析和运营搜索框，显示所有客服工具
                if (operationsElements.summary) operationsElements.summary.style.display = 'none';
                if (operationsElements.searchContainer) {{
                    operationsElements.searchContainer.classList.remove('active');
                }}
                if (customerServiceElements.toolbar) customerServiceElements.toolbar.style.display = 'flex';
                if (customerServiceElements.listTabs) customerServiceElements.listTabs.style.display = 'block';
                if (customerServiceElements.filters) customerServiceElements.filters.style.display = 'flex';
                if (customerServiceElements.table) customerServiceElements.table.style.display = 'table';
            }}
        }}

        // 监听角色切换事件
        if (roleOperations) {{
            roleOperations.addEventListener('change', function() {{
                if (this.checked) {{
                    switchRole('operations');
                }}
            }});
        }}

        if (roleCustomerService) {{
            roleCustomerService.addEventListener('change', function() {{
                if (this.checked) {{
                    switchRole('customer-service');
                }}
            }});
        }}

        // 初始化：根据默认选中的角色设置显示
        const initialRole = roleCustomerService.checked ? 'customer-service' : 'operations';
        switchRole(initialRole);

        // ============ 运营搜索框功能 ============
        const operationsSearchBox = document.getElementById('operationsSearchBox');
        if (operationsSearchBox) {{
            operationsSearchBox.addEventListener('input', applyFilters);
        }}
    }})();
    </script>
</body>
</html>
"""
    # 将原始 HTML 包装为基于「未命名.html」风格的 SaaS 布局
    # 若包装过程中出现异常，则回退使用原始布局
    try:
        import re as _re  # 局部引入，避免与上方使用冲突

        legacy_html = html_template
        body_match = _re.search(r"<body[^>]*>(.*)</body>", legacy_html, _re.DOTALL)
        body_content = body_match.group(1) if body_match else ""
        # 提取所有脚本内容，并从 body_content 中移除脚本，避免重复执行
        scripts = _re.findall(r"<script[^>]*>(.*?)</script>", legacy_html, _re.DOTALL)
        body_content = _re.sub(r"<script[^>]*>.*?</script>", "", body_content, flags=_re.DOTALL)

        gen_date = today.isoformat()

        # 计算统计卡片数据
        high_priority_count = sum(1 for row in action_rows if row.get('priority_score', 0) >= 80)
        mid_priority_count = sum(1 for row in action_rows if 50 <= row.get('priority_score', 0) < 80)
        total_customers = len(action_rows)

        # SKU预警数量（从overview_rows或global_details计算）
        # 这里使用一个估算值，实际数量在SKU卡片区域已有详细统计
        sku_alert_count = cooldown_total if cooldown_total > 0 else 0

        new_html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>客户触达仪表盘 - {gen_date}</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@400;500;600;700&family=IBM+Plex+Sans:wght@300;400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">

    <script>
        tailwind.config = {{
            theme: {{
                extend: {{
                    fontFamily: {{
                        sans: ['Noto Sans SC', 'IBM Plex Sans', 'system-ui', '-apple-system', 'BlinkMacSystemFont', 'Segoe UI', 'sans-serif'],
                        mono: ['IBM Plex Mono', 'ui-monospace', 'SFMono-Regular', 'Menlo', 'Monaco', 'Consolas', 'Liberation Mono', 'Courier New', 'monospace'],
                    }},
                    colors: {{
                        brand: {{
                            50: '#eff6ff',
                            100: '#dbeafe',
                            500: '#3b82f6',
                            600: '#2563eb',
                            700: '#1d4ed8',
                            900: '#1e3a8a',
                        }}
                    }}
                }}
            }}
        }}
    </script>

    <style>
        /* 滚动条样式 */
        ::-webkit-scrollbar {{ width: 6px; height: 6px; }}
        ::-webkit-scrollbar-track {{ background: transparent; }}
        ::-webkit-scrollbar-thumb {{ background: #cbd5e1; border-radius: 3px; }}
        ::-webkit-scrollbar-thumb:hover {{ background: #94a3b8; }}

        /* 导航激活态 */
        .nav-item.active {{
            background-color: #eff6ff;
            color: #2563eb;
            border-right: 3px solid #2563eb;
        }}

        /* 兼容原有内容的基础字体 */
        body {{
            font-family: 'Noto Sans SC', 'IBM Plex Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
        }}

        /* 隐藏原始标题与角色切换器（新布局中已有对应组件） */
        #originalContent > h1:first-of-type {{
            display: none;
        }}

        #originalContent .role-selector {{
            display: none;
        }}

        /* 原有 meta/summary 等区域样式微调以适配新布局 */
        .meta {{
            background: white;
            border: 1px solid #e2e8f0;
            padding: 16px;
            border-radius: 12px;
            margin-bottom: 24px;
            font-size: 13px;
            color: #64748b;
        }}

        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(max(320px, calc((100% - 32px) / 3)), 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}

        @media (max-width: 1200px) {{
            .summary {{
                grid-template-columns: repeat(auto-fill, minmax(max(280px, calc((100% - 16px) / 2)), 1fr));
            }}
        }}

        @media (max-width: 768px) {{
            .summary {{
                grid-template-columns: 1fr;
            }}
        }}

        .card {{
            background: white;
            border: 1px solid #e2e8f0;
            border-radius: 12px;
            padding: 16px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }}

        .card h3 {{
            color: #1e293b;
            font-size: 14px;
            font-weight: 600;
            margin-bottom: 12px;
        }}

        /* 工具栏样式 */
        .toolbar {{
            background: white;
            border: 1px solid #e2e8f0;
            padding: 12px 16px;
            border-radius: 12px;
            margin-bottom: 16px;
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
            align-items: center;
        }}

        .toolbar button {{
            background: #f8fafc;
            border: 1px solid #e2e8f0;
            color: #475569;
            padding: 6px 12px;
            border-radius: 6px;
            font-size: 13px;
            cursor: pointer;
            transition: all 0.2s;
        }}

        .toolbar button:hover {{
            background: #3b82f6;
            color: white;
            border-color: #3b82f6;
        }}

        /* 表格样式改造 */
        table {{
            width: 100%;
            border-collapse: collapse;
            background: white;
        }}

        thead {{
            background: #f8fafc;
            border-bottom: 1px solid #e2e8f0;
        }}

        th {{
            padding: 12px 16px;
            text-align: left;
            font-size: 12px;
            font-weight: 600;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}

        td {{
            padding: 12px 16px;
            border-bottom: 1px solid #f1f5f9;
            font-size: 13px;
            color: #1e293b;
        }}

        tbody tr:hover {{
            background: #f8fafc;
        }}

        /* 标签页样式 */
        .list-tabs {{
            background: white;
            border: 1px solid #e2e8f0;
            padding: 12px;
            border-radius: 12px;
            margin-bottom: 16px;
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
        }}

        .list-tab {{
            padding: 8px 16px !important;
            border: 1px solid #e2e8f0 !important;
            background: #f8fafc !important;
            border-radius: 6px !important;
            font-size: 13px !important;
            font-weight: 500 !important;
            color: #64748b !important;
            cursor: pointer;
            transition: all 0.2s !important;
            border-bottom: none !important;
        }}

        .list-tab:hover {{
            background: #eff6ff !important;
            color: #2563eb !important;
            border-color: #3b82f6 !important;
        }}

        .list-tab.active,
        .list-tab[style*="color: #00f2ff"],
        .list-tab[style*="color:#00f2ff"] {{
            background: #eff6ff !important;
            color: #2563eb !important;
            border-color: #3b82f6 !important;
            font-weight: 600 !important;
        }}

        /* 筛选器样式 */
        .filters {{
            background: white;
            border: 1px solid #e2e8f0;
            padding: 14px;
            border-radius: 12px;
            margin-bottom: 16px;
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
            align-items: center;
        }}

        .filters label {{
            font-size: 13px;
            color: #64748b;
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .filters input,
        .filters select {{
            background: white;
            border: 1px solid #e2e8f0;
            color: #1e293b;
            padding: 6px 12px;
            border-radius: 6px;
            font-size: 13px;
        }}

        .filters input:focus,
        .filters select:focus {{
            outline: none;
            border-color: #3b82f6;
            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
        }}

        /* 隐藏类 */
        .hidden {{
            display: none !important;
        }}

        /* 运营搜索框 */
        .operations-search-container {{
            background: white;
            border: 1px solid #e2e8f0;
            padding: 12px;
            border-radius: 12px;
            margin-bottom: 16px;
        }}

        .operations-search-container.active {{
            display: block;
        }}

        .operations-search-box {{
            width: 100%;
            padding: 10px 16px;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
            font-size: 14px;
        }}

        .operations-search-box:focus {{
            outline: none;
            border-color: #3b82f6;
            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
        }}

        /* 优先级边框 */
        .priority-high {{ border-left: 3px solid #ef4444; }}
        .priority-mid {{ border-left: 3px solid #f59e0b; }}
        .priority-low {{ border-left: 3px solid #3b82f6; }}
        .priority-other {{ border-left: 3px solid #94a3b8; }}

        /* 已完成样式 */
        tr.completed {{
            opacity: 0.6;
        }}
        tr.completed td {{
            text-decoration: line-through;
        }}

        /* 详情面板 */
        .detail-backdrop {{
            position: fixed;
            inset: 0;
            background: rgba(0, 0, 0, 0.5);
            z-index: 999;
            backdrop-filter: blur(4px);
        }}

        .detail-panel {{
            position: fixed;
            left: 0;
            right: 0;
            bottom: 0;
            height: 80vh;
            background: white;
            border-top-left-radius: 16px;
            border-top-right-radius: 16px;
            box-shadow: 0 -4px 30px rgba(0, 0, 0, 0.2);
            z-index: 1000;
            display: flex;
            flex-direction: column;
        }}

        .detail-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 16px 20px;
            border-bottom: 1px solid #e2e8f0;
        }}

        .detail-header h3 {{
            margin: 0;
            font-size: 18px;
            font-weight: 600;
            color: #1e293b;
        }}

        .detail-close {{
            width: 32px;
            height: 32px;
            border-radius: 8px;
            border: none;
            background: #f1f5f9;
            color: #64748b;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-center;
            transition: all 0.2s;
        }}

        .detail-close:hover {{
            background: #e2e8f0;
            color: #1e293b;
        }}

        .detail-body {{
            flex: 1;
            overflow: auto;
            padding: 16px 20px;
        }}

        /* 冷却期容器 */
        #cooldownContainer {{
            background: white;
            border: 1px solid #e2e8f0;
            border-radius: 12px;
            padding: 16px;
            margin-bottom: 16px;
        }}

        /* Mini表格 */
        .mini-table {{
            width: 100%;
            border-collapse: collapse;
        }}

        .mini-table th,
        .mini-table td {{
            padding: 8px 12px;
            border-bottom: 1px solid #f1f5f9;
            font-size: 12px;
        }}

        .mini-table th {{
            background: #f8fafc;
            color: #64748b;
            font-weight: 600;
            cursor: pointer;
            user-select: none;
            position: relative;
            transition: background-color 0.15s ease, color 0.15s ease;
        }}

        .mini-table th:hover {{
            background: #e2e8f0;
            color: #475569;
        }}

        /* 排序箭头样式 */
        .mini-table th::after {{
            content: '⇅';
            margin-left: 6px;
            font-size: 10px;
            color: #cbd5e1;
            opacity: 0;
            transition: opacity 0.2s ease;
        }}

        .mini-table th:hover::after {{
            opacity: 1;
        }}

        .mini-table th[data-sort-order='asc']::after {{
            content: '▲';
            color: #3b82f6;
            opacity: 1;
        }}

        .mini-table th[data-sort-order='desc']::after {{
            content: '▼';
            color: #3b82f6;
            opacity: 1;
        }}

        /* 禁止排序的表头样式 */
        .mini-table th.no-sort {{
            cursor: default;
        }}

        .mini-table th.no-sort::after {{
            content: none;
        }}

        /* SKU导航按钮 */
        .sku-nav {{
            display: flex;
            gap: 8px;
            justify-content: flex-end;
            margin-bottom: 8px;
        }}

        .sku-nav button {{
            padding: 4px 10px;
            border: 1px solid #e2e8f0;
            background: white;
            border-radius: 4px;
            font-size: 12px;
            cursor: pointer;
        }}

        .sku-nav button:hover {{
            background: #f8fafc;
        }}

        /* 滚动区域 */
        .scroll-pane {{
            max-height: 400px;
            overflow-y: auto;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
        }}

        /* 页脚 */
        .footer {{
            margin-top: 24px;
            color: #94a3b8;
            font-size: 12px;
        }}

        /* AI Chatbox */
        .ai-chat-btn {{
            position: fixed;
            bottom: 24px;
            right: 24px;
            width: 56px;
            height: 56px;
            border-radius: 28px;
            background: linear-gradient(135deg, #6366f1 0%, #a855f7 100%);
            box-shadow: 0 4px 14px rgba(99, 102, 241, 0.4);
            color: #fff;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 24px;
            cursor: pointer;
            z-index: 2000;
            transition: transform 0.2s;
        }}
        .ai-chat-btn:hover {{ transform: scale(1.05); }}
        
        .ai-chat-panel {{
            position: fixed;
            bottom: 90px;
            right: 24px;
            width: 360px;
            height: 500px;
            background: #fff;
            border-radius: 12px;
            box-shadow: 0 8px 30px rgba(0,0,0,0.12);
            z-index: 2000;
            display: flex;
            flex-direction: column;
            border: 1px solid #e2e8f0;
        }}
        .ai-chat-panel.hidden {{ display: none; }}
        .ai-chat-header {{
            padding: 12px 16px;
            background: #f8fafc;
            border-bottom: 1px solid #e2e8f0;
            border-radius: 12px 12px 0 0;
            display: flex;
            justify-content: space-between;
            align-items: center;
            font-weight: 600;
            color: #334155;
        }}
        .ai-chat-body {{
            flex: 1;
            padding: 16px;
            overflow-y: auto;
            display: flex;
            flex-direction: column;
            gap: 12px;
            font-size: 13px;
        }}
        .chat-msg {{
            padding: 8px 12px;
            border-radius: 8px;
            max-width: 85%;
            line-height: 1.5;
        }}
        .chat-msg.user {{
            align-self: flex-end;
            background: #eff6ff;
            color: #1e40af;
            border: 1px solid #dbeafe;
        }}
        .chat-msg.ai {{
            align-self: flex-start;
            background: #f1f5f9;
            color: #334155;
            white-space: pre-wrap;
            font-family: monospace;
        }}
        .ai-chat-input-area {{
            padding: 12px;
            border-top: 1px solid #e2e8f0;
            display: flex;
            gap: 8px;
        }}
        .ai-chat-input {{
            flex: 1;
            padding: 8px;
            border: 1px solid #cbd5e1;
            border-radius: 6px;
            font-size: 13px;
            outline: none;
        }}
        .ai-chat-input:focus {{ border-color: #6366f1; }}
        .ai-chat-send {{
            padding: 6px 12px;
            background: #6366f1;
            color: #fff;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 13px;
        }}
        .ai-chat-send:disabled {{ background: #cbd5e1; cursor: not-allowed; }}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 font-sans h-screen flex overflow-hidden">

    <!-- 左侧导航栏 -->
    <aside class="w-64 bg-white border-r border-slate-200 flex flex-col flex-shrink-0 z-20 shadow-sm">
        <!-- Logo -->
        <div class="h-16 flex items-center gap-3 px-6 border-b border-slate-100">
            <div class="w-8 h-8 rounded-lg bg-gradient-to-br from-brand-600 to-brand-500 flex items-center justify-center text-white shadow-lg">
                <i class="fa-solid fa-chart-line text-sm"></i>
            </div>
            <span class="font-bold text-lg tracking-tight text-slate-800">客户预警 <span class="text-xs font-normal text-slate-400">v3.0</span></span>
        </div>

        <!-- 菜单 -->
        <nav class="flex-1 py-6 px-3 space-y-1 overflow-y-auto">
            <div class="px-3 mb-2 text-xs font-semibold text-slate-400 uppercase tracking-wider">总览</div>
            <a href="#" onclick="document.getElementById('roleCustomerService')?.click(); setTimeout(() => window.scrollTo({{top: 0, behavior: 'smooth'}}), 100); return false;" class="nav-item active flex items-center gap-3 px-3 py-2.5 rounded-lg text-sm font-medium text-slate-600 hover:bg-slate-50 hover:text-slate-900 transition-colors">
                <i class="fa-solid fa-chart-pie w-5 text-center"></i> 触达仪表盘
            </a>
            <a href="#" onclick="document.getElementById('roleCustomerService')?.click(); setTimeout(() => document.querySelector('.list-tab[data-list=\\'冷却期\\']')?.click(), 100); return false;" class="nav-item flex items-center gap-3 px-3 py-2.5 rounded-lg text-sm font-medium text-slate-600 hover:bg-slate-50 hover:text-slate-900 transition-colors">
                <i class="fa-solid fa-clock w-5 text-center"></i> 冷却期客户
                <span id="sidebarCooldownBadge" class="ml-auto bg-blue-100 text-blue-600 py-0.5 px-2 rounded-full text-xs font-bold">0</span>
            </a>

            <div class="px-3 mt-6 mb-2 text-xs font-semibold text-slate-400 uppercase tracking-wider">SKU分析</div>
            <a href="#" onclick="(function(){{ const opsInput=document.getElementById('operationsSearchBox'); if(opsInput) opsInput.value=''; const opsContainer=document.querySelector('.operations-search-container'); if(opsContainer) opsContainer.classList.add('active'); const summary=document.getElementById('skuSummary'); const table=document.getElementById('actionTable'); if(summary) summary.style.display='grid'; if(table) table.style.display='none'; }})(); document.getElementById('roleOperations')?.click(); setTimeout(() => document.getElementById('skuSummary')?.scrollIntoView({{behavior: 'smooth'}}), 100); return false;" class="nav-item flex items-center gap-3 px-3 py-2.5 rounded-lg text-sm font-medium text-slate-600 hover:bg-slate-50 hover:text-slate-900 transition-colors">
                <i class="fa-solid fa-fire w-5 text-center"></i> 加推SKU
            </a>
            <a href="#" onclick="(function(){{ const opsInput=document.getElementById('operationsSearchBox'); if(opsInput) opsInput.value=''; const opsContainer=document.querySelector('.operations-search-container'); if(opsContainer) opsContainer.classList.add('active'); const summary=document.getElementById('skuSummary'); const table=document.getElementById('actionTable'); if(summary) summary.style.display='grid'; if(table) table.style.display='none'; }})(); document.getElementById('roleOperations')?.click(); setTimeout(() => document.getElementById('skuReturnAlertTable')?.scrollIntoView({{behavior: 'smooth'}}), 100); return false;" class="nav-item flex items-center gap-3 px-3 py-2.5 rounded-lg text-sm font-medium text-slate-600 hover:bg-slate-50 hover:text-slate-900 transition-colors">
                <i class="fa-solid fa-triangle-exclamation w-5 text-center"></i> 高退货预警
            </a>
            <a href="#" onclick="(function(){{ const opsInput=document.getElementById('operationsSearchBox'); if(opsInput) opsInput.value=''; const opsContainer=document.querySelector('.operations-search-container'); if(opsContainer) opsContainer.classList.add('active'); const summary=document.getElementById('skuSummary'); const table=document.getElementById('actionTable'); if(summary) summary.style.display='grid'; if(table) table.style.display='none'; }})(); document.getElementById('roleOperations')?.click(); setTimeout(() => document.getElementById('lowProfitTable')?.scrollIntoView({{behavior: 'smooth'}}), 100); return false;" class="nav-item flex items-center gap-3 px-3 py-2.5 rounded-lg text-sm font-medium text-slate-600 hover:bg-slate-50 hover:text-slate-900 transition-colors">
                <i class="fa-solid fa-circle-exclamation w-5 text-center"></i> 低毛利预警
            </a>

            <div class="px-3 mt-6 mb-2 text-xs font-semibold text-slate-400 uppercase tracking-wider">操作</div>
            <a href="#" onclick="document.getElementById('exportCsv')?.click(); return false;" class="nav-item flex items-center gap-3 px-3 py-2.5 rounded-lg text-sm font-medium text-slate-600 hover:bg-slate-50 hover:text-slate-900 transition-colors">
                <i class="fa-solid fa-download w-5 text-center"></i> 导出记录
            </a>
            <a href="#" onclick="document.getElementById('clearMarks')?.click(); return false;" class="nav-item flex items-center gap-3 px-3 py-2.5 rounded-lg text-sm font-medium text-slate-600 hover:bg-slate-50 hover:text-slate-900 transition-colors">
                <i class="fa-solid fa-eraser w-5 text-center"></i> 清除标记
            </a>
        </nav>

        <!-- 底部用户信息 -->
        <div class="p-4 border-t border-slate-100">
            <div class="flex items-center gap-3 p-2 rounded-lg hover:bg-slate-50 cursor-pointer transition">
                <div class="w-9 h-9 rounded-full bg-brand-100 text-brand-600 flex items-center justify-center font-bold text-sm">
                    客
                </div>
                <div class="flex-1 min-w-0">
                    <p class="text-sm font-medium text-slate-900 truncate">客服团队</p>
                    <p class="text-xs text-slate-500 truncate">当前在线</p>
                </div>
                <i class="fa-solid fa-chevron-right text-xs text-slate-400"></i>
            </div>
        </div>
    </aside>

    <!-- 主内容区 -->
    <div class="flex-1 flex flex-col min-w-0 overflow-hidden">

        <!-- 顶部栏 -->
        <header class="h-16 bg-white border-b border-slate-200 flex items-center justify-between px-6 flex-shrink-0 z-10">
            <div class="flex items-center text-sm text-slate-500">
                <span class="hover:text-slate-800 cursor-pointer">首页</span>
                <i class="fa-solid fa-chevron-right text-xs mx-2 text-slate-300"></i>
                <span class="font-semibold text-slate-800">客户触达</span>
            </div>

            <div class="flex items-center gap-4">
                <!-- 角色切换 -->
                <div class="bg-slate-100 p-1 rounded-lg flex items-center">
                    <button id="topRoleService" onclick="switchTopRole('customer-service')" class="px-3 py-1.5 rounded-md text-xs font-semibold transition-all shadow-sm bg-white text-brand-600">
                        <i class="fa-solid fa-headset mr-1.5"></i>客服
                    </button>
                    <button id="topRoleOps" onclick="switchTopRole('operations')" class="px-3 py-1.5 rounded-md text-xs font-semibold transition-all text-slate-500 hover:text-slate-700">
                        <i class="fa-solid fa-user-gear mr-1.5"></i>运营
                    </button>
                </div>

                <div class="h-4 w-px bg-slate-200 mx-1"></div>

                <button onclick="location.reload()" class="relative p-2 text-slate-400 hover:text-slate-600 transition" title="刷新数据">
                    <i class="fa-solid fa-rotate text-lg"></i>
                </button>
            </div>
        </header>

        <!-- 可滚动内容区 -->
        <main class="flex-1 overflow-y-auto p-6 md:p-8 scroll-smooth">
            <div class="max-w-[1600px] mx-auto">
                <!-- 标题 -->
                <div class="mb-6">
                    <h2 class="text-2xl font-bold text-slate-900">客户触达仪表盘</h2>
                    <p class="text-sm text-slate-500 mt-1">生成日期: {gen_date}</p>
                </div>

                <!-- 统计卡片网格 -->
                <div id="dashboardStats" class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-4 gap-4 mb-6">
                    <!-- 卡片1: 高优先级客户 -->
                    <div class="bg-white p-5 rounded-2xl border border-slate-200 shadow-sm hover:shadow-md transition-shadow group">
                        <div class="flex justify-between items-start">
                            <div>
                                <p class="text-xs font-bold text-slate-400 uppercase tracking-wide">高优先级客户</p>
                                <p class="text-3xl font-mono font-bold text-slate-900 mt-2">{high_priority_count}</p>
                            </div>
                            <div class="w-10 h-10 rounded-xl bg-red-50 text-red-500 flex items-center justify-center group-hover:scale-110 transition-transform">
                                <i class="fa-solid fa-triangle-exclamation text-lg"></i>
                            </div>
                        </div>
                        <div class="mt-4 flex items-center text-xs font-medium text-red-600 bg-red-50 w-fit px-2 py-1 rounded-full">
                            <span>优先分 ≥ 80</span>
                        </div>
                    </div>

                    <!-- 卡片2: 中优先级客户 -->
                    <div class="bg-white p-5 rounded-2xl border border-slate-200 shadow-sm hover:shadow-md transition-shadow group">
                        <div class="flex justify-between items-start">
                            <div>
                                <p class="text-xs font-bold text-slate-400 uppercase tracking-wide">中优先级客户</p>
                                <p class="text-3xl font-mono font-bold text-slate-900 mt-2">{mid_priority_count}</p>
                            </div>
                            <div class="w-10 h-10 rounded-xl bg-amber-50 text-amber-500 flex items-center justify-center group-hover:scale-110 transition-transform">
                                <i class="fa-regular fa-clock text-lg"></i>
                            </div>
                        </div>
                        <div class="mt-4 flex items-center text-xs font-medium text-amber-600 bg-amber-50 w-fit px-2 py-1 rounded-full">
                            <span>优先分 50-79</span>
                        </div>
                    </div>

                    <!-- 卡片3: 触达客户总数 -->
                    <div class="bg-white p-5 rounded-2xl border border-slate-200 shadow-sm hover:shadow-md transition-shadow group">
                        <div class="flex justify-between items-start">
                            <div>
                                <p class="text-xs font-bold text-slate-400 uppercase tracking-wide">触达客户总数</p>
                                <p class="text-3xl font-mono font-bold text-slate-900 mt-2">{total_customers:,}</p>
                            </div>
                            <div class="w-10 h-10 rounded-xl bg-brand-50 text-brand-500 flex items-center justify-center group-hover:scale-110 transition-transform">
                                <i class="fa-solid fa-users text-lg"></i>
                            </div>
                        </div>
                        <div class="mt-4 flex items-center text-xs font-medium text-brand-600 bg-brand-50 w-fit px-2 py-1 rounded-full">
                            <i class="fa-solid fa-check mr-1"></i> 系统运行中
                        </div>
                    </div>

                    <!-- 卡片4: 冷却期客户 -->
                    <div class="bg-white p-5 rounded-2xl border border-slate-200 shadow-sm hover:shadow-md transition-shadow group">
                        <div class="flex justify-between items-start">
                            <div>
                                <p class="text-xs font-bold text-slate-400 uppercase tracking-wide">冷却期客户</p>
                                <p class="text-3xl font-mono font-bold text-slate-900 mt-2">{cooldown_total}</p>
                            </div>
                            <div class="w-10 h-10 rounded-xl bg-purple-50 text-purple-500 flex items-center justify-center group-hover:scale-110 transition-transform">
                                <i class="fa-solid fa-clock-rotate-left text-lg"></i>
                            </div>
                        </div>
                        <div class="mt-4 flex items-center text-xs font-medium text-purple-600 bg-purple-50 w-fit px-2 py-1 rounded-full">
                            <span>近{cooldown_days}天已联系</span>
                        </div>
                    </div>
                </div>

                <!-- 原有内容 -->
                <div id="originalContent">
{body_content}
                </div>
            </div>
        </main>
    </div>

    <!-- 原有脚本 -->
'''
        for script in scripts:
            script_body = script.strip()
            if script_body:
                new_html += f"    <script>\n{script_body}\n    </script>\n"

        new_html += """
    <!-- AI Chat Components -->
    <div id="aiChatBtn" class="ai-chat-btn">🤖</div>
    <div id="aiChatPanel" class="ai-chat-panel hidden">
        <div class="ai-chat-header">
            <span>AI 数据分析师</span>
            <span id="closeAiChat" class="detail-close" style="font-size:12px;">✖</span>
        </div>
        <div id="aiChatBody" class="ai-chat-body">
            <div class="chat-msg ai">你好！我是您的智能数据分析师。您可以问我类似：
- “找出消费最高的5个客户”
- “统计各省份的销售额”
            </div>
        </div>
        <div class="ai-chat-input-area">
            <input type="text" id="aiChatInput" class="ai-chat-input" placeholder="输入问题..." />
            <button id="aiChatSend" class="ai-chat-send">发送</button>
        </div>
    </div>

    <script>
    // AI Chat Logic
    (function() {
        const btn = document.getElementById('aiChatBtn');
        const panel = document.getElementById('aiChatPanel');
        const close = document.getElementById('closeAiChat');
        const input = document.getElementById('aiChatInput');
        const send = document.getElementById('aiChatSend');
        const body = document.getElementById('aiChatBody');
        
        if (btn) btn.onclick = () => panel.classList.remove('hidden');
        if (close) close.onclick = () => panel.classList.add('hidden');
        
        async function sendMsg() {
            const query = input.value.trim();
            if (!query) return;
            
            // Add user message
            addMsg(query, 'user');
            input.value = '';
            input.disabled = true;
            send.disabled = true;
            
            // Show loading
            const loadingId = addMsg('思考中...', 'ai');
            
            try {
                const res = await fetch(`http://127.0.0.1:5005/ask`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ 'query': query })
                });
                const data = await res.json();
                
                // Remove loading
                const loadingEl = document.getElementById(loadingId);
                if (loadingEl) loadingEl.remove();
                
                if (data.ok) {
                    addMsg(data.answer, 'ai');
                } else {
                    addMsg('出错啦: ' + data.error, 'ai');
                }
            } catch (e) {
                const loadingEl = document.getElementById(loadingId);
                if (loadingEl) loadingEl.remove();
                addMsg('网络请求失败，请确认后台服务已启动。', 'ai');
            } finally {
                input.disabled = false;
                send.disabled = false;
                input.focus();
            }
        }
        
        function addMsg(text, type) {
            const div = document.createElement('div');
            div.className = 'chat-msg ' + type;
            div.textContent = text;
            div.id = 'msg-' + Date.now();
            body.appendChild(div);
            body.scrollTop = body.scrollHeight;
            return div.id;
        }
        
        if (send) send.onclick = sendMsg;
        if (input) input.onkeypress = (e) => {
            if (e.key === 'Enter') sendMsg();
        };
    })();
    </script>

    <!-- 布局适配脚本 -->
    <script>
        // 同步顶部角色切换到原有的 radio 按钮
        function switchTopRole(role) {
            const btnService = document.getElementById('topRoleService');
            const btnOps = document.getElementById('topRoleOps');

            const activeClass = "bg-white text-brand-600 shadow-sm";
            const inactiveClass = "text-slate-500 hover:text-slate-700 bg-transparent shadow-none";

            if (role === 'customer-service') {
                btnService.className = `px-3 py-1.5 rounded-md text-xs font-semibold transition-all ${activeClass}`;
                btnOps.className = `px-3 py-1.5 rounded-md text-xs font-semibold transition-all ${inactiveClass}`;
                const radio = document.getElementById('roleCustomerService');
                if (radio) radio.click();
                const dashboardStats = document.getElementById('dashboardStats');
                if (dashboardStats) dashboardStats.style.display = '';
            } else {
                btnService.className = `px-3 py-1.5 rounded-md text-xs font-semibold transition-all ${inactiveClass}`;
                btnOps.className = `px-3 py-1.5 rounded-md text-xs font-semibold transition-all ${activeClass}`;
                const radio = document.getElementById('roleOperations');
                if (radio) radio.click();
                const dashboardStats = document.getElementById('dashboardStats');
                if (dashboardStats) dashboardStats.style.display = 'none';
            }
        }

        // 更新侧边栏的冷却期徽章
        document.addEventListener('DOMContentLoaded', function() {
            const cooldownSpan = document.getElementById('cooldownTotalSpan');
            const badge = document.getElementById('sidebarCooldownBadge');
            if (cooldownSpan && badge) {
                badge.textContent = cooldownSpan.textContent;
            }
        });

        // 侧边栏导航高亮切换
        document.querySelectorAll('.nav-item').forEach(item => {
            item.addEventListener('click', function(e) {
                // 移除所有导航项的 active 类
                document.querySelectorAll('.nav-item').forEach(nav => {
                    nav.classList.remove('active');
                });
                // 给当前点击项添加 active 类
                this.classList.add('active');
            });
        });

        console.log('✅ SaaS 布局已加载');
    </script>
</body>
</html>
"""
        html_template = new_html
    except Exception as e:
        # 包装失败时使用原始布局，并打印错误信息供调试
        print(f"警告：SaaS布局包装失败，使用原始布局。错误: {e}")
        import traceback
        traceback.print_exc()
        pass

    output_path.write_text(html_template, encoding="utf-8")

def main():
    args = parse_arguments()
    source_path = Path(args.source)
    if not source_path.exists():
        raise FileNotFoundError(f"Source workbook not found: {source_path}")
    if args.churn_multiplier <= 0:
        raise ValueError("churn multiplier must be greater than 0.")

    today = (
        datetime.strptime(args.today, "%Y-%m-%d").date()
        if args.today
        else datetime.today().date()
    )

    month_offsets = parse_month_offsets(args.anniversary_months) if args.anniversary_months else []
    anniversary_dates = (
        build_anniversary_dates(today, month_offsets) if month_offsets else []
    )

    wb, ws = common_resolve_sheet(source_path, args.sheet)
    try:
        customers = load_customers(ws, today)
    finally:
        wb.close()

    cooldown_days = max(0, args.cooldown_days or 0)
    cooldown_scope = str(getattr(args, 'cooldown_scope', 'action') or 'action').strip().lower()
    # 支持通过环境变量覆盖：COOLDOWN_SCOPE=all/action
    cooldown_scope_env = os.getenv('COOLDOWN_SCOPE')
    if cooldown_scope_env:
        cooldown_scope = cooldown_scope_env.strip().lower() in ('all', 'full', 'any') and 'all' or 'action'
    contact_log_path = Path(args.contact_log) if args.contact_log else None
    contact_log: Dict[str, date] = {}
    contact_log_active = False
    # 优先：从飞书读取联系记录（若环境变量提供）
    feishu_app_token = os.getenv('FEISHU_CONTACT_APP_TOKEN') or os.getenv('FEISHU_APP_TOKEN')
    feishu_table_id = os.getenv('FEISHU_CONTACT_TABLE_ID') or os.getenv('FEISHU_TABLE_ID')
    feishu_token = os.getenv('FEISHU_USER_ACCESS_TOKEN') or os.getenv('FEISHU_TENANT_ACCESS_TOKEN')
    feishu_view_id = os.getenv('FEISHU_CONTACT_VIEW_ID')
    if feishu_app_token and feishu_table_id and feishu_token:
        try:
            contact_log = fetch_feishu_contact_log(
                feishu_app_token, feishu_table_id, today,
                token=feishu_token, view_id=feishu_view_id
            )
            if contact_log:
                contact_log_active = True
                try:
                    _total = len(contact_log or {})
                    _recent = 0
                    for _ph, _dt in (contact_log or {}).items():
                        try:
                            if (today - _dt).days >= 0 and (today - _dt).days < max(0, cooldown_days):
                                _recent += 1
                        except Exception:
                            pass
                    print(f"ℹ️  飞书联系记录加载完成：总计 {_total} 条；近{max(0, cooldown_days)}天 {_recent} 条")
                except Exception:
                    pass
            else:
                print("ℹ️  飞书联系记录为空，改用本地联系记录。")
                contact_log_active = False
        except Exception as e:
            print(f"⚠️  读取飞书联系记录失败：{e}，将尝试读取本地联系记录。")
    # 回退：读取本地 Excel 联系记录（扩展字段）
    contact_info_map: Dict[str, Dict[str, Any]] = {}
    if not contact_log_active:
        if contact_log_path and contact_log_path.exists():
            try:
                contact_log, contact_info_map = load_contact_log_extended(contact_log_path, today)
            except Exception:
                contact_log = load_contact_log(contact_log_path, today)
                contact_info_map = {}
            contact_log_active = True
            try:
                _total = len(contact_log or {})
                _recent = 0
                for _ph, _dt in (contact_log or {}).items():
                    try:
                        if (today - _dt).days >= 0 and (today - _dt).days < max(0, cooldown_days):
                            _recent += 1
                    except Exception:
                        pass
                print(f"ℹ️  本地联系记录加载完成：总计 {_total} 条；近{max(0, cooldown_days)}天 {_recent} 条")
            except Exception:
                pass
        elif args.contact_log and args.contact_log != "contact_log.xlsx":
            print(f"⚠️  联系记录表未找到：{contact_log_path}")

    config_path = Path(args.config) if args.config else None
    if not config_path:
        raise ValueError("Config path must be provided.")
    config_model = load_config(config_path)

    high_value_threshold = determine_threshold(customers, args.value_top)

    overview_rows, action_rows, snoozed_total, meta_map, cooldown_customers = build_alert_rows(
        customers=customers,
        today=today,
        high_value_threshold=high_value_threshold,
        churn_days=args.churn_days,
        churn_multiplier=args.churn_multiplier,
        drop_threshold=args.drop_threshold,
        config_model=config_model,
        anniversary_dates=anniversary_dates,
        anniversary_window=args.anniversary_window,
        anniversary_only=args.anniversary_only,
        contact_log=contact_log,
        contact_info=contact_info_map,
        cooldown_days=cooldown_days,
        cooldown_scope=cooldown_scope,
        exclude_recent_days=max(0, int(args.exclude_recent_days or 0)),
        allow_high_return=bool(args.allow_high_return),
    )

    if args.max_action and args.max_action > 0:
        action_rows = action_rows[: args.max_action]

    if not overview_rows:
        raise RuntimeError("No valid customer records were generated. Check source data and filters.")

    output_path = Path(args.output)
    write_workbook(
        output_path=output_path,
        overview_rows=overview_rows,
        action_rows=action_rows,
        today=today,
        high_value_threshold=high_value_threshold,
        config=args,
        anniversary_dates=anniversary_dates if anniversary_dates else None,
        contact_log_used=contact_log_active,
        cooldown_days=cooldown_days,
        snoozed_total=snoozed_total,
    )
    # Build global details map for all customers，用于 HTML 全库单号检索
    global_details: Dict[str, List[Dict[str, Any]]] = {}
    for key, stats in customers.items():
        details: List[Dict[str, Any]] = []
        for d in getattr(stats, "order_details", []) or []:
            try:
                # 字段映射：支持新旧字段名
                details.append(
                    {
                        "姓名": str(d.get("姓名", "")),
                        "手机号": str(d.get("手机号", "") or getattr(stats, "phone", "") or key),
                        "下单时间": str(d.get("顾客付款日期") or d.get("下单时间", "")),
                        "下单平台": str(d.get("出售平台") or d.get("下单平台", "")),
                        "厂家": str(d.get("厂家", "")),
                        "货品名": str(d.get("货品名", "")),
                        "商品名称": str(d.get("商品名称", "")),
                        "颜色": str(d.get("颜色", "")),
                        "尺码": str(d.get("尺码", "")),
                        "付款金额": float(d.get("收款额") or d.get("付款金额", 0.0)) if isinstance(d.get("收款额") or d.get("付款金额"), (int, float)) else common_to_float(d.get("收款额") or d.get("付款金额")),
                        "打款金额": float(d.get("打款金额", 0.0)) if isinstance(d.get("打款金额", 0.0), (int, float)) else common_to_float(d.get("打款金额")),
                        "负责人": str(d.get("负责人", "") or d.get("跟进人", "")),
                        "订单号": str(d.get("单号") or d.get("订单号", "")),
                        "退货单号": str(d.get("退货单号", "")),
                        "退款类型": str(d.get("退款类型") or d.get("状态", "")),
                        "退款原因": str(d.get("退款原因", "")),
                        "备注": str(d.get("备注", "")),
                        "数据来源": str(d.get("数据来源", "")),
                    }
                )
            except Exception:
                # best-effort; skip corrupt detail rows
                continue
        global_details[str(key)] = details

    # 构建货品名搜索索引
    product_search_index = build_product_search_index(global_details)

    # 计算冷却期手机号集合（用于 HTML 显示"已联系"名单）
    cooldown_keys: List[str] = []
    if contact_log_active and cooldown_days > 0:
        for ph, last_dt in (contact_log or {}).items():
            try:
                delta = (today - last_dt).days
            except Exception:
                continue
            if delta >= 0 and delta < cooldown_days:
                cooldown_keys.append(str(ph))
    
    # global_meta 已经在 build_alert_rows 中生成
    global_meta = meta_map

    if args.html_output:
        html_path = Path(args.html_output)
        write_html_dashboard(
            output_path=html_path,
            today=today,
            action_rows=action_rows,
            overview_rows=overview_rows,
            high_value_threshold=high_value_threshold,
            config=args,
            anniversary_dates=anniversary_dates if anniversary_dates else None,
            contact_log_used=contact_log_active,
            cooldown_days=cooldown_days,
            snoozed_total=snoozed_total,
            global_details=global_details,
            global_meta=global_meta,
            cooldown_keys=cooldown_keys,
            cooldown_customers=cooldown_customers,
            product_search_index=product_search_index,
        )
        print(f"HTML 可视化已生成: {html_path}")
    print(f"生成完成: {output_path} (高价值阈值= {high_value_threshold:.2f})")


if __name__ == "__main__":
    main()
