#!/usr/bin/env python3
from __future__ import annotations

import os
import sys
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent


def digits_only(s: str) -> str:
    return ''.join(ch for ch in str(s) if ch.isdigit())


def dedup_phone(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    d = digits_only(str(raw))
    return d or None


def auth_headers() -> Dict[str, str]:
    token = os.getenv('FEISHU_USER_ACCESS_TOKEN') or os.getenv('FEISHU_TENANT_ACCESS_TOKEN')
    return {
        'Authorization': f'Bearer {token}' if token else '',
        'Content-Type': 'application/json; charset=utf-8',
    }


def list_fields(app_token: str, table_id: str) -> List[str]:
    url = f'https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields'
    headers = auth_headers()
    params: Dict[str, Any] = {}
    names: List[str] = []
    page_token: Optional[str] = None
    while True:
        if page_token:
            params['page_token'] = page_token
        resp = requests.get(url, headers=headers, params=params, timeout=30)
        try:
            resp.raise_for_status()
            data = resp.json()
            if isinstance(data, dict) and data.get('code') not in (None, 0):
                break
        except Exception:
            break
        items = (data.get('data', {}) or {}).get('items', []) or []
        for it in items:
            n = it.get('field_name')
            if isinstance(n, str) and n:
                names.append(n)
        page_token = (data.get('data', {}) or {}).get('page_token')
        if not (data.get('data', {}) or {}).get('has_more'):
            break
    return names


def pick_name(candidates: List[str], available: List[str]) -> Optional[str]:
    aset = set(available)
    for c in candidates:
        if c in aset:
            return c
    return None


def fetch_records(app_token: str, table_id: str, view_id: Optional[str] = None) -> List[Dict[str, Any]]:
    url = f'https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records'
    headers = auth_headers()
    params: Dict[str, Any] = {'page_size': 200}
    if view_id:
        params['view_id'] = view_id
    records: List[Dict[str, Any]] = []
    page_token: Optional[str] = None
    while True:
        if page_token:
            params['page_token'] = page_token
        resp = requests.get(url, headers=headers, params=params, timeout=30)
        try:
            resp.raise_for_status()
            data = resp.json()
            if isinstance(data, dict) and data.get('code') not in (None, 0):
                break
        except Exception:
            break
        items = (data.get('data', {}) or {}).get('items', []) or []
        for it in items:
            rid = it.get('record_id') or it.get('id')
            fields = (it.get('fields') or {})
            if rid:
                records.append({'record_id': rid, 'fields': fields})
        page_token = (data.get('data', {}) or {}).get('page_token')
        if not (data.get('data', {}) or {}).get('has_more'):
            break
    return records


def update_record(app_token: str, table_id: str, record_id: str, fields: Dict[str, Any]) -> bool:
    url = f'https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/{record_id}'
    headers = auth_headers()
    body = {'fields': fields}
    resp = requests.put(url, headers=headers, json=body, timeout=30)
    try:
        resp.raise_for_status()
        data = resp.json()
        if isinstance(data, dict) and data.get('code') not in (None, 0):
            return False
        return True
    except Exception:
        return False


def load_phone_name_platform_map(ledger_path: Path) -> Tuple[Dict[str, str], Dict[str, str]]:
    if not ledger_path.exists():
        raise FileNotFoundError(f'Ledger not found: {ledger_path}')
    wb = load_workbook(ledger_path, data_only=True, read_only=True)
    try:
        ws = None
        # 优先选“汇总(全部)”工作表
        if '汇总(全部)' in wb.sheetnames:
            ws = wb['汇总(全部)']
        else:
            ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        rows = list(rows)
    finally:
        wb.close()
    if not rows:
        return {}, {}
    header = rows[0]
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    phone_idx = None
    name_idx = None
    platform_idx = None
    for k in ('手机号', '电话', '联系方式'):
        if k in idx:
            phone_idx = idx[k]
            break
    for k in ('姓名', '客户名称', '顾客姓名'):
        if k in idx:
            name_idx = idx[k]
            break
    for k in ('出售平台', '主要平台', '平台'):
        if k in idx:
            platform_idx = idx[k]
            break
    name_counter: Dict[str, Counter] = defaultdict(Counter)
    plat_counter: Dict[str, Counter] = defaultdict(Counter)
    for r in rows[1:]:
        try:
            ph = dedup_phone(r[phone_idx]) if phone_idx is not None and len(r) > phone_idx else None
        except Exception:
            ph = None
        if not ph:
            continue
        if name_idx is not None and len(r) > name_idx and r[name_idx]:
            name_counter[ph][str(r[name_idx]).strip()] += 1
        if platform_idx is not None and len(r) > platform_idx and r[platform_idx]:
            plat_counter[ph][str(r[platform_idx]).strip()] += 1
    phone_to_name = {ph: cnt.most_common(1)[0][0] for ph, cnt in name_counter.items() if cnt}
    phone_to_plat = {ph: cnt.most_common(1)[0][0] for ph, cnt in plat_counter.items() if cnt}
    return phone_to_name, phone_to_plat


def main() -> int:
    app_token = os.getenv('FEISHU_CONTACT_APP_TOKEN') or os.getenv('FEISHU_APP_TOKEN')
    table_id = os.getenv('FEISHU_CONTACT_TABLE_ID') or os.getenv('FEISHU_TABLE_ID')
    view_id = os.getenv('FEISHU_CONTACT_VIEW_ID')
    if not (app_token and table_id):
        print('Missing FEISHU_CONTACT_APP_TOKEN / FEISHU_CONTACT_TABLE_ID')
        return 2
    # 构建映射：手机号 -> 姓名/主要平台
    ledger = ROOT / 'tech' / '账单汇总_全部.xlsx'
    try:
        phone_to_name, phone_to_plat = load_phone_name_platform_map(ledger)
        if not phone_to_name and not phone_to_plat:
            print('No mapping extracted from ledger; aborting to avoid empty updates.')
            return 3
    except Exception as e:
        print('Load ledger failed:', e)
        return 3
    # 取目标表字段
    fields_available = list_fields(app_token, table_id)
    phone_col = pick_name(['手机号', '手机', '手机号码', '联系电话', '电话', '联系方式'], fields_available) or '手机号'
    name_col = pick_name(['姓名', '客户名称', '顾客姓名', 'Name', 'name'], fields_available)
    platform_col = pick_name(['联系平台', '主要平台', '平台', 'Platform', 'platform'], fields_available)
    if not phone_col:
        print('Cannot locate phone column in contact table; found:', fields_available)
        return 4
    # 拉取记录
    records = fetch_records(app_token, table_id, view_id=None)  # 不限定视图，避免筛选丢失
    if not records:
        print('No records fetched; check permissions or table id.')
        return 5
    updated = 0
    skipped = 0
    for rec in records:
        rid = rec['record_id']
        f = rec.get('fields') or {}
        ph = dedup_phone(f.get(phone_col))
        if not ph:
            skipped += 1
            continue
        patch: Dict[str, Any] = {}
        # 填姓名
        if name_col:
            cur_name = (f.get(name_col) or '').strip() if isinstance(f.get(name_col), str) else f.get(name_col)
            if not cur_name:
                new_name = phone_to_name.get(ph)
                if new_name:
                    patch[name_col] = new_name
        # 填联系平台
        if platform_col:
            cur_plat = (f.get(platform_col) or '').strip() if isinstance(f.get(platform_col), str) else f.get(platform_col)
            if not cur_plat:
                new_plat = phone_to_plat.get(ph)
                if new_plat:
                    patch[platform_col] = new_plat
        if patch:
            ok = update_record(app_token, table_id, rid, patch)
            if ok:
                updated += 1
            else:
                skipped += 1
        else:
            skipped += 1
    print(f'Done. Updated {updated} records, skipped {skipped}.')
    return 0


if __name__ == '__main__':
    sys.exit(main())

