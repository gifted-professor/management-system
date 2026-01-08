from __future__ import annotations
from datetime import date, datetime, timedelta
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path
from openpyxl import load_workbook

LEDGER_COLUMNS: List[str] = [
    '姓名', '顾客付款日期', '负责人', '出售平台', '商品名称', '货品名', '收款额', '成本价',
    '打款金额', '打款日期', '是否打款', '厂家', '报单日期', '出单号日期', '单号', '退货地址',
    '状态', '手机号', '地址', '客户信息', '快递公司', '备注', '附件', '退款负责人', '退款日',
    '退款金额', '退款类型', '退款原因', '退货单号', '发给厂家日期', '厂家确认日期', '厂家退回金额',
    '退货状态', '退芋圆入库日期', '是否出库', '出库单号', '库存情况', '尺码', '颜色', '当天日期',
    '父记录 3', '打款信息公式', '打款信息', '退货物流', '有货厂家', '货品备注', '报单异常确认',
    '公式出库', '优化利润', '优化利润率', '公式勿删', '优化成本', '单品销售量',
    '退货未找厂家处理', '未付货款', '退货未找厂家退款', '有效', 'temp【看到了就删掉这栏】', '父记录',
    '对账勿删', '父记录 (1)', '对账勿删 (1)', '父记录 4', '净收款', '毛利', '利润估算',
    '数据来源',
]

def resolve_sheet(path: Path, sheet_name: Optional[str]):
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in {path}.")
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]
    return wb, ws

def normalize(value: Optional[str]) -> str:
    if value is None:
        return ''
    return str(value).strip()

def digits_only(s: str) -> str:
    return ''.join(ch for ch in s if ch.isdigit())

def deduplicate_phone(raw: Optional[str]) -> Optional[str]:
    if raw is None:
        return None
    digits = ''.join(ch for ch in str(raw) if ch.isdigit())
    return digits or None

def to_float(value) -> float:
    """Parse a value into float robustly.

    Key behavior change: when a string contains multiple numeric fragments
    (e.g. "105~110", "105-110", "105/110"), we no longer concatenate digits
    which used to yield 105110.0. Instead, we extract numeric tokens and take
    the first one by default.

    This makes columns like 打款金额/成本价 safer when the sheet contains
    ranges or remarks such as "105-110(按款式)".
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return 0.0
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return 0.0
        # Normalize common full-width symbols and currency signs
        normalized = (
            stripped.replace('￥', '').replace('¥', '').replace(',', '').replace('，', ',')
        )
        # Extract numeric tokens (supports leading +/- and decimals)
        nums = re.findall(r"[+-]?\d+(?:\.\d+)?", normalized)
        if not nums:
            return 0.0
        try:
            return float(nums[0])
        except Exception:
            return 0.0
    return 0.0

def parse_excel_date(raw, today: date) -> Optional[date]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, (int, float)):
        origin = datetime(1899, 12, 30)
        try:
            return (origin + timedelta(days=float(raw))).date()
        except OverflowError:
            return None
    if isinstance(raw, str):
        text = str(raw).strip()
        if not text:
            return None
        m = re.search(r"(\d{4}|\d{2})[./-](\d{1,2})[./-](\d{1,2})", text)
        if m:
            y, mo, d = m.groups()
            yy = int(y)
            if len(y) == 2:
                yy = 2000 + yy if yy <= 68 else 1900 + yy
            try:
                return date(yy, int(mo), int(d))
            except ValueError:
                pass
        m = re.search(r"(\d{4}|\d{2})年(\d{1,2})月(\d{1,2})日", text)
        if m:
            y, mo, d = m.groups()
            yy = int(y)
            if len(y) == 2:
                yy = 2000 + yy if yy <= 68 else 1900 + yy
            try:
                return date(yy, int(mo), int(d))
            except ValueError:
                pass
        m = re.search(r"(\d{1,2})[./-](\d{1,2})(?!\d)", text)
        if m:
            mo, d = m.groups()
            try:
                assumed = date(today.year, int(mo), int(d))
            except ValueError:
                assumed = None
            if assumed:
                if assumed > today:
                    try:
                        assumed = date(today.year - 1, int(mo), int(d))
                    except ValueError:
                        assumed = None
                if assumed:
                    return assumed
        cleaned = ''.join(ch for ch in text if ch.isdigit() or ch in "-./年月日")
        formats = (
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y.%m.%d",
            "%Y%m%d",
            "%Y年%m月%d日",
            "%y-%m-%d",
            "%y/%m/%d",
            "%y.%m.%d",
            "%y年%m月%d日",
            "%m-%d",
            "%m/%d",
            "%m.%d",
        )
        for fmt in formats:
            try:
                parsed = datetime.strptime(cleaned, fmt)
                if fmt in ("%m-%d", "%m/%d", "%m.%d"):
                    assumed = parsed.replace(year=today.year)
                    if assumed.date() > today:
                        assumed = assumed.replace(year=today.year - 1)
                    return assumed.date()
                return parsed.date()
            except ValueError:
                continue
        digits = ''.join(ch for ch in cleaned if ch.isdigit())
        if len(digits) >= 6:
            m6 = re.search(r"(?<!\d)(\d{6})(?!\d)", digits)
            if m6:
                try:
                    return datetime.strptime(m6.group(1), "%y%m%d").date()
                except ValueError:
                    pass
    return None

def build_header_index(header_row: Iterable[Optional[str]]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        name = str(cell).strip() if cell is not None else ''
        if not name:
            continue
        mapping[name] = idx
    return mapping

def lookup_index(header_index: Dict[str, int], candidates: Tuple[str, ...]) -> Optional[int]:
    for candidate in candidates:
        if candidate in header_index:
            return header_index[candidate]
    return None
